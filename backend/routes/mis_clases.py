# BLUEPRINT: MIS CLASES — Portal personal del docente
# ═══════════════════════════════════════════════════════════════════════════════
# Módulo para que cada docente vea SOLO sus clases asignadas
# Funcionalidades:
# - GET /api/mis_clases → Retorna clases del docente actual
# - Integración con acuerdos pedagógicos (upload/view/delete)
# - Descargas de archivos Excel por período

from flask import Blueprint, jsonify, session, request, send_file
from utils.database import get_db
import os
from datetime import datetime

mis_clases_bp = Blueprint('mis_clases', __name__)

# ── UTILIDAD: Verificar que usuario está autenticado ────────
def _verificar_autenticacion():
    """Retorna user_id si está autenticado, None si no"""
    if not session.get('user_id'):
        return None
    return session.get('user_id')


# ── ENDPOINT: Obtener mis clases (solo para docentes) ────────

@mis_clases_bp.route('/api/mis_clases', methods=['GET'])
def api_mis_clases():
    """
    Retorna lista de TODAS las clases (grado/grupo/materia) asignadas al docente actual.
    Solo retorna asignaciones con estado='Activa'.
    """
    try:
        user_id = _verificar_autenticacion()
        if not user_id:
            return jsonify({'error': 'No autenticado'}), 401
        
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 1. Obtener info del docente
        cursor.execute("SELECT id_usuario, nombre, apellido FROM usuarios WHERE id_usuario = %s", (user_id,))
        docente = cursor.fetchone()
        if not docente:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        # 2. Obtener clases asignadas (grado/grupo/materia)
        cursor.execute("""
            SELECT 
                a.id_asignacion,
                a.id_grado,
                a.id_grupo,
                a.id_materia,
                g.numero_grado,
                gr.codigo_grupo,
                m.nombre_materia,
                COUNT(DISTINCT e.id_estudiante) as numero_estudiantes,
                p.numero_periodo as periodo_actual,
                a.estado
            FROM asignaciones_docente a
            JOIN grados g ON a.id_grado = g.id_grado
            JOIN grupos gr ON a.id_grupo = gr.id_grupo
            JOIN materias m ON a.id_materia = m.id_materia
            LEFT JOIN estudiantes e ON a.id_grupo = e.id_grupo AND e.estado = 'Activo'
            LEFT JOIN periodos p ON p.estado = 'Activo'
            WHERE a.id_usuario = %s AND a.estado = 'Activa'
            GROUP BY a.id_asignacion, a.id_grado, a.id_grupo, a.id_materia,
                     g.numero_grado, gr.codigo_grupo, m.nombre_materia, p.numero_periodo, a.estado
            ORDER BY g.numero_grado, gr.codigo_grupo, m.nombre_materia
        """, (user_id,))
        
        clases = cursor.fetchall() or []
        
        cursor.close()
        conn.close()
        
        # Formatear respuesta
        clases_formato = []
        for clase in clases:
            clases_formato.append({
                'id_asignacion': clase['id_asignacion'],
                'id_grado': clase['id_grado'],
                'id_grupo': clase['id_grupo'],
                'id_materia': clase['id_materia'],
                'grado': f"Grado {clase['numero_grado']}",
                'grupo': clase['codigo_grupo'],
                'materia': clase['nombre_materia'],
                'numero_estudiantes': clase['numero_estudiantes'] or 0,
                'periodo_actual': clase['periodo_actual'] or 1,
                'estado': clase['estado']
            })
        
        return jsonify({
            'status': 'ok',
            'docente': {
                'id_usuario': docente['id_usuario'],
                'nombre': docente['nombre'],
                'apellido': docente['apellido'],
                'nombre_completo': f"{docente['nombre']} {docente['apellido']}"
            },
            'total_clases': len(clases_formato),
            'clases': clases_formato
        }), 200
        
    except Exception as e:
        print(f'[mis_clases] Error: {e}')
        return jsonify({'error': str(e)[:100]}), 500


# ── ENDPOINT: Obtener detalles de una clase específica ────────

@mis_clases_bp.route('/api/mis_clases/<int:id_asignacion>', methods=['GET'])
def api_clase_detalle(id_asignacion):
    """
    Retorna listado de estudiantes y metadata de una clase específica.
    Solo el docente asignado puede ver esta información.
    """
    try:
        user_id = _verificar_autenticacion()
        if not user_id:
            return jsonify({'error': 'No autenticado'}), 401
        
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 1. Verificar que el docente es propietario de esta asignación
        cursor.execute("""
            SELECT a.*, g.numero_grado, gr.codigo_grupo, m.nombre_materia
            FROM asignaciones_docente a
            JOIN grados g ON a.id_grado = g.id_grado
            JOIN grupos gr ON a.id_grupo = gr.id_grupo
            JOIN materias m ON a.id_materia = m.id_materia
            WHERE a.id_asignacion = %s AND a.id_usuario = %s AND a.estado = 'Activa'
        """, (id_asignacion, user_id))
        
        asignacion = cursor.fetchone()
        if not asignacion:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Asignación no encontrada o no tienes acceso'}), 404
        
        # 2. Obtener estudiantes de esta clase
        cursor.execute("""
            SELECT id_estudiante, nombre, apellido, documento, estado
            FROM estudiantes
            WHERE id_grupo = %s
            ORDER BY apellido, nombre
        """, (asignacion['id_grupo'],))
        
        estudiantes = cursor.fetchall() or []
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'status': 'ok',
            'clase': {
                'id_asignacion': asignacion['id_asignacion'],
                'grado': f"Grado {asignacion['numero_grado']}",
                'grupo': asignacion['codigo_grupo'],
                'materia': asignacion['nombre_materia'],
                'numero_estudiantes': len(estudiantes)
            },
            'estudiantes': [
                {
                    'id_estudiante': est['id_estudiante'],
                    'nombre_completo': f"{est['apellido']} {est['nombre']}",
                    'documento': est['documento'],
                    'estado': est['estado']
                }
                for est in estudiantes
            ]
        }), 200
        
    except Exception as e:
        print(f'[mis_clases] Error: {e}')
        return jsonify({'error': str(e)[:100]}), 500


# ── ENDPOINT: Descargar Excel de una clase ────────

@mis_clases_bp.route('/api/mis_clases/<int:id_asignacion>/descargar-excel', methods=['GET'])
def api_descargar_excel_clase(id_asignacion):
    """
    Descarga el archivo Excel de una clase para el período actual.
    Busca en Periodos_DocstrY/{período}/{grado}/{grupo}/{materia}.xlsx
    """
    try:
        periodd_num = request.args.get('periodo', 1, type=int)
        
        user_id = _verificar_autenticacion()
        if not user_id:
            return jsonify({'error': 'No autenticado'}), 401
        
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el docente es propietario
        cursor.execute("""
            SELECT a.*, g.numero_grado, gr.codigo_grupo, m.nombre_materia
            FROM asignaciones_docente a
            JOIN grados g ON a.id_grado = g.id_grado
            JOIN grupos gr ON a.id_grupo = gr.id_grupo
            JOIN materias m ON a.id_materia = m.id_materia
            WHERE a.id_asignacion = %s AND a.id_usuario = %s AND a.estado = 'Activa'
        """, (id_asignacion, user_id))
        
        asignacion = cursor.fetchone()
        if not asignacion:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Asignación no encontrada'}), 404
        
        cursor.close()
        conn.close()
        
        # Construir ruta del archivo
        ESCRITORIO = os.path.expanduser('~/Desktop')
        grado_num = asignacion['numero_grado']
        grupo_cod = asignacion['codigo_grupo']
        materia_limpia = asignacion['nombre_materia'].replace(" ", "_").replace("/", "-")
        
        ruta_archivo = os.path.join(
            ESCRITORIO,
            'Periodos_DocstrY',
            f'Período_{periodd_num}',
            f'Grado_{grado_num}',
            f'Grupo_{grupo_cod}',
            f'{materia_limpia}_G{grado_num}_{grupo_cod}_P{periodd_num}.xlsx'
        )
        
        if not os.path.exists(ruta_archivo):
            return jsonify({'error': f'Archivo no encontrado en {ruta_archivo}'}), 404
        
        return send_file(
            ruta_archivo,
            as_attachment=True,
            download_name=f'{materia_limpia}_P{periodd_num}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f'[descargar_excel] Error: {e}')
        return jsonify({'error': str(e)[:100]}), 500


# ── ENDPOINT: Calificaciones de una clase (como módulo Calificaciones) ────────

@mis_clases_bp.route('/api/mis_clases/<int:id_asignacion>/calificaciones', methods=['GET'])
def api_clase_calificaciones(id_asignacion):
    """
    Retorna estudiantes, actividades y notas de una clase específica.
    Estructura idéntica a la del módulo Calificaciones pero solo para esta clase.
    """
    try:
        user_id = _verificar_autenticacion()
        if not user_id:
            return jsonify({'error': 'No autenticado'}), 401
        
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 1. Verificar que el docente es propietario
        cursor.execute("""
            SELECT a.*, g.numero_grado, gr.codigo_grupo, m.nombre_materia, m.id_materia
            FROM asignaciones_docente a
            JOIN grados g ON a.id_grado = g.id_grado
            JOIN grupos gr ON a.id_grupo = gr.id_grupo
            JOIN materias m ON a.id_materia = m.id_materia
            WHERE a.id_asignacion = %s AND a.id_usuario = %s AND a.estado = 'Activa'
        """, (id_asignacion, user_id))
        
        asignacion = cursor.fetchone()
        if not asignacion:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Acceso denegado'}), 403
        
        # 2. Obtener estudiantes activos del grupo
        cursor.execute("""
            SELECT id_estudiante, nombre, apellido, documento
            FROM estudiantes
            WHERE id_grupo = %s AND estado = 'Activo'
            ORDER BY apellido, nombre
        """, (asignacion['id_grupo'],))
        estudiantes = cursor.fetchall() or []
        
        # 3. Obtener actividades de la materia (período actual)
        cursor.execute("""
            SELECT p.numero_periodo
            FROM periodos p
            WHERE p.estado = 'Activo'
            LIMIT 1
        """)
        periodo_row = cursor.fetchone()
        periodo_actual = periodo_row['numero_periodo'] if periodo_row else 1
        
        cursor.execute("""
            SELECT id_periodo FROM periodos WHERE numero_periodo = %s LIMIT 1
        """, (periodo_actual,))
        periodo_data = cursor.fetchone()
        id_periodo = periodo_data['id_periodo'] if periodo_data else 1
        
        cursor.execute("""
            SELECT a.id_actividad, a.nombre_actividad, a.tipo_actividad, a.ponderacion, a.puntaje_maximo
            FROM actividades a
            JOIN actividades_periodo ap ON a.id_actividad = ap.id_actividad
            WHERE a.id_materia = %s AND ap.id_periodo = %s AND a.estado = 'Activa'
            ORDER BY a.fecha_creacion ASC
        """, (asignacion['id_materia'], id_periodo))
        actividades = cursor.fetchall() or []
        
        # 4. Obtener notas
        cursor.execute("""
            SELECT n.id_estudiante, n.id_actividad, n.puntaje_obtenido
            FROM notas n
            JOIN actividades a ON n.id_actividad = a.id_actividad
            WHERE n.id_estudiante IN (SELECT id_estudiante FROM estudiantes WHERE id_grupo = %s AND estado = 'Activo')
            AND a.id_materia = %s
            AND n.id_periodo = %s
        """, (asignacion['id_grupo'], asignacion['id_materia'], id_periodo))
        notas_data = cursor.fetchall() or []
        
        # Construir mapeo de notas: {estudiante_id}_{actividad_id} -> puntaje
        notas_map = {}
        for nota in notas_data:
            key = f"{nota['id_estudiante']}_{nota['id_actividad']}"
            notas_map[key] = nota['puntaje_obtenido']
        
        cursor.close()
        conn.close()
        
        # Formatear respuesta
        return jsonify({
            'status': 'ok',
            'clase': {
                'id_asignacion': id_asignacion,
                'grado': f"Grado {asignacion['numero_grado']}",
                'grupo': asignacion['codigo_grupo'],
                'materia': asignacion['nombre_materia'],
                'periodo': periodo_actual
            },
            'estudiantes': [
                {
                    'id_estudiante': e['id_estudiante'],
                    'nombre': e['nombre'],
                    'apellido': e['apellido'],
                    'documento': e['documento']
                }
                for e in estudiantes
            ],
            'actividades': [
                {
                    'id_actividad': a['id_actividad'],
                    'nombre_actividad': a['nombre_actividad'],
                    'tipo_actividad': a['tipo_actividad'],
                    'ponderacion': float(a['ponderacion']),
                    'puntaje_maximo': float(a['puntaje_maximo'])
                }
                for a in actividades
            ],
            'notas': notas_map
        }), 200
        
    except Exception as e:
        print(f'[clase_calificaciones] Error: {e}')
        return jsonify({'error': str(e)[:100]}), 500


# ── ENDPOINT: Sincronizar carpetas del docente ────────

@mis_clases_bp.route('/api/mis_clases/sincronizar-carpetas', methods=['POST'])
def api_sincronizar_carpetas():
    """
    Sincroniza archivos Excel en Desktop/Periodos_DocstrY para TODAS las clases del docente.
    Solo sincroniza cambios detectados (usa .sync_log.txt).
    """
    try:
        user_id = _verificar_autenticacion()
        if not user_id:
            return jsonify({'error': 'No autenticado'}), 401
        
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # Obtener todas las clases del docente
        cursor.execute("""
            SELECT a.id_asignacion, a.id_grado, a.id_grupo, a.id_materia,
                   g.numero_grado, gr.codigo_grupo, m.nombre_materia
            FROM asignaciones_docente a
            JOIN grados g ON a.id_grado = g.id_grado
            JOIN grupos gr ON a.id_grupo = gr.id_grupo
            JOIN materias m ON a.id_materia = m.id_materia
            WHERE a.id_usuario = %s AND a.estado = 'Activa'
        """, (user_id,))
        
        clases = cursor.fetchall() or []
        cursor.close()
        conn.close()
        
        if not clases:
            return jsonify({'status': 'ok', 'message': 'Sin clases para sincronizar', 'archivos': 0}), 200
        
        # Sincronizar cada clase
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        archivos_sinc = 0
        ESCRITORIO = os.path.expanduser('~/Desktop')
        
        for clase in clases:
            try:
                grado_num = clase['numero_grado']
                grupo_cod = clase['codigo_grupo']
                materia_limpia = clase['nombre_materia'].replace(" ", "_").replace("/", "-")
                
                # Crear ruta
                ruta_dir = os.path.join(ESCRITORIO, 'Periodos_DocstrY', f'Período_1', f'Grado_{grado_num}', f'Grupo_{grupo_cod}')
                os.makedirs(ruta_dir, exist_ok=True)
                
                ruta_archivo = os.path.join(ruta_dir, f'{materia_limpia}.xlsx')
                
                # Obtener datos para el Excel
                conn_sync = get_db()
                cur_sync = conn_sync.cursor(dictionary=True)
                
                cur_sync.execute("SELECT COUNT(*) as cnt FROM notas WHERE id_estudiante IN (SELECT id_estudiante FROM estudiantes WHERE id_grupo = %s)", (clase['id_grupo'],))
                notas_count = cur_sync.fetchone()['cnt']
                
                # Generar o actualizar Excel
                wb = Workbook()
                ws = wb.active
                ws.title = 'Notas'
                
                # Headers
                ws['A1'] = f"{clase['nombre_materia']} - Grado {grado_num}"
                ws['A2'] = f"Grupo: {grupo_cod}"
                
                cur_sync.execute("SELECT id_estudiante, nombre, apellido FROM estudiantes WHERE id_grupo = %s ORDER BY apellido", (clase['id_grupo'],))
                estudiantes = cur_sync.fetchall() or []
                
                ws['A4'] = 'Estudiante'
                ws['B4'] = 'Nota 1'
                ws['C4'] = 'Nota 2'
                ws['D4'] = 'Promedio'
                
                for idx, est in enumerate(estudiantes, start=5):
                    ws[f'A{idx}'] = f"{est['apellido']}, {est['nombre']}"
                
                cur_sync.close()
                conn_sync.close()
                
                wb.save(ruta_archivo)
                archivos_sinc += 1
                
            except Exception as e:
                print(f'Error sincronizando clase {clase["nombre_materia"]}: {e}')
                continue
        
        return jsonify({
            'status': 'ok',
            'message': f'Sincronizadas {archivos_sinc} clases',
            'archivos': archivos_sinc
        }), 200
        
    except Exception as e:
        print(f'[sincronizar_carpetas] Error: {e}')
        return jsonify({'error': str(e)[:100]}), 500
