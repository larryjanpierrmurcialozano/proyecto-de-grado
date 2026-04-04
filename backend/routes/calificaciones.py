# Se requiere que se crean 4 carpetas para cuatro periodos y dentro de cada carpeta crear los documentos correspondientes que ya estén diseñados (Los excels importados ya definidos).
# 
# indicaciones by larry.ai.chan.uwu.rmppnochas3000 :D 
# el id de periodo es el que permite modificar el archivo excel, si el id de periodo esta desactivado desde el futuro modulo de periodos, el cual la idea es que cierre el permiso de modificaciones, el cual en algun futuro se añadira
# en calificaciones se requiere que los excel se rellenen con los nombre de los estudiantes y los partes de las notas las cuales deben ser modificables dentro de la app
# Calificaciones sistema de gestion de calificaciones, CRUD de calificaciones, listado de calificaciones por curso, etc. 
# ═══════════════════════════════════════════════════════════════════════════════
# BLUEPRINT: CALIFICACIONES — CRUD de calificaciones, listado de calificaciones por curso
# todo el modulo de calificaciones funciona como un sistema local, el cual usara los archivos, tanto a nivel local como a nivel de servidor, 
# para almacenar las calificaciones, esto con el fin de evitar problemas de concurrencia y bloqueo de base de datos, 
# ademas de que se podra usar en modo offline, y luego sincronizar con el servidor cuando se tenga conexion a internet.
# todo se tratara de guardar en formato excel, usando libreria openpyxl, y se guardara en una carpeta especifica para cada curso, 
# dentro de una carpeta general de calificaciones, esto con el fin de mantener un orden y facilitar la busqueda de las calificaciones. 
#   todo se tendra una estructura de carpetas asi:
#   calificaciones/
#       grado #/ 
#                grupo #/ 
#                          materias #/
#           materia_#_#_(en este espacio de numeral para materias se decidira poner 2 valores tipo el grado y grupo perteneciente, siendo tal que se vea "materia_g6_A_calificaciones.xlsx")calificaciones.xlsx
#           calificaciones.xlsx / todo para cada curso por grupos
#  el chiste de todo es usar archivos excel como referencia, que sean editable dentro de la aplicacion 
# y que se puedan descargar o subir al sistema para asi mantener cierta flexibilidad y facilidad de uso.
# este sistema tambien podra incluir funcionalidades como el calculo de promedios, la generacion de reportes, la exportacion a pdf, etc.
# todo se tendra una ruta para cada funcionalidad, por ejemplo:
#   /calificaciones/grado_#_calificaciones/grupo_#_calificaciones.xlsx
#       GET: devuelve el archivo excel con las calificaciones del grado # y del grupo #
#       POST: recibe un archivo excel con las calificaciones del grado # y del grupo #, y lo guarda en el archivo correspondiente
#       PUT: actualiza las calificaciones del grado # y del grupo #, basadas en el archivo excel recibido
#       DELETE: elimina el archivo excel con las calificaciones del grado #
# grado (#) es literal como señalo el grado tipo 1 2 3 4 5 6 7 8 9 10 11, y se usara como parte del nombre del archivo y de la carpeta.
# sin embargo la idea es que se autogeneren estas funcionalidad de manera dinamica, es decir, 
# que se tenga una ruta general para cada curso, y que se maneje el archivo excel de manera dinamica, 
# sin necesidad de tener una ruta especifica para cada curso, esto con el fin de evitar la redundancia y facilitar el mantenimiento del codigo.
# tambien hay que tener claro que lo que se busca es que al generar el excel de cada grado se autogenere con el valor seleccionado 
# salido de la lista, tipo en caso de seleccionar el grado 7, que se genere la carpeta con el nombre,
# /grado_7_calificaciones(calificaciones/grado_7_calificaciones/grupo_A_calificaciones.xlsx(la idea en la que pongo grupo es para,
# tener la identificacion de que esa planilla es para ese grupo en especifico, con la idea de que al crear la carpeta para los archivos .xlsx ,
# seria que el propio coso genere segun los grupos existentes, osea seria tipo si hay un grupo, que se autogenere el xlsx para el grupo, 
# si se elimina el grupo, se prioriza que quede un precedente del archivo, para esto es la idea de que al crear estas carpetas, 
# dentro del propio dispositivo se tenga un archivo de respaldo donde se guarden todos los excel que esten hechos con sus modificaciones 
# hasta el ultimo momento en el que existieron o hubo un ultimo cambio)), y que al descargarlo se descargue con ese nombre, y al subirlo se suba 
# con ese nombre
# que se guarde en la carpeta correspondiente, y que al descargarlo se descargue con ese nombre, y al subirlo se suba con ese nombre, 
# claro, se creara es una carpeta del grado con el nombre del grado, y dentro de esa carpeta se creara un archivo excel con el nombre del grupo, 
# esto con el fin de mantener un orden y facilitar la busqueda de las calificaciones.
# se puede mantener un control de archivos dentro del modulo grafico de calificaciones, no se muy bien como se deberia ver, pero tengo una idea,
# podria recrear una especia de explorador de archivos dentro del modulo de calificaciones, donde se pueda ver la estructura de carpetas y archivos,
# y se pueda navegar por ella y se pueda selecionar todos las carpetas y rutas, sin importar si es profesor o admin, cada uno tendra un sistema de
# gestion de carpetas propio, claro la idea es que el profesor solo vea las carpetas y archivos que le corresponden a sus grados y grupos de
# materias asignadas, y el admin/director/cordinador, tendra una vista de todas las carpetas y archivos dentro del sistema, todo organizado
# en escala de grados y grupos como esta ya diseñado, la idea es que cada grupo con una materia asiganada tenga un archivo de calificaciones,
# claro, seria una estructura mucho mas compleja, puesto que cada grado deberia tener su planilla de cada materia que tiene asignada, 
# entonces seria algo asi: 
# /calificaciones/grado_#(# = numero de grado sea sexto septimo y asi)/grupo_# (siendo el # el grupo que sea tipo a, b, o c)/materias_#/materia_#_#_calificaciones.xlsx
# regulacion, lee la base de datos para entender estructura completa, te especifico las tablas que usaras para construir estas indicaciones

# ═══════════════════════════════════════════════════════════════════════════════
# BLUEPRINT: CALIFICACIONES — Sistema Híbrido (Archivos Locales + Sincronización DB)
# ═══════════════════════════════════════════════════════════════════════════════
#
# CONCEPTO ARQUITECTÓNICO DEL MÓDULO: "Tolerancia a fallos y Trabajo Offline"
# --------------------------------------------------------------------------------
# Este módulo se basa en un enfoque híbrido, combinando la flexibilidad de uso de 
# archivos Excel (carpetas locales) con la rapidez y seguridad de una Base de Datos (MySQL).
#
# 1. ALMACENAMIENTO FÍSICO JERÁRQUICO (El corazón local/offline):
#    - El sistema será el encargado ÚNICO de CREAR los archivos Excel.
#    - Mantiene una estructura de carpetas física y persistente: EJEMPLO:
#      /planillas_locales/Grado_7/Grupo_A/materia_Matematicas_G7_A.xlsx
#    - Esto asegura que la App/Web pueda funcionar incluso si no hay conexión 
#      momentánea a la base de datos principal, manteniendo los Excel en su directorio.
#
# 2. SISTEMA DE RESPALDO Y VERSIONADO ("El Archivador Histórico"):
#    - NUNCA se sobreescribe destructivamente un Excel modificado. 
#    - Al subir un Excel con nuevas calificaciones, la versión anterior se mueve a 
#      una subcarpeta /historial_respaldos/ con marca de tiempo (ej. ..._v20260313.xlsx).
#    - Estos archivos se guardan permanentemente y SOLO se borran por mantenimiento 
#      directo de sistema o DB (nunca desde la interfaz normal del usuario).
#
# 3. SINCRONIZACIÓN AUTOMÁTICA HACIA LA BASE DE DATOS:
#    - Cuando hay internet/conexión a la base de datos, el sistema lee el archivo 
#      Excel ingresado/modificado y "sincroniza" esos datos, rellenando las tablas:
#      `notas`, `estudiantes`, `actividades`, etc.
#    - La base de datos mantiene lectura constante de estos envíos. Generar los Excel
#      siempre se hace en base a la info maestra.
#
# 4. EXCEL CONTROLADOS:
#    - Los archivos generados por el sistema bloquearán las zonas de nombres/id, 
#      permitiendo al docente llenar solo las notas.
#    - Los IDs de estudiantes / notas, en lugar de estar ocultos en el Excel, serán
#      manejables o visibles de forma opcional DENTRO de la interfaz web/app del 
#      sistema, no estorbando en el propio archivo físico .xlsx esto con el fin de 
#      que al desear imprimir o descargar el archivo, se tenga un formato limpio y 
#      profesional, pero a su vez, al subirlo, el sistema pueda reconocer los datos y 
#      sincronizarlos con la base de datos sin problemas.
#
# RUTAS CORE A IMPLEMENTAR:
# - GET/POST `/api/calificaciones/estructura_carpetas`: Explora y crea la jerarquía.
# - GET `/api/calificaciones/generar_planilla`: Crea el Excel bloqueado físicamente.
# - POST `/api/calificaciones/subir_planilla`: Recibe Excel, versiona el anterior, guarda y sincroniza DB.
#
#
# LIBRERÍAS CLAVE: os (rutas), shutil (movimiento de historiales), openpyxl (creación/lectura de Excel).ahora quiero que ya funcionando la creacion de excel, sirva la sincronizacion del escritorio con el "server", osea que si se crean archivos o excel en el escritorio principal, se carguen en el server ahi te mostre donde tengo los archivos dentro del "server" para que se sincronice los archivos del escritorio y se suban a donde corresponden automaticamente, evitando lo de tener que crear el get de cada excel, tipo la idea es que los que crean la planilla son la coordinadora y roles superiores, el rol de profesores no lo vera pero no nos preocupemos por eso de momento, por ahora concentremonos en hacer funcionar lo que hay
# 1. El Concepto de "Plantilla Maestra"

# En lugar de generar el diseño desde cero con código (que es muy difícil), haz esto:
#    Crea un archivo Excel que tenga todo el diseño bonito: logos, bordes, colores y los textos fijos.
#    En las celdas donde van los datos variables (Nombre del Docente, Grado, Institución), escribe etiquetas especiales como {{DOCENTE}}, {{GRADO}}, {{INSTITUCION}}.
#    Guarda este archivo en una carpeta de tu backend llamada templates/.
#
# 2. Automatización de Relleno (Lógica Simple)
# Cuando el docente quiera descargar la planilla o subirla al Drive, tu código hará lo siguiente:
#    Carga la Plantilla: Abre el archivo .xlsx maestro usando la librería openpyxl.
#    Busca y Reemplaza: El código busca la celda que dice {{DOCENTE}} y la cambia por el nombre del usuario logueado en la plataforma.
#    Inyecta los Estudiantes: Recorre tu lista de estudiantes de la base de datos y los escribe a partir de la fila donde comienza la tabla.
#    Conexión con Asistencias: Como mencionas que tienes el módulo de asistencias, el código hace un COUNT de las fallas del estudiante en la DB y pone ese número en la columna "FALLAS" de la planilla automáticamente.
#
# 3. ¿Quién edita la plantilla? (Tu duda entre Admin o Docente)
# Para mantenerlo simple como pides, te sugiero el Modelo Híbrido:
#    La Institución/Admin: Sube el "Formato Base" (el diseño general con los logos de la escuela). Esto se hace una sola vez.
#    La Plataforma: Se encarga de llenar automáticamente: Institución, Sede, Jornada, Grado, Periodo, Docente y Estudiantes.
#    El Docente: Solo se encarga de lo que realmente le toca: las notas de las actividades.
#
# 4. Flujo de Trabajo Sugerido para tu Web
# Para organizar esto en tu página, podrías estructurarlo así:
# Acción	            Quién lo hace	      Resultado
# Configuración	      Admin / Secretaria	  Sube el archivo Excel con el diseño de la institución a la plataforma.
# Generación	          Docente	          Al entrar a "Calificaciones", el sistema detecta su nombre y grado, toma la plantilla del Admin y genera el archivo ya pre-llenado.
# Cierre de Periodo      Automático	          Al finalizar el periodo, el sistema genera el PDF/Excel final y lo envía a la carpeta correspondiente en Google Drive (Docstry > 2026 > Periodo 1...).
# ═══════════════════════════════════════════════════════════════════════════════
import sys
import os
import mysql.connector
import shutil
import tempfile
from copy import copy
import openpyxl
from flask import Blueprint, jsonify, request, send_file, session, after_this_request
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter

# Importaciones del proyecto
from utils.database import get_db
from utils.helpers import _error_interno

calificaciones_bp = Blueprint('calificaciones', __name__)

# ── GESTIÓN DE RUTAS ABSOLUTAS ────────
# _here es: .../backend/routes
_here = os.path.dirname(os.path.abspath(__file__))
# _parent es: .../backend (la primera carpeta)
_parent = os.path.dirname(_here)

if _parent not in sys.path:
    sys.path.insert(0, _parent)

# ── CONFIGURACIÓN DE LA PLANTILLA INSTITUCIONAL ────────
# Intentamos la ruta basada en tu estructura de doble carpeta: backend/backend/templates
PLANTILLA_PATH = os.path.join(_parent, 'backend', 'templates', 'PlantillaCalificaciones.xlsx')

# Verificación de seguridad: si no existe en la ruta doble, probamos la ruta simple
if not os.path.exists(PLANTILLA_PATH):
    PLANTILLA_PATH = os.path.join(_parent, 'templates', 'PlantillaCalificaciones.xlsx')

# VALIDACIÓN CRÍTICA PARA LA TERMINAL
if not os.path.exists(PLANTILLA_PATH):
    print(f" ERROR CRÍTICO: No se encuentra la plantilla en ninguna de estas rutas:")
    print(f"   Intentado: {PLANTILLA_PATH}")
    print("   Acción: Verifica que el archivo esté en 'backend/backend/templates/'")
else:
    print(f" ¡LOGRADO! Plantilla detectada correctamente en: {PLANTILLA_PATH}")

# ── CONFIGURACIÓN DE CARPETAS EN EL ESCRITORIO (Local) ────────
def obtener_ruta_escritorio():
    usuario_dir = os.path.expanduser("~")
    rutas_posibles = [
        os.path.join(usuario_dir, "OneDrive", "Escritorio"),
        os.path.join(usuario_dir, "Desktop"),
        os.path.join(usuario_dir, "Escritorio")
    ]
    for ruta in rutas_posibles:
        if os.path.exists(ruta):
            return ruta
    return usuario_dir

ESCRITORIO = obtener_ruta_escritorio()
PLANILLAS_DIR = os.path.join(ESCRITORIO, 'Planillas_DocstrY')
HISTORIAL_DIR = os.path.join(ESCRITORIO, 'Planillas_DocstrY_Historial')
ACUERDOS_DIR = os.path.join(_parent, 'backend', 'uploads', 'acuerdos_pedagogicos')

os.makedirs(PLANILLAS_DIR, exist_ok=True)
os.makedirs(HISTORIAL_DIR, exist_ok=True)
os.makedirs(ACUERDOS_DIR, exist_ok=True)


def _to_int(value, default=None):
    try:
        return int(value)
    except Exception:
        return default


def _to_float(value, default=None):
    if value is None:
        return default
    try:
        s = str(value).strip().replace(',', '.')
        if s == '':
            return default
        return float(s)
    except Exception:
        return default


def _to_iso_date_text(value, default=''):
    """Convierte fecha (datetime/str) a texto YYYY-MM-DD."""
    if value is None:
        return default

    try:
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
    except Exception:
        pass

    s = str(value).strip()
    if not s:
        return default

    formatos = (
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%d/%m/%Y',
        '%d-%m-%Y',
    )

    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except Exception:
            continue

    try:
        return datetime.fromisoformat(s.replace('Z', '')).strftime('%Y-%m-%d')
    except Exception:
        return default


def _acuerdo_pdf_filename(grupo_id, materia_id, periodo_id):
    return f"acuerdo_g{int(grupo_id)}_m{int(materia_id)}_p{int(periodo_id)}.pdf"


def _acuerdo_pdf_path(grupo_id, materia_id, periodo_id):
    return os.path.join(ACUERDOS_DIR, _acuerdo_pdf_filename(grupo_id, materia_id, periodo_id))


_ROLES_ADMIN_EDICION = {'server_admin', 'admin_server', 'admin'}


def _es_rol_admin_edicion():
    rol = str(session.get('user_role') or '').strip().lower()
    return rol in _ROLES_ADMIN_EDICION


def _estado_periodo(cursor, id_periodo):
    cursor.execute("SELECT estado FROM periodos WHERE id_periodo = %s LIMIT 1", (id_periodo,))
    row = cursor.fetchone()
    if not row:
        return None
    return str(row.get('estado') or '').strip().lower()


def _bloquear_docente_si_periodo_cerrado(cursor, id_periodo):
    estado = _estado_periodo(cursor, id_periodo)
    if estado is None:
        return jsonify({'error': 'Período no encontrado'}), 404
    if estado == 'cerrado' and not _es_rol_admin_edicion():
        return jsonify({'error': 'El período está cerrado. Solo el administrador puede realizar cambios.'}), 403
    return None


def _resolver_id_usuario_para_actividad(cursor, grupo_id, materia_id):
    """Resuelve el id_usuario para crear actividades de forma segura."""
    usuario_sesion = session.get('user_id')
    if usuario_sesion:
        return int(usuario_sesion)

    cursor.execute(
        """
        SELECT id_usuario
        FROM asignaciones_docente
        WHERE id_grupo = %s AND id_materia = %s AND estado = 'Activa'
        ORDER BY id_asignacion DESC
        LIMIT 1
        """,
        (grupo_id, materia_id)
    )
    row = cursor.fetchone()
    if row and row.get('id_usuario'):
        return int(row['id_usuario'])

    cursor.execute("SELECT id_usuario FROM usuarios WHERE is_activo = 1 ORDER BY id_usuario LIMIT 1")
    row = cursor.fetchone()
    if row and row.get('id_usuario'):
        return int(row['id_usuario'])

    return None


def _asegurar_tabla_actividades_periodo(cursor):
    """Crea la tabla de mapeo actividad-periodo si aún no existe."""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS actividades_periodo (
            id_actividad_periodo INT NOT NULL AUTO_INCREMENT,
            id_actividad INT NOT NULL,
            id_periodo INT NOT NULL,
            created_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (id_actividad_periodo),
            UNIQUE KEY uk_actividad_periodo (id_actividad, id_periodo),
            KEY idx_ap_periodo (id_periodo),
            KEY idx_ap_actividad (id_actividad)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )


def _seed_actividades_para_periodo(cursor, grupo_id, materia_id, id_periodo):
    """Inicializa mapeo mínimo: solo migra actividades legacy al periodo 1.

    Regla funcional: los periodos nuevos (2,3,4,...) inician sin actividades.
    """
    _asegurar_tabla_actividades_periodo(cursor)

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM actividades_periodo ap
        JOIN actividades a ON a.id_actividad = ap.id_actividad
        WHERE ap.id_periodo = %s
          AND a.id_grupo = %s
          AND a.id_materia = %s
        """,
        (id_periodo, grupo_id, materia_id)
    )
    row_actual = cursor.fetchone() or {}
    if int(row_actual.get('total') or 0) > 0:
        return

    # Solo para periodo 1: migrar actividades legacy (sin mapeo) a actividades_periodo.
    if int(id_periodo) != 1:
        return

    cursor.execute(
        """
        SELECT id_actividad
        FROM actividades
        WHERE id_grupo = %s AND id_materia = %s
        ORDER BY fecha_creacion, id_actividad
        """,
        (grupo_id, materia_id)
    )
    actividades_legacy = cursor.fetchall() or []

    for act in actividades_legacy:
        cursor.execute(
            """
            INSERT IGNORE INTO actividades_periodo (id_actividad, id_periodo)
            VALUES (%s, 1)
            """,
            (act['id_actividad'],)
        )


def _obtener_actividades_periodo(cursor, grupo_id, materia_id, id_periodo):
    """Obtiene actividades del periodo actual (aisladas de otros periodos)."""
    _seed_actividades_para_periodo(cursor, grupo_id, materia_id, id_periodo)
    cursor.execute(
        """
        SELECT a.id_actividad, a.nombre_actividad, a.tipo_actividad,
               a.puntaje_maximo, a.ponderacion, a.fecha_vencimiento
        FROM actividades_periodo ap
        JOIN actividades a ON a.id_actividad = ap.id_actividad
        WHERE ap.id_periodo = %s
          AND a.id_grupo = %s
          AND a.id_materia = %s
        ORDER BY a.fecha_creacion, a.id_actividad
        """,
        (id_periodo, grupo_id, materia_id)
    )
    return cursor.fetchall() or []


def _upsert_nota(cursor, id_estudiante, id_actividad, id_materia, id_periodo, puntaje_obtenido):
    """Inserta/actualiza (o elimina) nota para mantener consistencia con la tabla `notas`."""
    cursor.execute(
        "SELECT id_nota FROM notas WHERE id_estudiante = %s AND id_actividad = %s AND id_periodo = %s LIMIT 1",
        (id_estudiante, id_actividad, id_periodo)
    )
    existing = cursor.fetchone()

    if puntaje_obtenido is None:
        if existing:
            cursor.execute("DELETE FROM notas WHERE id_nota = %s", (existing['id_nota'],))
        return

    if existing:
        cursor.execute(
            """
            UPDATE notas
            SET puntaje_obtenido = %s,
                id_materia = %s,
                id_periodo = %s,
                fecha_calificacion = NOW()
            WHERE id_nota = %s
            """,
            (puntaje_obtenido, id_materia, id_periodo, existing['id_nota'])
        )
    else:
        cursor.execute(
            """
            INSERT INTO notas (id_estudiante, id_actividad, id_materia, id_periodo, puntaje_obtenido, fecha_calificacion)
            VALUES (%s, %s, %s, %s, %s, NOW())
            """,
            (id_estudiante, id_actividad, id_materia, id_periodo, puntaje_obtenido)
        )


def _build_planillas_tree(root_path, max_depth=4):
    """Construye un árbol simple (lista) de carpetas y archivos bajo root_path."""
    root_path = os.path.abspath(root_path)
    tree = []
    for dirpath, dirnames, filenames in os.walk(root_path):
        rel = os.path.relpath(dirpath, root_path)
        depth = 0 if rel == '.' else rel.count(os.sep) + 1
        if depth > max_depth:
            dirnames[:] = []
            continue
        entry = {
            'path': rel if rel != '.' else '',
            'dirs': dirnames.copy(),
            'files': filenames.copy()
        }
        tree.append(entry)
    return tree


def _resolver_ruta_plantilla_calificaciones():
    """Resuelve la ruta de la plantilla base de calificaciones en diferentes layouts del proyecto."""
    rutas_posibles = [
        PLANTILLA_PATH,
        os.path.join(_parent, 'templates', 'PlantillaCalificaciones.xlsx'),
        os.path.join(_parent, 'backend', 'templates', 'PlantillaCalificaciones.xlsx')
    ]
    for ruta in rutas_posibles:
        if os.path.exists(ruta):
            return ruta
    return None


def _replace_placeholders_worksheet(ws, mapping):
    """Reemplaza tokens {{TOKEN}} en todas las celdas de tipo texto."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                nuevo_valor = cell.value
                for token, valor in mapping.items():
                    if token in nuevo_valor:
                        nuevo_valor = nuevo_valor.replace(token, str(valor or ''))
                if nuevo_valor != cell.value:
                    cell.value = nuevo_valor


def _detectar_layout_plantilla(ws):
    """Layout fijo de la plantilla institucional."""
    return {
        'header_row': 6,
        'start_row': 8,
        'col_cod': 1,
        'col_estudiantes': 2,
        'col_final_nota': 11,
        'col_fallas': 12,
        'actividad_cols': list(range(4, 10))
    }


# === DESHABILITADO PARA TECNÓLOGO: Integración con Google Drive ===
# def _subir_excel_a_drive_calificaciones(ruta_archivo, nombre_archivo, periodo_id, user_id):
#     """Sube/actualiza un Excel en Drive en Docstry > Año > Periodo X > Calificaciones."""
#     from routes.servicio_drive import get_service, setup_folder_structure
#     from googleapiclient.http import MediaFileUpload
# 
#     service = get_service(user_id)
#     anio_actual = datetime.now().year
#     subcarpetas = setup_folder_structure(user_id, anio_actual, f"Periodo {periodo_id}")
#     carpeta_calificaciones_id = subcarpetas.get('Calificaciones')
# 
#     query = f"name='{nombre_archivo}' and '{carpeta_calificaciones_id}' in parents and trashed=false"
#     results = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
#     items = results.get('files', [])
# 
#     media = MediaFileUpload(
#         ruta_archivo,
#         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#         resumable=True
#     )
# 
#     if items:
#         resp = service.files().update(fileId=items[0]['id'], media_body=media, fields='id,name').execute()
#     else:
#         file_metadata = {'name': nombre_archivo, 'parents': [carpeta_calificaciones_id]}
#         resp = service.files().create(body=file_metadata, media_body=media, fields='id,name').execute()
#     return resp
# ================================================================


def _generar_planilla_desde_plantilla(grupo_id, materia_id, periodo_id, fecha_actividad=None, user_id=None, actividad_ids_order=None):
    """Genera una planilla desde PlantillaCalificaciones.xlsx y la guarda como archivo temporal."""
    plantilla_path = _resolver_ruta_plantilla_calificaciones()
    if not plantilla_path:
        raise FileNotFoundError(
            'No se encontró PlantillaCalificaciones.xlsx en backend/templates ni backend/backend/templates'
        )

    conn = None
    cursor = None
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT gr.id_grupo, gr.codigo_grupo, g.numero_grado, g.nombre_grado
            FROM grupos gr
            JOIN grados g ON g.id_grado = gr.id_grado
            WHERE gr.id_grupo = %s
            LIMIT 1
            """,
            (grupo_id,)
        )
        info_grupo = cursor.fetchone()
        if not info_grupo:
            raise ValueError('Grupo no encontrado')

        cursor.execute(
            """
            SELECT id_materia, nombre_materia, codigo_materia
            FROM materias
            WHERE id_materia = %s
            LIMIT 1
            """,
            (materia_id,)
        )
        info_materia = cursor.fetchone()
        if not info_materia:
            raise ValueError('Materia no encontrada')

        cursor.execute(
            """
            SELECT id_periodo, numero_periodo, nombre_periodo, fecha_inicio
            FROM periodos
            WHERE id_periodo = %s
            LIMIT 1
            """,
            (periodo_id,)
        )
        info_periodo = cursor.fetchone()
        if not info_periodo:
            raise ValueError('Período no encontrado')

        docente_id = user_id
        if not docente_id:
            cursor.execute(
                """
                SELECT id_usuario
                FROM asignaciones_docente
                WHERE id_grupo = %s AND id_materia = %s AND estado = 'Activa'
                ORDER BY id_asignacion DESC
                LIMIT 1
                """,
                (grupo_id, materia_id)
            )
            asig = cursor.fetchone()
            if asig and asig.get('id_usuario'):
                docente_id = asig.get('id_usuario')

        info_usuario = {}
        if docente_id:
            cursor.execute("SELECT * FROM usuarios WHERE id_usuario = %s LIMIT 1", (docente_id,))
            info_usuario = cursor.fetchone() or {}

        cursor.execute(
            """
            SELECT e.id_estudiante, e.nombre, e.apellido, COALESCE(ap.total_ausencias, 0) AS fallas
            FROM estudiantes e
            LEFT JOIN asistencias_por_periodo ap ON ap.id_estudiante = e.id_estudiante
                AND ap.id_materia = %s
                AND ap.id_periodo = %s
            WHERE e.id_grupo = %s AND e.estado = 'Activo'
            ORDER BY e.apellido, e.nombre
            """,
            (materia_id, periodo_id, grupo_id)
        )
        estudiantes = cursor.fetchall() or []

        actividades = _obtener_actividades_periodo(cursor, grupo_id, materia_id, periodo_id)
        conn.commit()

        # Si la UI envía un orden explícito visible en pantalla, respetarlo.
        if actividad_ids_order:
            orden_map = {int(act_id): idx for idx, act_id in enumerate(actividad_ids_order) if _to_int(act_id) is not None}
            actividades.sort(key=lambda a: orden_map.get(int(a.get('id_actividad')), 10**9))

        notas_map = {}
        if estudiantes and actividades:
            estudiantes_ids = [int(e['id_estudiante']) for e in estudiantes]
            actividades_ids = [int(a['id_actividad']) for a in actividades]
            placeholders_est = ','.join(['%s'] * len(estudiantes_ids))
            placeholders_act = ','.join(['%s'] * len(actividades_ids))

            query_notas = f"""
                SELECT id_estudiante, id_actividad, puntaje_obtenido
                FROM notas
                WHERE id_periodo = %s
                  AND id_materia = %s
                  AND id_estudiante IN ({placeholders_est})
                  AND id_actividad IN ({placeholders_act})
            """
            params_notas = [periodo_id, materia_id] + estudiantes_ids + actividades_ids
            cursor.execute(query_notas, params_notas)
            for n in cursor.fetchall() or []:
                notas_map[(int(n['id_estudiante']), int(n['id_actividad']))] = n.get('puntaje_obtenido')

        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active

        def _pick(d, keys, default=''):
            for k in keys:
                if d.get(k):
                    return d.get(k)
            return default

        nombre_docente = (
            f"{info_usuario.get('apellido', '')} {info_usuario.get('nombre', '')}".strip()
            or _pick(info_usuario, ['nombre_completo', 'usuario', 'email'], '')
        )
        nombre_grado = (
            info_grupo.get('nombre_grado')
            or f"Grado {info_grupo.get('numero_grado', '')} - Grupo {info_grupo.get('codigo_grupo', '')}"
        )
        grupo_label = f"Grupo {info_grupo.get('codigo_grupo', '')}".strip()

        fecha_inicio = info_periodo.get('fecha_inicio')
        anio = str(getattr(fecha_inicio, 'year', datetime.now().year))
        fecha_actividad = _to_iso_date_text(fecha_actividad, datetime.now().strftime('%Y-%m-%d'))

        for act in actividades:
            act['fecha_actividad'] = _to_iso_date_text(act.get('fecha_vencimiento'), fecha_actividad)

        fecha_reporte = next((a.get('fecha_actividad') for a in actividades if a.get('fecha_actividad')), fecha_actividad)

        periodo_label = f"PERIODO {info_periodo.get('numero_periodo', periodo_id)}"

        mapeo_datos = {
            '{{INSTITUCION}}': _pick(info_usuario, ['institucion'], 'DocstrY'),
            '{{FECHA}}': fecha_reporte,
            '{{SEDE}}': _pick(info_usuario, ['sede', 'nombre_sede'], 'Principal'),
            '{{JORNADA}}': _pick(info_usuario, ['jornada', 'tipo_jornada'], ''),
            '{{GRADO}}': nombre_grado,
            '{{GRUPO}}': grupo_label,
            '{{DOCENTE}}': nombre_docente,
            '{{MATERIA}}': info_materia.get('nombre_materia', ''),
            '{{PERIODO}}': periodo_label,
            '{{ANIO}}': anio,
        }

        _replace_placeholders_worksheet(ws, mapeo_datos)

        # Fallback explícito para la nueva casilla de Grupo (fila 2, columnas G-H).
        # Si el placeholder no existe en plantilla, dejar el valor directamente en G2.
        try:
            ws.cell(row=2, column=7, value=grupo_label)
        except Exception:
            pass

        layout = _detectar_layout_plantilla(ws)

        # La plantilla nueva define 10 columnas fijas para actividades (D..M).
        # No insertar columnas para no romper diseño ni formato de impresión.
        actividades = actividades[:10]
        actividades_total = len(actividades)
        actividad_cols = list(range(4, 4 + actividades_total)) if actividades_total else []
        cols_actividades_plantilla = list(range(4, 14))  # D..M (10 actividades)
        col_equivalencia = 14  # N
        col_final_nota = 15    # O
        col_fallas = 16        # P

        def _formula_equivalencia(row_num):
            if not actividad_cols or not actividades:
                return None

            partes = []
            total_pond = sum(float(a.get('ponderacion') or 0) for a in actividades)
            pond_por_defecto = (100.0 / len(actividades)) if (actividades and total_pond <= 0) else None

            for idx, col in enumerate(actividad_cols):
                act = actividades[idx]
                puntaje_max = float(act.get('puntaje_maximo') or 1)
                ponderacion = float(act.get('ponderacion') or 0)
                if pond_por_defecto is not None:
                    ponderacion = pond_por_defecto

                col_letter = get_column_letter(col)
                partes.append(f"IFERROR(({col_letter}{row_num}/{puntaje_max}),0)*{ponderacion}")

            return f"=IFERROR(({'+'.join(partes)})/100,\"\")"

        def _formula_nota_final(row_num):
            if not actividad_cols:
                return None
            ultima_col = get_column_letter(actividad_cols[-1])
            return f"=IFERROR(AVERAGE(D{row_num}:{ultima_col}{row_num}),\"\")"

        max_clear_row = max(layout['start_row'] + max(len(estudiantes), 1) + 80, layout['start_row'] + 1)
        columnas_a_limpiar = [
            layout['col_cod'],
            layout['col_estudiantes'],
            col_equivalencia,
            col_final_nota,
            col_fallas,
        ] + cols_actividades_plantilla

        for fila in range(layout['start_row'], max_clear_row + 1):
            for col in columnas_a_limpiar:
                ws.cell(row=fila, column=col, value=None)

        # Encabezados/fechas: limpiar solo zona de actividades para que no herede basura previa.
        for col in cols_actividades_plantilla:
            ws.cell(row=layout['header_row'], column=col, value=None)
            ws.cell(row=5, column=col, value=None)

        for idx, col in enumerate(actividad_cols):
            ws.cell(row=layout['header_row'], column=col, value=actividades[idx].get('nombre_actividad'))
            ws.cell(row=5, column=col, value=actividades[idx].get('fecha_actividad') or fecha_actividad)

        for offset, est in enumerate(estudiantes):
            row_num = layout['start_row'] + offset
            est_id = int(est['id_estudiante'])
            nombre = f"{(est.get('apellido') or '').strip()} {(est.get('nombre') or '').strip()}".strip()

            ws.cell(row=row_num, column=layout['col_cod'], value=est_id)
            ws.cell(row=row_num, column=layout['col_estudiantes'], value=nombre)
            ws.cell(row=row_num, column=col_fallas, value=int(est.get('fallas') or 0))

            formula_equiv = _formula_equivalencia(row_num)
            if formula_equiv:
                ws.cell(row=row_num, column=col_equivalencia, value=formula_equiv)

            formula_final = _formula_nota_final(row_num)
            if formula_final:
                ws.cell(row=row_num, column=col_final_nota, value=formula_final)

            for idx, col in enumerate(actividad_cols):
                id_actividad = int(actividades[idx]['id_actividad'])
                ws.cell(row=row_num, column=col, value=notas_map.get((est_id, id_actividad)))

        prefijo = f"Planilla_G{info_grupo.get('numero_grado')}_{info_grupo.get('codigo_grupo')}_{info_materia.get('nombre_materia', 'Materia')}"
        prefijo = prefijo.replace(' ', '_').replace('/', '-')
        with tempfile.NamedTemporaryFile(prefix=prefijo + '_', suffix='.xlsx', delete=False) as tmp:
            ruta_temporal = tmp.name
        wb.save(ruta_temporal)

        return {
            'ruta_temporal': ruta_temporal,
            'nombre_archivo': os.path.basename(ruta_temporal),
            'mapeo_datos': mapeo_datos,
            'estudiantes_total': len(estudiantes),
            'actividades_total': len(actividades)
        }
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@calificaciones_bp.route('/api/calificaciones/plantilla/generar', methods=['POST'])
def api_generar_plantilla_desde_base():
    """
    Genera un Excel desde PlantillaCalificaciones.xlsx y lo devuelve como descarga directa.
    """
    data = request.get_json() or {}
    grupo_id = _to_int(data.get('grupo_id'))
    materia_id = _to_int(data.get('materia_id'))
    periodo_id = _to_int(data.get('periodo_id'), 1)
    fecha_actividad = str(data.get('fecha_actividad') or datetime.now().strftime('%Y-%m-%d')).strip()
    actividad_ids_order = data.get('actividad_ids_order') if isinstance(data.get('actividad_ids_order'), list) else None

    if not grupo_id or not materia_id or not periodo_id:
        return jsonify({'error': 'Faltan parámetros (grupo_id, materia_id, periodo_id)'}), 400

    try:
        resultado = _generar_planilla_desde_plantilla(
            grupo_id=grupo_id,
            materia_id=materia_id,
            periodo_id=periodo_id,
            fecha_actividad=fecha_actividad,
            user_id=session.get('user_id'),
            actividad_ids_order=actividad_ids_order
        )

        @after_this_request
        def _cleanup_temp_file(resp):
            try:
                if os.path.exists(resultado['ruta_temporal']):
                    os.remove(resultado['ruta_temporal'])
            except Exception:
                pass
            return resp

        return send_file(
            resultado['ruta_temporal'],
            as_attachment=True,
            download_name='Planilla_Calificaciones.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 404
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return _error_interno(e)

def _crear_excel_fisico(grado_id, grupo_id, materia_id, periodo_id=1, force_recreate=False, save_path=None):
    """Función interna para crear un archivo Excel si no existe o se requiere forzar con notas y fórmulas"""
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    # 1. Obtener la info del Grado, Grupo y Materia para el nombre del archivo
    info = {}
    try:
        cursor.execute("SELECT id_grado, numero_grado FROM grados WHERE id_grado = %s", (grado_id,))
        g = cursor.fetchone()
        if g:
            info['id_grado'] = g.get('id_grado')
            info['numero_grado'] = g.get('numero_grado')

        cursor.execute("SELECT id_grupo, codigo_grupo, id_grado FROM grupos WHERE id_grupo = %s", (grupo_id,))
        gr = cursor.fetchone()
        if gr:
            info['id_grupo'] = gr.get('id_grupo')
            info['codigo_grupo'] = gr.get('codigo_grupo')
            if 'numero_grado' not in info and gr.get('id_grado'):
                cursor.execute("SELECT numero_grado FROM grados WHERE id_grado=%s", (gr.get('id_grado'),))
                gg = cursor.fetchone()
                if gg:
                    info['numero_grado'] = gg.get('numero_grado')

        cursor.execute("SELECT id_materia, nombre_materia, codigo_materia FROM materias WHERE id_materia = %s", (materia_id,))
        m = cursor.fetchone()
        if m:
            info['id_materia'] = m.get('id_materia')
            info['nombre_materia'] = m.get('nombre_materia')
            info['codigo_materia'] = m.get('codigo_materia')
    except Exception as e:
        print('[calificaciones] Error consultando metadata:', e)

    # 2. Obtener ACTIVIDADES dinámicamente para este grupo/materia
    actividades = []
    try:
        actividades = _obtener_actividades_periodo(cursor, grupo_id, materia_id, periodo_id)
        conn.commit()
    except Exception as e:
        print('[calificaciones] Error consultando actividades:', e)

    # 3. Estudiantes: intentar varias estrategias para obtener el listado
    estudiantes = []
    try:
        cursor.execute("""
            SELECT id_estudiante, nombre, apellido, documento
            FROM estudiantes
            WHERE id_grupo = %s AND estado = 'Activo'
            ORDER BY apellido, nombre
        """, (grupo_id,))
        estudiantes = cursor.fetchall() or []

        if not estudiantes and grado_id:
            try:
                cursor.execute("""
                    SELECT id_estudiante, nombre, apellido, documento
                    FROM estudiantes
                    WHERE id_grado = %s AND estado = 'Activo'
                    ORDER BY apellido, nombre
                """, (grado_id,))
                estudiantes = cursor.fetchall() or []
            except Exception as e2:
                print('[calificaciones] Error fallback por grado:', e2)

        if not estudiantes:
            try:
                cursor.execute("""
                    SELECT id_estudiante, nombre, apellido, documento
                    FROM estudiantes
                    WHERE estado = 'Activo'
                    ORDER BY apellido, nombre
                    LIMIT 500
                """)
                estudiantes = cursor.fetchall() or []
            except Exception as e3:
                print('[calificaciones] Error fallback amplio estudiantes:', e3)

    except Exception as e:
        print('[calificaciones] Error consultando estudiantes:', e)

    # 4. Obtener NOTAS existentes de la BD
    notas_map = {}
    try:
        cursor.execute("""
            SELECT n.id_estudiante, n.id_actividad, n.puntaje_obtenido
            FROM notas n
            JOIN actividades a ON a.id_actividad = n.id_actividad
            WHERE a.id_grupo = %s AND a.id_materia = %s AND n.id_periodo = %s
        """, (grupo_id, materia_id, periodo_id))
        notas_rows = cursor.fetchall() or []
        for nota in notas_rows:
            key = (nota['id_estudiante'], nota['id_actividad'])
            notas_map[key] = nota['puntaje_obtenido']
    except Exception as e:
        print('[calificaciones] Error consultando notas:', e)

    cursor.close(); conn.close()

    # 5. Preparar la ruta
    if save_path:
        # Si se proporciona una ruta personalizada, usarla directamente
        ruta_archivo = save_path
        os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)
    else:
        # Usar la estructura estándar de PLANILLAS_DIR
        nombre_carpeta_grado = f"Grado_{info.get('numero_grado', grado_id)}"
        nombre_carpeta_grupo = f"Grupo_{info.get('codigo_grupo', grupo_id)}"
        ruta_directorio = os.path.join(PLANILLAS_DIR, nombre_carpeta_grado, nombre_carpeta_grupo)
        os.makedirs(ruta_directorio, exist_ok=True)
        
        numero_grado = str(info.get('numero_grado') or f"Gr{grado_id}")
        codigo_grupo = str(info.get('codigo_grupo') or f"G{grupo_id}")
        nombre_materia = str(info.get('nombre_materia') or f"Materia_{materia_id}")

        materia_limpia = nombre_materia.replace(" ", "_").replace("/", "-")
        nombre_archivo = f"{materia_limpia}_G{numero_grado}_{codigo_grupo}_P{periodo_id}.xlsx"
        ruta_archivo = os.path.join(ruta_directorio, nombre_archivo)

    # 6. Crear o actualizar Excel (forzado si force_recreate=True)
    if os.path.exists(ruta_archivo) and not force_recreate:
        return ruta_archivo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notas - P{periodo_id}"
    
    header_fill = PatternFill(start_color="8B6F47", end_color="8B6F47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    
    # Construir encabezados dinámicamente
    headers = ['ID_Estudiante', 'Estudiante']
    header_col_map = {}  # Mapeo de id_actividad -> columna Excel
    
    for idx, act in enumerate(actividades):
        col_idx = 3 + idx
        header_col_map[act['id_actividad']] = col_idx
        headers.append(f"{act['nombre_actividad']} (P:{act['ponderacion']}%)")
    
    final_col = 3 + len(actividades)
    headers.append('Nota Final')
    
    ws.append(headers)
    
    # Aplicar estilos al encabezado
    for col, title in enumerate(headers, start=1):
        celda = ws.cell(row=1, column=col)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Rellenar datos de estudiantes y notas
    for row_num, est in enumerate(estudiantes, start=2):
        est_id = None
        if isinstance(est, dict):
            for key in ('id_estudiante', 'id', 'id_alumno', 'estudiante_id'):
                if key in est and est.get(key) is not None:
                    est_id = est.get(key)
                    break

        # Columna ID
        ws.cell(row=row_num, column=1, value=est_id)

        # Columna Nombre
        nombre_val = None
        if isinstance(est, dict):
            if est.get('apellido') and est.get('nombre'):
                nombre_val = f"{est.get('apellido')} {est.get('nombre')}"
            elif est.get('nombre'):
                nombre_val = est.get('nombre')
        ws.cell(row=row_num, column=2, value=nombre_val)

        # Rellenar notas de actividades
        for act in actividades:
            col_idx = header_col_map[act['id_actividad']]
            nota_key = (est_id, act['id_actividad'])
            nota_val = notas_map.get(nota_key)
            
            celda = ws.cell(row=row_num, column=col_idx, value=nota_val)
            if nota_val is not None:
                celda.number_format = '0.00'
            celda.alignment = Alignment(horizontal='center')

        # Columna Nota Final con fórmula de ponderación
        celda_final = ws.cell(row=row_num, column=final_col)
        
        if actividades:
            # Construir fórmula: SUM((nota1/max1)*pond1, (nota2/max2)*pond2, ...)
            formula_parts = []
            for act in actividades:
                col_idx = header_col_map[act['id_actividad']]
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                puntaje_max = act.get('puntaje_maximo') or 1
                ponderacion = act.get('ponderacion') or 0
                
                # Fórmula: (celda/puntaje_max)*ponderacion
                formula_parts.append(f"(({col_letter}{row_num}/{puntaje_max})*{ponderacion})")
            
            formula = f"=({'+'.join(formula_parts)})/100"
            celda_final.value = formula
            celda_final.number_format = '0.00'
        
        celda_final.alignment = Alignment(horizontal='center')
        celda_final.font = Font(bold=True)

    # Formateo de columnas
    ws.column_dimensions['A'].hidden = True
    ws.column_dimensions['B'].width = 35
    for col in range(3, final_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    ws.row_dimensions[1].height = 30

    wb.save(ruta_archivo)
    return ruta_archivo


# Ruta de utilidad: devolver estructura de carpetas y archivos bajo PLANILLAS_DIR
@calificaciones_bp.route('/api/calificaciones/estructura_carpetas', methods=['GET'])
def api_estructura_carpetas():
    try:
        if not os.path.exists(PLANILLAS_DIR):
            return jsonify({'status': 'ok', 'tree': [], 'message': 'No existe la carpeta de planillas en este equipo.'}), 200

        tree = _build_planillas_tree(PLANILLAS_DIR, max_depth=4)
        return jsonify({'status': 'ok', 'tree': tree, 'base': PLANILLAS_DIR}), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/estructura_carpetas/public', methods=['GET'])
def api_estructura_carpetas_public():
    """Endpoint público de diagnóstico: lista la estructura de `PLANILLAS_DIR` sin requerir sesión.
    Atención: dejar temporalmente solo para depuración local; no recomendable en producción.
    """
    try:
        if not os.path.exists(PLANILLAS_DIR):
            return jsonify({'status': 'ok', 'tree': [], 'message': 'No existe la carpeta de planillas en este equipo.'}), 200
        tree = _build_planillas_tree(PLANILLAS_DIR, max_depth=4)
        return jsonify({'status': 'ok', 'tree': tree, 'base': PLANILLAS_DIR}), 200
    except Exception as e:
        return _error_interno(e)


# ── 1. SINCRONIZADOR MASIVO (Desde la DB hacia el sistema local) ────────

@calificaciones_bp.route('/api/calificaciones/sincronizar_carpetas', methods=['POST'])
def api_sincronizar_carpetas():
    """
    Crea la estructura de carpetas para todos los grados y grupos que existan en el sistema.
    Ahora también crea la estructura por períodos: /uploads/periodos/periodo_1/Grado_X/Grupo_Y/
    """
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 1️⃣ OBTENER TODOS LOS PERÍODOS
        cursor.execute("SELECT id_periodo, numero_periodo FROM periodos ORDER BY numero_periodo")
        periodos = cursor.fetchall()
        
        if not periodos:
            cursor.close()
            conn.close()
            return jsonify({'error': 'No hay períodos en el sistema'}), 400
        
        # 2️⃣ OBTENER TODOS LOS GRUPOS
        cursor.execute("""
            SELECT g.numero_grado, gr.codigo_grupo, g.id_grado, gr.id_grupo 
            FROM grupos gr
            JOIN grados g ON gr.id_grado = g.id_grado
        """)
        todos_los_grupos = cursor.fetchall()
        
        # 3️⃣ CREAR ESTRUCTURA POR PERÍODO
        PERIODOS_DIR = os.path.join(ESCRITORIO, 'Periodos_DocstrY')
        os.makedirs(PERIODOS_DIR, exist_ok=True)
        
        carpetas_creadas = 0
        archivos_creados = 0
        created_files = []
        errors = []
        
        for periodo in periodos:
            periodo_id = periodo['id_periodo']
            numero_periodo = periodo['numero_periodo']
            periodo_folder = os.path.join(PERIODOS_DIR, f"Período_{numero_periodo}")
            os.makedirs(periodo_folder, exist_ok=True)
            
            # Crear estructura Grado/Grupo dentro de cada período
            for grupo in todos_los_grupos:
                grado_num = grupo['numero_grado']
                grupo_cod = grupo['codigo_grupo']
                id_grado = grupo['id_grado']
                id_grupo = grupo['id_grupo']
                
                ruta_grupo = os.path.join(periodo_folder, f"Grado_{grado_num}", f"Grupo_{grupo_cod}")
                if not os.path.exists(ruta_grupo):
                    os.makedirs(ruta_grupo, exist_ok=True)
                    carpetas_creadas += 1
                
                # Obtener asignaciones para este grupo en este período
                cursor.execute("""
                    SELECT DISTINCT id_materia FROM asignaciones_docente 
                    WHERE id_grupo = %s AND estado = 'Activa'
                """, (id_grupo,))
                asignaciones = cursor.fetchall()
                
                # Si no hay asignaciones, crear para todas las materias
                if not asignaciones:
                    cursor.execute("SELECT id_materia FROM materias")
                    asignaciones = cursor.fetchall()
                
                # Crear Excel para cada materia en este período
                for asig in asignaciones:
                    # Obtener nombre de materia para el nombre del archivo
                    cursor.execute("SELECT nombre_materia FROM materias WHERE id_materia = %s", (asig['id_materia'],))
                    mat_row = cursor.fetchone()
                    nombre_materia = mat_row['nombre_materia'] if mat_row else f"Materia_{asig['id_materia']}"
                    materia_limpia = nombre_materia.replace(" ", "_").replace("/", "-")
                    
                    nombre_archivo = f"{materia_limpia}_G{grado_num}_{grupo_cod}_P{numero_periodo}.xlsx"
                    ruta_archivo = os.path.join(ruta_grupo, nombre_archivo)
                    
                    try:
                        # Solo crear si no existe
                        if not os.path.exists(ruta_archivo):
                            _crear_excel_fisico(id_grado, id_grupo, asig['id_materia'], 
                                              periodo_id=periodo_id,
                                              force_recreate=True,
                                              save_path=ruta_archivo)
                            archivos_creados += 1
                            created_files.append(ruta_archivo)
                    except Exception as ex:
                        errors.append({
                            'periodo': numero_periodo,
                            'grado': grado_num,
                            'grupo': grupo_cod,
                            'materia_id': asig['id_materia'],
                            'error': str(ex)
                        })
        
        # También crear estructura antigua para compatibilidad
        for grupo in todos_los_grupos:
            ruta = os.path.join(PLANILLAS_DIR, f"Grado_{grupo['numero_grado']}", f"Grupo_{grupo['codigo_grupo']}")
            if not os.path.exists(ruta):
                os.makedirs(ruta, exist_ok=True)
        
        cursor.close()
        conn.close()
                
        return jsonify({
            'status': 'ok',
            'message': f'✅ Sincronización completa. {carpetas_creadas} carpetas de período creadas.',
            'archivos_creados': archivos_creados,
            'periodos_procesados': len(periodos),
            'created_files': created_files[:10],  # Mostrar primeros 10
            'errors': errors
        }), 200

    except Exception as e:
        return _error_interno(e)


# ── 2. GENERADOR DE PLANILLAS EXCEL CONGELADA (Offline First) ────────────────

@calificaciones_bp.route('/api/calificaciones/generar_planilla', methods=['GET'])
def api_generar_planilla():
    """
    Endpoint maestro.
    Recibe por parámetros el ID de grado, grupo, materia y periodo.
    Consulta la DB, crea la carpeta si falta, y genera un Excel bloqueado donde
    el profesor solo puede editar las notas.
    """
    grado_id = request.args.get('grado_id')
    grupo_id = request.args.get('grupo_id')
    materia_id = request.args.get('materia_id')
    periodo_id = _to_int(request.args.get('periodo_id', 1), 1)

    if not all([grado_id, grupo_id, materia_id]):
        return jsonify({'error': 'Faltan parámetros (grado_id, grupo_id, materia_id)'}), 400

    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, periodo_id)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo

        # 1. Obtener la info del Grado, Grupo y Materia para el nombre del archivo
        cursor.execute("""
            SELECT g.numero_grado, gr.codigo_grupo, m.nombre_materia, m.codigo_materia
            FROM grados g, grupos gr, materias m
            WHERE g.id_grado = %s AND gr.id_grupo = %s AND m.id_materia = %s
        """, (grado_id, grupo_id, materia_id))
        info = cursor.fetchone()
        
        if not info:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Datos de grado, grupo o materia no existen.'}), 404

        # 2. Consultar todos los estudiantes activos en este grupo
        ruta_archivo = _crear_excel_fisico(grado_id, grupo_id, materia_id, periodo_id)
        
        if not ruta_archivo:
             return jsonify({'error': 'Error interno al generar archivo físico'}), 500
             
        nombre_archivo = os.path.basename(ruta_archivo)

        # Enviar archivo al profesor para descargar
        return send_file(
            ruta_archivo,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return _error_interno(e)

# ── NUEVA RUTA: LEER EXCEL PARA PANTALLA (Grilla Web) ────────────────────────
@calificaciones_bp.route('/api/calificaciones/leer_planilla', methods=['GET'])
def api_leer_planilla():
    grado_id = request.args.get('grado_id')
    grupo_id = request.args.get('grupo_id')
    materia_id = request.args.get('materia_id')
    periodo_id = request.args.get('periodo_id', 1)

    try:
        ruta_archivo = _crear_excel_fisico(grado_id, grupo_id, materia_id, periodo_id)
        if not ruta_archivo or not os.path.exists(ruta_archivo):
            return jsonify({'error': 'No se encontro ni se pudo generar el Excel físico'}), 404
            
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if row[0]: # Porque hay ID del estudiante
                data.append({
                    'id_estudiante': row[0],
                    'nombre': row[1],
                    'act1': row[2] if row[2] is not None else '',
                    'act2': row[3] if row[3] is not None else '',
                    'act3': row[4] if row[4] is not None else '',
                    'act4': row[5] if row[5] is not None else '',
                    'nota': row[6] if row[6] is not None else '' # Nota_Final
                })
                
        return jsonify({'status': 'ok', 'alumnos': data}), 200
        
    except Exception as e:
         return _error_interno(e)

# ── NUEVA RUTA: GUARDAR DESDE PANTALLA WEB DIRECTO ─────────────────────────────
@calificaciones_bp.route('/api/calificaciones/guardar_planilla_web', methods=['POST'])
def api_guardar_planilla_web():
    data_json = request.get_json()
    grado_id = data_json.get('grado_id')
    grupo_id = data_json.get('grupo_id')
    materia_id = data_json.get('materia_id')
    periodo_id = _to_int(data_json.get('periodo_id', 1), 1)
    alumnos = data_json.get('alumnos', [])
    
    try:
        # Encontramos el archivo físico
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, periodo_id)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo

        cursor.execute("SELECT g.numero_grado, gr.codigo_grupo, m.nombre_materia FROM grados g, grupos gr, materias m WHERE g.id_grado=%s AND gr.id_grupo=%s AND m.id_materia=%s", (grado_id, grupo_id, materia_id))
        info = cursor.fetchone()
        
        materia_limpia = str(info['nombre_materia']).replace(" ", "_").replace("/", "-")
        nombre_carpeta_grado = f"Grado_{info['numero_grado']}"
        nombre_carpeta_grupo = f"Grupo_{info['codigo_grupo']}"
        nombre_archivo = f"{materia_limpia}_G{info['numero_grado']}_{info['codigo_grupo']}_P{periodo_id}.xlsx"
        
        ruta_archivo_actual = os.path.join(PLANILLAS_DIR, nombre_carpeta_grado, nombre_carpeta_grupo, nombre_archivo)
        
        if not os.path.exists(ruta_archivo_actual):
             return jsonify({'error': 'El Excel fisico no existe para ser actualizado'}), 404

        # Versionado (Archivador) - igual que al subir archivo manual
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_historico = f"{nombre_archivo.replace('.xlsx', '')}_v{timestamp}.xlsx"
        ruta_historial_dir = os.path.join(HISTORIAL_DIR, nombre_carpeta_grado, nombre_carpeta_grupo)
        os.makedirs(ruta_historial_dir, exist_ok=True)
        shutil.copy2(ruta_archivo_actual, os.path.join(ruta_historial_dir, nombre_historico))

        # Modificamos el Excel Físico
        wb = openpyxl.load_workbook(ruta_archivo_actual)
        ws = wb.active
        
        # Mapeamos los alumos llegados desde la web a un diccionario
        alumnos_map = {int(a['id_estudiante']): a for a in alumnos}
        
        for fila_idx in range(2, ws.max_row + 1):
            id_est = ws.cell(row=fila_idx, column=1).value
            if id_est and int(id_est) in alumnos_map:
                alumno_data = alumnos_map[int(id_est)]
                
                # Helper para convertir a float seguro
                def set_float_val(col, key):
                    val = str(alumno_data.get(key, '')).strip()
                    if val == "": ws.cell(row=fila_idx, column=col).value = None
                    else:
                        try: ws.cell(row=fila_idx, column=col).value = float(val)
                        except: pass
                
                set_float_val(3, 'act1')
                set_float_val(4, 'act2')
                set_float_val(5, 'act3')
                set_float_val(6, 'act4')
                set_float_val(7, 'nota')
                
        wb.save(ruta_archivo_actual)
        
        # Sincronizamos la DB
        actividades_periodo = _obtener_actividades_periodo(cursor, int(grupo_id), int(materia_id), int(periodo_id))
        conn.commit()
        act_row = actividades_periodo[0] if actividades_periodo else None
        if not act_row:
             id_usuario = _resolver_id_usuario_para_actividad(cursor, grupo_id, materia_id)
             if not id_usuario:
                 return jsonify({'error': 'No existe un usuario activo para crear actividad'}), 400

             cursor.execute(
                 """
                 INSERT INTO actividades (id_usuario, id_grupo, id_materia, nombre_actividad, descripcion, tipo_actividad, puntaje_maximo, ponderacion)
                 VALUES (%s, %s, %s, 'Carga desde Web', 'Sincronización desde planilla web', 'Tarea', 5, 100)
                 """,
                 (id_usuario, grupo_id, materia_id)
             )
             id_actividad = cursor.lastrowid
             _asegurar_tabla_actividades_periodo(cursor)
             cursor.execute(
                 """
                 INSERT IGNORE INTO actividades_periodo (id_actividad, id_periodo)
                 VALUES (%s, %s)
                 """,
                 (id_actividad, int(periodo_id))
             )
             conn.commit()
        else:
             id_actividad = act_row['id_actividad']
             
        for est_id, alumno_data in alumnos_map.items():
            n_float = _to_float(alumno_data.get('nota'))
            _upsert_nota(cursor, est_id, id_actividad, materia_id, int(periodo_id), n_float)
                 
        conn.commit()
        cursor.close(); conn.close()
        
        return jsonify({'status': 'ok', 'message': 'Guardado en Excel físico y BD exitosamente'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return _error_interno(e)


# ── 3. SUBIDA Y PROCESAMIENTO CON HISTORIAL (El Archivador Inmortal) ─────────

@calificaciones_bp.route('/api/calificaciones/subir_planilla', methods=['POST'])
def api_subir_planilla():
    """
    Ruta que recibe el archivo Excel que el profesor editó.
    1. Si ya existía uno viejo, lo mueve a /historial_respaldos/ de forma permanente.
    2. Guarda el nuevo en la ruta local correspondiente.
    3. Lee el Excel, extrae las notas, y hace INSERT/UPDATE en la base de datos `notas`
       validando con el ID oculto del estudiante en la Columna 1.
    """
    try:
        # Validación de parámetros y archivo
        grado_id = request.form.get('grado_id')
        grupo_id = request.form.get('grupo_id')
        materia_id = request.form.get('materia_id')
        periodo_id = _to_int(request.form.get('periodo_id', 1), 1)
        
        if 'archivo_excel' not in request.files:
            return jsonify({'error': 'No se envió el archivo de Excel'}), 400
            
        archivo = request.files['archivo_excel']
        if archivo.filename == '':
            return jsonify({'error': 'Archivo sin nombre o inválido'}), 400
            
        if not all([grado_id, grupo_id, materia_id]):
            return jsonify({'error': 'Faltan parámetros (grado_id, grupo_id, materia_id)'}), 400

        # Conectar a la DB para sacar los nombres de carpetas
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, periodo_id)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo
        
        cursor.execute("""
            SELECT g.numero_grado, gr.codigo_grupo, m.nombre_materia
            FROM grados g, grupos gr, materias m
            WHERE g.id_grado = %s AND gr.id_grupo = %s AND m.id_materia = %s
        """, (grado_id, grupo_id, materia_id))
        info = cursor.fetchone()
        
        if not info:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Datos de grado, grupo o materia no válidos'}), 404

        # Construir rutas físicas exactas
        materia_limpia = str(info['nombre_materia']).replace(" ", "_")
        nombre_carpeta_grado = f"Grado_{info['numero_grado']}"
        nombre_carpeta_grupo = f"Grupo_{info['codigo_grupo']}"
        nombre_archivo = f"{materia_limpia}_G{info['numero_grado']}_{info['codigo_grupo']}_P{periodo_id}.xlsx"
        
        ruta_directorio = os.path.join(PLANILLAS_DIR, nombre_carpeta_grado, nombre_carpeta_grupo)
        ruta_archivo_actual = os.path.join(ruta_directorio, nombre_archivo)
        
        os.makedirs(ruta_directorio, exist_ok=True)

        # ---------------------------------------------------------
        # FASE 1: El Archivador Inmortal (Backup del archivo viejo)
        # ---------------------------------------------------------
        if os.path.exists(ruta_archivo_actual):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_historico = f"{nombre_archivo.replace('.xlsx', '')}_v{timestamp}.xlsx"
            
            ruta_historial_dir = os.path.join(HISTORIAL_DIR, nombre_carpeta_grado, nombre_carpeta_grupo)
            os.makedirs(ruta_historial_dir, exist_ok=True)
            
            ruta_historial_final = os.path.join(ruta_historial_dir, nombre_historico)
            
            # Copiar en lugar de mover para asegurar que la descarga/sobrescritura no falle
            shutil.copy2(ruta_archivo_actual, ruta_historial_final)

        # ---------------------------------------------------------
        # FASE 2: Guardar el nuevo Excel maestro localmente
        # ---------------------------------------------------------
        archivo.save(ruta_archivo_actual)

        # ---------------------------------------------------------
        # FASE 3: Sincronización Inversa (Excel -> Base de Datos)
        # ---------------------------------------------------------
        notas_procesadas = 0
        fallas_leidas = 0
        try:
            wb = openpyxl.load_workbook(ruta_archivo_actual, data_only=True)
            ws = wb.active

            header_row = 6
            start_row = 8
            col_nota = 11
            col_fallas = 12

            # Intentar detectar encabezados reales (si la plantilla movió columnas por expansión)
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col_idx).value
                if not isinstance(val, str):
                    continue
                txt = val.strip().upper()
                if 'NOTA' in txt and 'FINAL' in txt:
                    col_nota = col_idx
                elif 'FALLA' in txt:
                    col_fallas = col_idx

            actividades_db = _obtener_actividades_periodo(cursor, int(grupo_id), int(materia_id), int(periodo_id))
            conn.commit()

            # Mapeo dinámico: asociamos las columnas a actividades existentes o creamos nuevas
            actividades_map = []
            max_actividad_col = col_nota if col_nota else ws.max_column + 1
            
            for idx, col_idx in enumerate(range(4, max_actividad_col)):
                if idx < len(actividades_db):
                    # Actividad previamente existente en BD
                    actividades_map.append({
                        'col_idx': col_idx,
                        'id_actividad': actividades_db[idx]['id_actividad']
                    })
                else:
                    # Posible NUEVA actividad añadida por el docente en el Excel
                    celda_header = ws.cell(row=header_row, column=col_idx).value
                    nombre_posible_actividad = str(celda_header).strip() if celda_header else ""
                    
                    if not nombre_posible_actividad or nombre_posible_actividad.upper() == 'NONE':
                        continue
                        
                    # Validar si al menos se ingresó alguna nota para asegurar que no sea columna vacía
                    tiene_notas = False
                    for r in range(start_row, ws.max_row + 1):
                        v_test = ws.cell(row=r, column=col_idx).value
                        if v_test is not None and str(v_test).strip() != "":
                            tiene_notas = True
                            break
                    
                    if tiene_notas:
                        # Crear actividad 'On the Fly' si cuenta con notas reales
                        id_usuario = _resolver_id_usuario_para_actividad(cursor, grupo_id, materia_id)
                        cursor.execute("""
                            INSERT INTO actividades (id_usuario, id_grupo, id_materia, nombre_actividad, descripcion, tipo_actividad, puntaje_maximo, ponderacion)
                            VALUES (%s, %s, %s, %s, 'Creada desde archivo Excel editado', 'Tarea', 5.0, 0)
                        """, (id_usuario, grupo_id, materia_id, nombre_posible_actividad[:100]))
                        nuevo_id = cursor.lastrowid

                        _asegurar_tabla_actividades_periodo(cursor)
                        cursor.execute(
                            """
                            INSERT IGNORE INTO actividades_periodo (id_actividad, id_periodo)
                            VALUES (%s, %s)
                            """,
                            (nuevo_id, int(periodo_id))
                        )
                        conn.commit()
                        
                        actividades_map.append({
                            'col_idx': col_idx,
                            'id_actividad': nuevo_id
                        })

            for row_idx in range(start_row, ws.max_row + 1):
                est_id = _to_int(ws.cell(row=row_idx, column=1).value)
                nombre = ws.cell(row=row_idx, column=2).value
                if not est_id and not nombre:
                    continue
                if not est_id:
                    continue

                for act in actividades_map:
                    val = ws.cell(row=row_idx, column=act['col_idx']).value
                    nota_val = _to_float(val)
                    _upsert_nota(
                        cursor,
                        int(est_id),
                        int(act['id_actividad']),
                        int(materia_id),
                        int(periodo_id),
                        nota_val
                    )
                    if nota_val is not None:
                        notas_procesadas += 1

                # Mantener separado: Fallas. No lo ignoramos.
                if col_fallas:
                    fallas_val = _to_int(ws.cell(row=row_idx, column=col_fallas).value, 0)
                    if fallas_val is not None:
                        fallas_leidas += 1

            conn.commit()

        except Exception as db_err:
            print(f"Error procesando el archivo por layout D..I/K/L: {db_err}")

        # ---------------------------------------------------------
        # FASE 4: Respaldar Simultáneamente en Google Drive
        # ---------------------------------------------------------
        # === DESHABILITADO PARA TECNÓLOGO ===
        # try:
        #     from routes.servicio_drive import get_service, setup_folder_structure
        #     from googleapiclient.http import MediaFileUpload
        #     
        #     # Intentar obtener el user_id de la sesión (o un fallback a 1 para pruebas locales del admin)
        #     user_id = session.get('user_id') or 1
        #     service = get_service(user_id)
        #     
        #     # Preparar carpeta en Drive -> 'Docstry' > Año > Periodo X > Calificaciones
        #     año_actual = datetime.now().year
        #     subcarpetas = setup_folder_structure(user_id, año_actual, f"Periodo {periodo_id}")
        #     carpeta_calificaciones_id = subcarpetas.get('Calificaciones')
        #     
        #     # Buscar si el archivo ya existe para sobreescribirlo sin duplicar
        #     query = f"name='{nombre_archivo}' and '{carpeta_calificaciones_id}' in parents and trashed=false"
        #     results = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
        #     items = results.get('files', [])
        #     
        #     media = MediaFileUpload(ruta_archivo_actual, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        #     
        #     if items:
        #         # Actualiza el archivo en Drive
        #         service.files().update(fileId=items[0]['id'], media_body=media).execute()
        #     else:
        #         # Crea uno nuevo en Drive
        #         file_metadata = {
        #             'name': nombre_archivo,
        #             'parents': [carpeta_calificaciones_id]
        #         }
        #         service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        # except Exception as drive_err:
        #     print(f"Advertencia: No se pudo respaldar en Drive: {drive_err}")
        # ================================================================

        cursor.close()
        conn.close()

        return jsonify({
            'status': 'ok',
            'message': 'Planilla subida exitosamente y respaldada en Drive.',
            'notas_procesadas': notas_procesadas,
            'fallas_leidas': fallas_leidas,
            'archivo_viejo_respaldado': os.path.exists(ruta_archivo_actual) # Confirmación de que se hizo backup si existía
        }), 200

    except Exception as e:
        return _error_interno(e)


# ── 4. TABLA EN LÍNEA (DINÁMICA) ──────────────────────────────────────────────

@calificaciones_bp.route('/api/calificaciones/filtros', methods=['GET'])
def api_calificaciones_filtros():
    """Datos base para filtros de la tabla de registro de notas."""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT id_grado, numero_grado, nombre_grado FROM grados ORDER BY numero_grado")
        grados = cursor.fetchall() or []

        cursor.execute("""
            SELECT id_periodo, numero_periodo, nombre_periodo, estado
            FROM periodos
            ORDER BY numero_periodo
        """)
        periodos = cursor.fetchall() or []

        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'grados': grados, 'periodos': periodos}), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/grupos/<int:grado_id>', methods=['GET'])
def api_calificaciones_grupos(grado_id):
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT id_grupo, codigo_grupo FROM grupos WHERE id_grado = %s ORDER BY codigo_grupo",
            (grado_id,)
        )
        grupos = cursor.fetchall() or []
        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'grupos': grupos}), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/materias/<int:grupo_id>', methods=['GET'])
def api_calificaciones_materias(grupo_id):
    """Materias preferiblemente asignadas al grupo; fallback a catálogo general."""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT DISTINCT m.id_materia, m.nombre_materia, m.codigo_materia
            FROM asignaciones_docente a
            JOIN materias m ON m.id_materia = a.id_materia
            WHERE a.id_grupo = %s AND a.estado = 'Activa'
            ORDER BY m.nombre_materia
            """,
            (grupo_id,)
        )
        materias = cursor.fetchall() or []

        if not materias:
            cursor.execute(
                """
                SELECT m.id_materia, m.nombre_materia, m.codigo_materia
                FROM materias m
                ORDER BY m.nombre_materia
                """
            )
            materias = cursor.fetchall() or []

        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'materias': materias}), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/tabla', methods=['GET'])
def api_calificaciones_tabla():
    """Retorna estudiantes, actividades y notas para renderizar tabla tipo Excel."""
    grupo_id = _to_int(request.args.get('grupo_id'))
    materia_id = _to_int(request.args.get('materia_id'))
    periodo_id = _to_int(request.args.get('periodo_id'))

    if not grupo_id or not materia_id or not periodo_id:
        return jsonify({'error': 'Parámetros incompletos (grupo_id, materia_id, periodo_id)'}), 400

    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT id_estudiante, nombre, apellido, documento
            FROM estudiantes
            WHERE id_grupo = %s AND estado = 'Activo'
            ORDER BY apellido, nombre
            """,
            (grupo_id,)
        )
        estudiantes = cursor.fetchall() or []

        actividades = _obtener_actividades_periodo(cursor, grupo_id, materia_id, periodo_id)
        conn.commit()
        for a in actividades:
            a['puntaje_maximo'] = float(a.get('puntaje_maximo') or 0)
            a['ponderacion'] = float(a.get('ponderacion') or 0)
            a['fecha_actividad'] = _to_iso_date_text(a.get('fecha_vencimiento'), '')

        cursor.execute(
            """
            SELECT n.id_estudiante, n.id_actividad, n.puntaje_obtenido
            FROM notas n
            JOIN actividades a ON a.id_actividad = n.id_actividad
            WHERE a.id_grupo = %s AND a.id_materia = %s AND n.id_periodo = %s
            """,
            (grupo_id, materia_id, periodo_id)
        )
        notas = cursor.fetchall() or []
        for n in notas:
            n['puntaje_obtenido'] = float(n.get('puntaje_obtenido')) if n.get('puntaje_obtenido') is not None else None

        total_ponderacion = sum(float(a.get('ponderacion') or 0) for a in actividades)

        cursor.close()
        conn.close()

        return jsonify({
            'status': 'ok',
            'estudiantes': estudiantes,
            'actividades': actividades,
            'notas': notas,
            'total_ponderacion': total_ponderacion
        }), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/actividades', methods=['POST'])
def api_calificaciones_actividad_crear():
    data = request.get_json() or {}
    try:
        nombre = (data.get('nombre_actividad') or '').strip()
        tipo = (data.get('tipo_actividad') or 'Tarea').strip()
        ponderacion = _to_float(data.get('ponderacion'))
        puntaje_maximo = _to_float(data.get('puntaje_maximo'))
        id_grupo = _to_int(data.get('id_grupo'))
        id_materia = _to_int(data.get('id_materia'))
        id_periodo = _to_int(data.get('id_periodo'))
        descripcion = (data.get('descripcion') or '').strip() or None
        fecha_actividad_txt = _to_iso_date_text(data.get('fecha_actividad'), '')
        fecha_vencimiento = f"{fecha_actividad_txt} 00:00:00" if fecha_actividad_txt else None

        if not nombre or not id_grupo or not id_materia or not id_periodo:
            return jsonify({'error': 'Faltan datos para crear actividad'}), 400
        if ponderacion is None or ponderacion <= 0:
            return jsonify({'error': 'La ponderación debe ser mayor que cero'}), 400
        if puntaje_maximo is None or puntaje_maximo <= 0:
            return jsonify({'error': 'El puntaje máximo debe ser mayor que cero'}), 400

        tipos_validos = {'Tarea', 'Quiz', 'Proyecto', 'Evaluación', 'Clase'}
        if tipo not in tipos_validos:
            tipo = 'Tarea'

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, id_periodo)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo

        id_usuario = _resolver_id_usuario_para_actividad(cursor, id_grupo, id_materia)
        if not id_usuario:
            cursor.close()
            conn.close()
            return jsonify({'error': 'No se pudo resolver docente para la actividad'}), 400

        cursor.execute(
            """
            INSERT INTO actividades (id_usuario, id_grupo, id_materia, nombre_actividad, descripcion, tipo_actividad, fecha_vencimiento, puntaje_maximo, ponderacion)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (id_usuario, id_grupo, id_materia, nombre, descripcion, tipo, fecha_vencimiento, puntaje_maximo, ponderacion)
        )
        id_nuevo = cursor.lastrowid

        _asegurar_tabla_actividades_periodo(cursor)
        cursor.execute(
            """
            INSERT IGNORE INTO actividades_periodo (id_actividad, id_periodo)
            VALUES (%s, %s)
            """,
            (id_nuevo, id_periodo)
        )
        conn.commit()

        cursor.execute(
            """
            SELECT id_actividad, nombre_actividad, tipo_actividad, puntaje_maximo, ponderacion, fecha_vencimiento
            FROM actividades WHERE id_actividad = %s
            """,
            (id_nuevo,)
        )
        actividad = cursor.fetchone()
        if actividad:
            actividad['puntaje_maximo'] = float(actividad.get('puntaje_maximo') or 0)
            actividad['ponderacion'] = float(actividad.get('ponderacion') or 0)
            actividad['fecha_actividad'] = _to_iso_date_text(actividad.get('fecha_vencimiento'), '')

        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'actividad': actividad}), 201
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/actividades/<int:id_actividad>', methods=['PUT'])

def api_calificaciones_actividad_actualizar(id_actividad):
    data = request.get_json() or {}
    try:
        nombre = (data.get('nombre_actividad') or '').strip()
        tipo = (data.get('tipo_actividad') or 'Tarea').strip()
        ponderacion = _to_float(data.get('ponderacion'))
        puntaje_maximo = _to_float(data.get('puntaje_maximo'))
        id_periodo = _to_int(data.get('id_periodo'))
        descripcion = (data.get('descripcion') or '').strip() or None
        fecha_actividad_txt = _to_iso_date_text(data.get('fecha_actividad'), '')
        fecha_vencimiento = f"{fecha_actividad_txt} 00:00:00" if fecha_actividad_txt else None

        if not nombre:
            return jsonify({'error': 'El nombre de actividad es obligatorio'}), 400
        if not id_periodo:
            return jsonify({'error': 'Falta id_periodo'}), 400
        if ponderacion is None or ponderacion <= 0:
            return jsonify({'error': 'La ponderación debe ser mayor que cero'}), 400
        if puntaje_maximo is None or puntaje_maximo <= 0:
            return jsonify({'error': 'El puntaje máximo debe ser mayor que cero'}), 400

        tipos_validos = {'Tarea', 'Quiz', 'Proyecto', 'Evaluación', 'Clase'}
        if tipo not in tipos_validos:
            tipo = 'Tarea'

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, id_periodo)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo

        _asegurar_tabla_actividades_periodo(cursor)
        cursor.execute(
            """
            SELECT a.id_actividad
            FROM actividades a
            JOIN actividades_periodo ap ON ap.id_actividad = a.id_actividad
            WHERE a.id_actividad = %s AND ap.id_periodo = %s
            LIMIT 1
            """,
            (id_actividad, id_periodo)
        )
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'error': 'Actividad no encontrada para el período seleccionado'}), 404

        cursor.execute(
            """
            UPDATE actividades
            SET nombre_actividad = %s,
                descripcion = %s,
                tipo_actividad = %s,
                fecha_vencimiento = %s,
                puntaje_maximo = %s,
                ponderacion = %s
            WHERE id_actividad = %s
            """,
            (nombre, descripcion, tipo, fecha_vencimiento, puntaje_maximo, ponderacion, id_actividad)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/actividades/<int:id_actividad>', methods=['DELETE'])
def api_calificaciones_actividad_eliminar(id_actividad):
    try:
        id_periodo = _to_int(request.args.get('id_periodo'))
        if not id_periodo:
            return jsonify({'error': 'Falta id_periodo'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, id_periodo)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo

        _asegurar_tabla_actividades_periodo(cursor)
        cursor.execute(
            """
            SELECT a.id_actividad
            FROM actividades a
            JOIN actividades_periodo ap ON ap.id_actividad = a.id_actividad
            WHERE a.id_actividad = %s AND ap.id_periodo = %s
            LIMIT 1
            """,
            (id_actividad, id_periodo)
        )
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'error': 'Actividad no encontrada para el período seleccionado'}), 404

        cursor.execute("DELETE FROM notas WHERE id_actividad = %s AND id_periodo = %s", (id_actividad, id_periodo))
        notas_eliminadas = cursor.rowcount or 0

        cursor.execute(
            "DELETE FROM actividades_periodo WHERE id_actividad = %s AND id_periodo = %s",
            (id_actividad, id_periodo)
        )

        cursor.execute("SELECT COUNT(*) AS total FROM actividades_periodo WHERE id_actividad = %s", (id_actividad,))
        restante = cursor.fetchone() or {}
        if int(restante.get('total') or 0) == 0:
            cursor.execute("DELETE FROM actividades WHERE id_actividad = %s", (id_actividad,))

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'notas_eliminadas': notas_eliminadas}), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/notas', methods=['POST'])
def api_calificaciones_nota_guardar():
    data = request.get_json() or {}
    try:
        id_estudiante = _to_int(data.get('id_estudiante'))
        id_actividad = _to_int(data.get('id_actividad'))
        id_materia = _to_int(data.get('id_materia'))
        id_periodo = _to_int(data.get('id_periodo'))
        puntaje = _to_float(data.get('puntaje_obtenido'))

        if not id_estudiante or not id_actividad or not id_materia or not id_periodo:
            return jsonify({'error': 'Datos incompletos para guardar nota'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, id_periodo)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo

        _asegurar_tabla_actividades_periodo(cursor)
        cursor.execute(
            """
            SELECT 1
            FROM actividades_periodo
            WHERE id_actividad = %s AND id_periodo = %s
            LIMIT 1
            """,
            (id_actividad, id_periodo)
        )
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'error': 'La actividad no pertenece al período seleccionado'}), 400

        _upsert_nota(cursor, id_estudiante, id_actividad, id_materia, id_periodo, puntaje)
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/notas/lote', methods=['POST'])
def api_calificaciones_notas_lote():
    data = request.get_json() or {}
    notas = data.get('notas') or []
    id_materia = _to_int(data.get('id_materia'))
    id_periodo = _to_int(data.get('id_periodo'))

    if not id_materia or not id_periodo:
        return jsonify({'error': 'Faltan id_materia o id_periodo'}), 400

    if not isinstance(notas, list):
        return jsonify({'error': 'El campo notas debe ser una lista'}), 400

    guardadas = 0
    errores = []

    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, id_periodo)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo

        _asegurar_tabla_actividades_periodo(cursor)

        actividades_cache = {}

        for idx, n in enumerate(notas):
            try:
                id_estudiante = _to_int(n.get('id_estudiante'))
                id_actividad = _to_int(n.get('id_actividad'))
                puntaje = _to_float(n.get('puntaje_obtenido'))

                if not id_estudiante or not id_actividad:
                    raise ValueError('id_estudiante o id_actividad inválidos')

                if id_actividad not in actividades_cache:
                    cursor.execute(
                        """
                        SELECT 1
                        FROM actividades_periodo
                        WHERE id_actividad = %s AND id_periodo = %s
                        LIMIT 1
                        """,
                        (id_actividad, id_periodo)
                    )
                    actividades_cache[id_actividad] = bool(cursor.fetchone())

                if not actividades_cache[id_actividad]:
                    raise ValueError('actividad fuera del período seleccionado')

                _upsert_nota(cursor, id_estudiante, id_actividad, id_materia, id_periodo, puntaje)
                guardadas += 1
            except Exception as inner:
                errores.append({'index': idx, 'error': str(inner)})

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'guardadas': guardadas, 'errores': errores}), 200
    except Exception as e:
        return _error_interno(e)
    
@calificaciones_bp.route('/api/calificaciones/descargar/institucional', methods=['GET'])
def descargar_planilla_institucional():
    """Genera la planilla usando la plantilla Excel con diseño institucional."""
    grupo_id = request.args.get('grupo_id')
    materia_id = request.args.get('materia_id')
    periodo_id = request.args.get('periodo_id')

    if not all([grupo_id, materia_id, periodo_id]):
        return jsonify({'error': 'Faltan parámetros'}), 400

    try:
        # 1. Cargar la plantilla profesional que ya detectamos
        wb = openpyxl.load_workbook(PLANTILLA_PATH)
        ws = wb.active

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        # 2. Obtener datos para los placeholders {{...}}
        cursor.execute("""
            SELECT g.codigo_grupo, m.nombre_materia, p.numero_periodo 
            FROM grupos g, materias m, periodos p 
            WHERE g.id_grupo = %s AND m.id_materia = %s AND p.id_periodo = %s
        """, (grupo_id, materia_id, periodo_id))
        info = cursor.fetchone()

        # 3. Reemplazo de placeholders (Basado en tu diseño)
        mapeo = {
            "{{INSTITUCION}}": "DOCSTRY ACADEMY", # Aquí puedes traer el nombre real de la DB
            "{{GRADO}}": info['codigo_grupo'],
            "{{MATERIA}}": info['nombre_materia'],
            "{{PERIODO}}": f"{info['numero_periodo']}"
        }

        for row in ws.iter_rows():
            for cell in row:
                if cell.value in mapeo:
                    cell.value = mapeo[cell.value]

        # 4. Llenar Estudiantes y Fallas (Empezando en Fila 7 como vimos en tu imagen)
        cursor.execute("""
            SELECT e.id_estudiante, e.apellido, e.nombre,
                   (SELECT total_ausencias FROM asistencias_por_periodo 
                    WHERE id_estudiante = e.id_estudiante AND id_materia = %s AND id_periodo = %s) as fallas
            FROM estudiantes e 
            WHERE e.id_grupo = %s AND e.estado = 'Activo'
            ORDER BY e.apellido, e.nombre
        """, (materia_id, periodo_id, grupo_id))
        
        estudiantes = cursor.fetchall()
        fila_inicio = 7
        for i, est in enumerate(estudiantes):
            fila = fila_inicio + i
            ws.cell(row=fila, column=1, value=i+1) # COD
            ws.cell(row=fila, column=2, value=f"{est['apellido']} {est['nombre']}") # ESTUDIANTES
            ws.cell(row=fila, column=13, value=est['fallas'] or 0) # Columna de FALLAS

        # 5. Guardar temporalmente y enviar
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            ruta_tmp = tmp.name

        cursor.close()
        conn.close()

        nombre_descarga = f"Planilla_{info['codigo_grupo']}_{info['nombre_materia']}.xlsx"
        return send_file(ruta_tmp, as_attachment=True, download_name=nombre_descarga)

    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/acuerdo-pedagogico', methods=['GET'])
def api_acuerdo_pedagogico_info():
    grupo_id = _to_int(request.args.get('grupo_id'))
    materia_id = _to_int(request.args.get('materia_id'))
    periodo_id = _to_int(request.args.get('periodo_id'))

    if not grupo_id or not materia_id or not periodo_id:
        return jsonify({'error': 'Faltan parámetros (grupo_id, materia_id, periodo_id)'}), 400

    ruta = _acuerdo_pdf_path(grupo_id, materia_id, periodo_id)
    existe = os.path.exists(ruta)
    return jsonify({
        'status': 'ok',
        'existe': existe,
        'filename': os.path.basename(ruta) if existe else None,
        'ver_url': (
            f"/api/calificaciones/acuerdo-pedagogico/ver?grupo_id={grupo_id}&materia_id={materia_id}&periodo_id={periodo_id}"
            if existe else None
        )
    }), 200


@calificaciones_bp.route('/api/calificaciones/acuerdo-pedagogico', methods=['POST'])
def api_acuerdo_pedagogico_subir():
    grupo_id = _to_int(request.form.get('grupo_id'))
    materia_id = _to_int(request.form.get('materia_id'))
    periodo_id = _to_int(request.form.get('periodo_id'))

    if not grupo_id or not materia_id or not periodo_id:
        return jsonify({'error': 'Faltan parámetros (grupo_id, materia_id, periodo_id)'}), 400

    if 'archivo_pdf' not in request.files:
        return jsonify({'error': 'No se envió archivo_pdf'}), 400

    archivo = request.files['archivo_pdf']
    if not archivo or not archivo.filename:
        return jsonify({'error': 'Archivo inválido'}), 400

    nombre = str(archivo.filename).strip().lower()
    if not nombre.endswith('.pdf'):
        return jsonify({'error': 'Solo se permiten archivos PDF'}), 400

    conn = None
    cursor = None
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, periodo_id)
        if bloqueo:
            return bloqueo

        ruta = _acuerdo_pdf_path(grupo_id, materia_id, periodo_id)
        archivo.save(ruta)

        return jsonify({
            'status': 'ok',
            'message': 'Acuerdo pedagógico guardado correctamente',
            'filename': os.path.basename(ruta),
            'ver_url': f"/api/calificaciones/acuerdo-pedagogico/ver?grupo_id={grupo_id}&materia_id={materia_id}&periodo_id={periodo_id}"
        }), 200
    except Exception as e:
        return _error_interno(e)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@calificaciones_bp.route('/api/calificaciones/acuerdo-pedagogico/ver', methods=['GET'])
def api_acuerdo_pedagogico_ver():
    grupo_id = _to_int(request.args.get('grupo_id'))
    materia_id = _to_int(request.args.get('materia_id'))
    periodo_id = _to_int(request.args.get('periodo_id'))

    if not grupo_id or not materia_id or not periodo_id:
        return jsonify({'error': 'Faltan parámetros (grupo_id, materia_id, periodo_id)'}), 400

    ruta = _acuerdo_pdf_path(grupo_id, materia_id, periodo_id)
    if not os.path.exists(ruta):
        return jsonify({'error': 'No hay acuerdo pedagógico cargado para este período'}), 404

    return send_file(
        ruta,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=os.path.basename(ruta)
    )

# ═══════════════════════════════════════════════════════════════════════════════
# NUEVAS RUTAS: GESTIÓN DE ESTRUCTURA DE PERIODOS
# ═══════════════════════════════════════════════════════════════════════════════

@calificaciones_bp.route('/api/calificaciones/periodos/estructura', methods=['GET'])
def api_periodos_estructura():
    """
    Obtiene la estructura de carpetas de periodos (años, períodos, módulos).
    Devuelve un árbol JSON con:
    - Años disponibles (2026, 2027, etc.)
    - Períodos dentro de cada año (1, 2, 3, 4)
    - Módulos dentro de cada período (Calificaciones, Asistencias, Reportes)
    - Archivos Excel dentro de cada módulo
    """
    try:
        estructura = {'status': 'ok', 'años': {}}
        
        if not os.path.exists(PLANILLAS_DIR):
            return jsonify(estructura), 200
        
        # Iterar a través de años
        for año_carpeta in sorted(os.listdir(PLANILLAS_DIR)):
            año_path = os.path.join(PLANILLAS_DIR, año_carpeta)
            if not os.path.isdir(año_path):
                continue
            
            # Extraer número del año (ej: "Año_2026" → "2026")
            try:
                año_num = año_carpeta.split('_')[1]
            except:
                continue
            
            estructura['años'][año_num] = {'periodos': {}}
            
            # Iterar a través de períodos
            for periodo_carpeta in sorted(os.listdir(año_path)):
                periodo_path = os.path.join(año_path, periodo_carpeta)
                if not os.path.isdir(periodo_path):
                    continue
                
                try:
                    periodo_num = periodo_carpeta.split('_')[1]
                except:
                    continue
                
                estructura['años'][año_num]['periodos'][periodo_num] = {'modulos': {}}
                
                # Iterar a través de módulos
                for modulo_carpeta in sorted(os.listdir(periodo_path)):
                    modulo_path = os.path.join(periodo_path, modulo_carpeta)
                    if not os.path.isdir(modulo_path):
                        continue
                    
                    # Listar archivos en el módulo
                    archivos = [f for f in os.listdir(modulo_path) if f.endswith('.xlsx')]
                    estructura['años'][año_num]['periodos'][periodo_num]['modulos'][modulo_carpeta] = {
                        'archivos': sorted(archivos),
                        'cantidad': len(archivos)
                    }
        
        return jsonify(estructura), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/periodos/listar', methods=['GET'])
def api_periodos_listar():
    """
    Lista todos los periodos con sus información (año, número, nombre, rango).
    Útil para mostrar en dropdowns y selectores de interfaz.
    """
    try:
        periodos_list = []
        
        if not os.path.exists(PLANILLAS_DIR):
            return jsonify({'status': 'ok', 'periodos': []}), 200
        
        for año_carpeta in sorted(os.listdir(PLANILLAS_DIR)):
            año_path = os.path.join(PLANILLAS_DIR, año_carpeta)
            if not os.path.isdir(año_path):
                continue
            
            try:
                año_num = int(año_carpeta.split('_')[1])
            except:
                continue
            
            for periodo_carpeta in sorted(os.listdir(año_path)):
                periodo_path = os.path.join(año_path, periodo_carpeta)
                if not os.path.isdir(periodo_path):
                    continue
                
                try:
                    periodo_num = int(periodo_carpeta.split('_')[1])
                    periodos_list.append({
                        'año': año_num,
                        'periodo': periodo_num,
                        'nombre': f'Período {periodo_num} - {año_num}',
                        'ruta': os.path.relpath(periodo_path, PLANILLAS_DIR)
                    })
                except:
                    continue
        
        return jsonify({'status': 'ok', 'periodos': periodos_list}), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/periodos/descargar/<path:ruta_archivo>', methods=['GET'])
def api_periodos_descargar(ruta_archivo):
    """
    Descarga un archivo Excel de la estructura de periodos.
    
    Parámetro: ruta_archivo (relativa desde PLANILLAS_DIR)
    Ejemplo: Año_2026/Periodo_1/Calificaciones/Calificaciones_2026_P1.xlsx
    """
    try:
        # Validar que la ruta esté dentro de PLANILLAS_DIR (seguridad)
        ruta_completa = os.path.normpath(os.path.join(PLANILLAS_DIR, ruta_archivo))
        
        # Verificar que sale de PLANILLAS_DIR
        if not ruta_completa.startswith(os.path.normpath(PLANILLAS_DIR)):
            return jsonify({'error': 'Acceso denegado: ruta inválida'}), 403
        
        if not os.path.exists(ruta_completa):
            return jsonify({'error': 'Archivo no encontrado'}), 404
        
        nombre_archivo = os.path.basename(ruta_completa)
        
        return send_file(
            ruta_completa,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/periodos/info', methods=['GET'])
def api_periodos_info():
    """
    Devuelve información general de la estructura de periodos:
    - Años disponibles
    - Total de períodos
    - Total de módulos
    - Total de archivos
    """
    try:
        info = {
            'status': 'ok',
            'directorio': PLANILLAS_DIR,
            'existe': os.path.exists(PLANILLAS_DIR),
            'años_total': 0,
            'periodos_total': 0,
            'modulos_total': 0,
            'archivos_total': 0,
            'años': []
        }
        
        if not os.path.exists(PLANILLAS_DIR):
            return jsonify(info), 200
        
        for año_carpeta in os.listdir(PLANILLAS_DIR):
            año_path = os.path.join(PLANILLAS_DIR, año_carpeta)
            if os.path.isdir(año_path):
                try:
                    año_num = int(año_carpeta.split('_')[1])
                    info['años'].append(año_num)
                    info['años_total'] += 1
                    
                    for periodo_carpeta in os.listdir(año_path):
                        periodo_path = os.path.join(año_path, periodo_carpeta)
                        if os.path.isdir(periodo_path):
                            info['periodos_total'] += 1
                            
                            for modulo_carpeta in os.listdir(periodo_path):
                                modulo_path = os.path.join(periodo_path, modulo_carpeta)
                                if os.path.isdir(modulo_path):
                                    info['modulos_total'] += 1
                                    archivos = [f for f in os.listdir(modulo_path) if f.endswith('.xlsx')]
                                    info['archivos_total'] += len(archivos)
                except:
                    pass
        
        info['años'] = sorted(info['años'])
        return jsonify(info), 200
    except Exception as e:
        return _error_interno(e)


# ════════════════════════════════════════════════════════════════════════════════
# FASE 2: ENDPOINTS DE DESCARGA POR PERÍODO
# ════════════════════════════════════════════════════════════════════════════════

@calificaciones_bp.route('/api/calificaciones/periodos/<int:periodo_id>/listar', methods=['GET'])
def api_listar_periodo(periodo_id):
    """Lista todos los archivos Excel de un período específico."""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # Validar que el período existe
        cursor.execute("SELECT numero_periodo FROM periodos WHERE id_periodo = %s", (periodo_id,))
        periodo = cursor.fetchone()
        if not periodo:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Período no encontrado'}), 404
        
        cursor.close()
        conn.close()
        
        PERIODOS_DIR = os.path.join(ESCRITORIO, 'Periodos_DocstrY')
        periodo_folder = os.path.join(PERIODOS_DIR, f"Período_{periodo['numero_periodo']}")
        
        if not os.path.exists(periodo_folder):
            return jsonify({'error': 'Carpeta del período no existe. Sincroniza primero.'}), 404
        
        # Construir lista de archivos
        archivos = []
        for root, dirs, files in os.walk(periodo_folder):
            for file in files:
                if file.endswith('.xlsx'):
                    ruta_completa = os.path.join(root, file)
                    ruta_relativa = os.path.relpath(ruta_completa, PERIODOS_DIR)
                    tamaño = os.path.getsize(ruta_completa)
                    archivos.append({
                        'nombre': file,
                        'ruta_relativa': ruta_relativa,
                        'tamaño_kb': round(tamaño / 1024, 2),
                        'modificado': str(os.path.getmtime(ruta_completa))
                    })
        
        return jsonify({
            'status': 'ok',
            'periodo_id': periodo_id,
            'numero_periodo': periodo['numero_periodo'],
            'total_archivos': len(archivos),
            'archivos': sorted(archivos, key=lambda x: x['nombre'])
        }), 200
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/periodos/descargar_archivo', methods=['GET'])
def api_descargar_archivo_periodo():
    """
    Descarga un archivo Excel individual de la estructura de períodos.
    Parámetros: periodo_id, ruta_relativa (desde Periodos_DocstrY)
    """
    try:
        from urllib.parse import unquote
        
        periodo_id = request.args.get('periodo_id', type=int)
        ruta_relativa = unquote(request.args.get('ruta', ''))
        
        if not periodo_id or not ruta_relativa:
            return jsonify({'error': 'Faltan parámetros: periodo_id y ruta'}), 400
        
        PERIODOS_DIR = os.path.join(ESCRITORIO, 'Periodos_DocstrY')
        
        # Validar que la ruta esté dentro de PERIODOS_DIR (seguridad)
        ruta_completa = os.path.normpath(os.path.join(PERIODOS_DIR, ruta_relativa))
        
        if not ruta_completa.startswith(os.path.normpath(PERIODOS_DIR)):
            return jsonify({'error': 'Acceso denegado: ruta inválida'}), 403
        
        if not os.path.exists(ruta_completa):
            return jsonify({'error': 'Archivo no encontrado'}), 404
        
        nombre_archivo = os.path.basename(ruta_completa)
        
        return send_file(
            ruta_completa,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
    except Exception as e:
        return _error_interno(e)


@calificaciones_bp.route('/api/calificaciones/periodos/<int:periodo_id>/descargar_zip', methods=['GET'])
def api_descargar_zip_periodo(periodo_id):
    """Descarga todos los archivos Excel de un período comprimidos en ZIP."""
    try:
        import zipfile
        from io import BytesIO
        from datetime import datetime
        
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # Validar que el período existe
        cursor.execute("SELECT numero_periodo FROM periodos WHERE id_periodo = %s", (periodo_id,))
        periodo = cursor.fetchone()
        if not periodo:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Período no encontrado'}), 404
        
        cursor.close()
        conn.close()
        
        PERIODOS_DIR = os.path.join(ESCRITORIO, 'Periodos_DocstrY')
        periodo_folder = os.path.join(PERIODOS_DIR, f"Período_{periodo['numero_periodo']}")
        
        if not os.path.exists(periodo_folder):
            return jsonify({'error': 'Carpeta del período no existe'}), 404
        
        # Crear ZIP en memoria
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(periodo_folder):
                for file in files:
                    if file.endswith('.xlsx'):
                        ruta_archivo = os.path.join(root, file)
                        # Ruta relativa dentro del ZIP (mantener estructura)
                        arcname = os.path.relpath(ruta_archivo, PERIODOS_DIR)
                        zip_file.write(ruta_archivo, arcname=arcname)
        
        zip_buffer.seek(0)
        fecha_hoy = datetime.now().strftime('%Y%m%d')
        nombre_zip = f"Calificaciones_Periodo_{periodo['numero_periodo']}_{fecha_hoy}.zip"
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=nombre_zip
        )
    except Exception as e:
        return _error_interno(e)
