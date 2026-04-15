# ════════════════════════════════════════════════════════════════════════════════
# BLUEPRINT: REPORTES — Reportes, correos por grupo, envío de correo, logs
# ════════════════════════════════════════════════════════════════════════════════

from flask import Blueprint, jsonify, request, session, send_file
import mysql.connector
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime
from copy import copy
import io
import json
import os
import shutil
import subprocess

import openpyxl
from openpyxl.utils import get_column_letter
from utils.database import get_db
from utils.helpers import _error_interno, log_action

reportes_bp = Blueprint('reportes', __name__)

# ── Logs ─────────────────────────────────────────────────────────────────────

@reportes_bp.route('/api/logs', methods=['GET'])
def api_logs():
    """Listar logs del sistema"""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                l.id_log,
                l.id_usuario,
                l.tipo_accion,
                l.tabla_afectada,
                l.registro_id,
                l.descripcion,
                l.ip_address,
                l.user_agent,
                l.exito,
                COALESCE(l.timestamp_accion, l.created_at) AS fecha,
                u.nombre as usuario_nombre,
                u.apellido as usuario_apellido
            FROM log_registro l
            LEFT JOIN usuarios u ON l.id_usuario = u.id_usuario
            ORDER BY COALESCE(l.timestamp_accion, l.created_at) DESC
            LIMIT 100
        """)
        logs = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'logs': logs}), 200
    except Exception as e:
        return _error_interno(e)

# ── Obtener grupos para reportes ─────────────────────────────────────────────

@reportes_bp.route('/api/reportes/grupos', methods=['GET'])
def api_obtener_grupos():
    """Obtiene todos los grupos con su información de grado asociado"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'No autenticado'}), 401

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                g.id_grupo,
                g.codigo_grupo AS codigo_grupo,
                gd.nombre_grado AS nombre_grado,
                gd.id_grado
            FROM grupos g
            INNER JOIN grados gd ON g.id_grado = gd.id_grado
            ORDER BY gd.id_grado ASC, g.codigo_grupo ASC
        """)

        grupos = cursor.fetchall()
        cursor.close()
        conn.close()

        grupos_lista = []
        for g in grupos:
            item = {
                'id_grupo': int(g['id_grupo']),
                'codigo_grupo': str(g['codigo_grupo']),
                'nombre_grado': str(g['nombre_grado']),
                'id_grado': int(g['id_grado'])
            }
            grupos_lista.append(item)

        return jsonify({'status': 'ok', 'grupos': grupos_lista}), 200

    except mysql.connector.Error:
        return _error_interno(Exception('Error de base de datos al obtener grupos'))
    except Exception as e:
        return _error_interno(e)

# ── Correos por grupo ────────────────────────────────────────────────────────

@reportes_bp.route('/api/reportes/correos-grupo/<int:id_grupo>', methods=['GET'])
def api_correos_por_grupo(id_grupo):
    """Obtiene todos los correos de estudiantes de un grupo específico"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'No autenticado'}), 401

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT DISTINCT e.correo
            FROM estudiantes e
            WHERE e.id_grupo = %s
            AND e.correo IS NOT NULL
            AND e.correo != ''
            AND e.estado = 'Activo'
        """, (id_grupo,))

        resultados = cursor.fetchall()
        correos = [row['correo'] for row in resultados]
        cursor.close()
        conn.close()

        return jsonify({
            'status': 'ok',
            'id_grupo': id_grupo,
            'cantidad': len(correos),
            'correos': correos
        }), 200

    except mysql.connector.Error:
        return _error_interno(Exception('Error de base de datos al obtener correos'))
    except Exception as e:
        return _error_interno(e)

# ── Envío de correo SMTP ─────────────────────────────────────────────────────

def enviar_correo_smtp(remitente, destinatarios, asunto, mensaje_html):
    """Envía correo usando SMTP (Gmail)."""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', 587))
        email_cuenta = os.getenv('EMAIL_CUENTA', 'tu_email@gmail.com')
        email_password = os.getenv('EMAIL_PASSWORD', 'tu_app_password')

        mensaje = MIMEMultipart('alternative')
        mensaje['Subject'] = asunto
        mensaje['From'] = email_cuenta
        mensaje['To'] = ', '.join(destinatarios)

        html = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f5f5f5; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
                    <h2 style="color: #8B6F47; border-bottom: 2px solid #8B6F47; padding-bottom: 10px;">
                        📄 Reporte de la Institución
                    </h2>
                    <div style="margin: 20px 0;">
                        <p><strong>De:</strong> {remitente}</p>
                        <p><strong>Fecha:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                    </div>
                    <div style="background-color: #f9f9f9; padding: 15px; border-radius: 5px; margin: 20px 0; border-left: 4px solid #8B6F47;">
                        {mensaje_html.replace(chr(10), '<br>')}
                    </div>
                    <hr style="border: none; border-top: 1px solid #ddd; margin: 20px 0;">
                    <p style="color: #999; font-size: 12px; text-align: center;">
                        Este es un correo automático del Sistema de Gestión Académica DOCSTRY
                    </p>
                </div>
            </body>
        </html>
        """

        parte_html = MIMEText(html, 'html')
        mensaje.attach(parte_html)

        with smtplib.SMTP(smtp_server, smtp_port) as servidor:
            servidor.starttls()
            servidor.login(email_cuenta, email_password)
            servidor.sendmail(email_cuenta, destinatarios, mensaje.as_string())

        return True
    except Exception as e:
        print(f"❌ Error enviando correo: {e}")
        return False


def enviar_correo_smtp_con_adjuntos(remitente, destinatarios, asunto, mensaje_html, adjuntos):
    """Envía correo usando SMTP con adjuntos binarios."""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', 587))
        email_cuenta = os.getenv('EMAIL_CUENTA', 'tu_email@gmail.com')
        email_password = os.getenv('EMAIL_PASSWORD', 'tu_app_password')

        mensaje = MIMEMultipart('mixed')
        mensaje['Subject'] = asunto
        mensaje['From'] = email_cuenta
        mensaje['To'] = ', '.join(destinatarios)

        html = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f5f5f5; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
                    <h2 style="color: #8B6F47; border-bottom: 2px solid #8B6F47; padding-bottom: 10px;">
                        📄 Reporte de la Institución
                    </h2>
                    <div style="margin: 20px 0;">
                        <p><strong>De:</strong> {remitente}</p>
                        <p><strong>Fecha:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                    </div>
                    <div style="background-color: #f9f9f9; padding: 15px; border-radius: 5px; margin: 20px 0; border-left: 4px solid #8B6F47;">
                        {mensaje_html.replace(chr(10), '<br>')}
                    </div>
                    <hr style="border: none; border-top: 1px solid #ddd; margin: 20px 0;">
                    <p style="color: #999; font-size: 12px; text-align: center;">
                        Este es un correo automático del Sistema de Gestión Académica DOCSTRY
                    </p>
                </div>
            </body>
        </html>
        """

        mensaje.attach(MIMEText(html, 'html'))

        for adjunto in adjuntos or []:
            nombre = adjunto.get('filename') or 'adjunto.pdf'
            data = adjunto.get('data') or b''
            mimetype = adjunto.get('mimetype') or 'application/octet-stream'
            parte = MIMEApplication(data, Name=nombre)
            parte['Content-Disposition'] = f'attachment; filename="{nombre}"'
            parte.add_header('Content-Type', mimetype)
            mensaje.attach(parte)

        with smtplib.SMTP(smtp_server, smtp_port) as servidor:
            servidor.starttls()
            servidor.login(email_cuenta, email_password)
            servidor.sendmail(email_cuenta, destinatarios, mensaje.as_string())

        return True
    except Exception as e:
        print(f"❌ Error enviando correo con adjuntos: {e}")
        return False


def _obtener_ruta_escritorio():
    usuario_dir = os.path.expanduser('~')
    rutas = [
        os.path.join(usuario_dir, 'OneDrive', 'Escritorio'),
        os.path.join(usuario_dir, 'Desktop'),
        os.path.join(usuario_dir, 'Escritorio')
    ]
    for ruta in rutas:
        if os.path.exists(ruta):
            return ruta
    return usuario_dir


ESCRITORIO = _obtener_ruta_escritorio()
BOLETINES_DIR = os.path.join(ESCRITORIO, 'Boletines_DocstrY')


def _resolver_ruta_plantilla_boletin():
    """Resuelve la ruta de PlantillaBoletin en diferentes layouts."""
    here = os.path.dirname(os.path.abspath(__file__))
    parent = os.path.dirname(here)
    rutas = [
        os.path.join(parent, 'backend', 'templates'),
        os.path.join(parent, 'templates')
    ]
    for base in rutas:
        if not os.path.isdir(base):
            continue
        for nombre in os.listdir(base):
            lower = nombre.lower()
            if lower.startswith('plantillabolet') and lower.endswith('.xlsx'):
                return os.path.join(base, nombre)
    return None


def _email_valido_simple(correo):
    if not correo or '@' not in correo:
        return False
    local, domain = correo.split('@', 1)
    if not local or '.' not in domain:
        return False
    return True


def _asegurar_tabla_boletin_puestos(cursor):
    """Crea la tabla de puestos si no existe (tolerante a fallos)."""
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS boletin_puestos (
                id_boletin_puesto INT NOT NULL AUTO_INCREMENT,
                id_estudiante INT NOT NULL,
                id_grupo INT NOT NULL,
                id_periodo INT NULL,
                es_definitiva TINYINT(1) DEFAULT 0,
                promedio DECIMAL(5,2) NULL,
                total_fallas INT DEFAULT 0,
                puesto INT DEFAULT NULL,
                fecha_generado DATETIME DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (id_boletin_puesto),
                UNIQUE KEY uk_boletin_puesto (id_estudiante, id_grupo, id_periodo, es_definitiva),
                KEY idx_boletin_grupo (id_grupo, id_periodo)
            )
            """
        )
    except Exception:
        pass


def _asegurar_tabla_actividades_periodo(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS actividades_periodo (
            id_actividad_periodo INT NOT NULL AUTO_INCREMENT,
            id_actividad INT NOT NULL,
            id_periodo INT NOT NULL,
            PRIMARY KEY (id_actividad_periodo),
            UNIQUE KEY uk_actividad_periodo (id_actividad, id_periodo),
            KEY idx_ap_periodo (id_periodo),
            KEY idx_ap_actividad (id_actividad)
        )
        """
    )


def _seed_actividades_para_periodo(cursor, grupo_id, materia_id, id_periodo):
    """Inicializa mapeo minimo de actividades para el periodo."""
    _asegurar_tabla_actividades_periodo(cursor)
    cursor.execute(
        """
        SELECT ap.id_actividad
        FROM actividades_periodo ap
        JOIN actividades a ON a.id_actividad = ap.id_actividad
        WHERE ap.id_periodo = %s AND a.id_grupo = %s AND a.id_materia = %s
        LIMIT 1
        """,
        (id_periodo, grupo_id, materia_id)
    )
    if cursor.fetchone():
        return

    if int(id_periodo) != 1:
        return

    cursor.execute(
        """
        SELECT id_actividad
        FROM actividades
        WHERE id_grupo = %s AND id_materia = %s
        """,
        (grupo_id, materia_id)
    )
    actividades = cursor.fetchall() or []
    for act in actividades:
        cursor.execute(
            "INSERT IGNORE INTO actividades_periodo (id_actividad, id_periodo) VALUES (%s, %s)",
            (act['id_actividad'], id_periodo)
        )


def _obtener_actividades_periodo(cursor, grupo_id, materia_id, id_periodo):
    _seed_actividades_para_periodo(cursor, grupo_id, materia_id, id_periodo)
    cursor.execute(
        """
        SELECT a.id_actividad, a.puntaje_maximo, a.ponderacion
        FROM actividades_periodo ap
        JOIN actividades a ON a.id_actividad = ap.id_actividad
        WHERE ap.id_periodo = %s AND a.id_grupo = %s AND a.id_materia = %s
        ORDER BY ap.id_actividad_periodo
        """,
        (id_periodo, grupo_id, materia_id)
    )
    actividades = cursor.fetchall() or []
    for act in actividades:
        act['puntaje_maximo'] = float(act.get('puntaje_maximo') or 0)
        act['ponderacion'] = float(act.get('ponderacion') or 0)
    return actividades


def _ajustar_nota_escala_0_5(valor):
    if valor is None:
        return None
    try:
        nota = float(valor)
    except Exception:
        return None
    if nota <= 0:
        nota = 1.0
    if nota > 5:
        nota = 5.0
    return round(nota, 2)


def _calcular_nota_periodo(cursor, notas_map, estudiante_id, grupo_id, materia_id, periodo_id):
    actividades = _obtener_actividades_periodo(cursor, grupo_id, materia_id, periodo_id)
    if not actividades:
        return None

    total_ponderacion = sum(float(a.get('ponderacion') or 0) for a in actividades)
    usar_default = total_ponderacion <= 0
    ponderacion_default = 100.0 / len(actividades) if actividades else 0

    acumulado = 0.0
    for act in actividades:
        puntaje_maximo = float(act.get('puntaje_maximo') or 0)
        if puntaje_maximo <= 0:
            continue
        peso = ponderacion_default if usar_default else float(act.get('ponderacion') or 0)
        nota = notas_map.get((estudiante_id, act['id_actividad'], periodo_id))
        nota = float(nota) if nota is not None else 0.0
        acumulado += (nota / puntaje_maximo) * peso

    # acumulado queda en escala 0-100, convertir a 0-5
    nota_0_5 = acumulado / 20.0 if acumulado is not None else None
    return _ajustar_nota_escala_0_5(nota_0_5)


def _obtener_fallas_periodo(cursor, estudiante_id, materia_id, periodo_id):
    cursor.execute(
        """
        SELECT COALESCE(total_ausencias, 0) AS fallas
        FROM asistencias_por_periodo
        WHERE id_estudiante = %s AND id_materia = %s AND id_periodo = %s
        LIMIT 1
        """,
        (estudiante_id, materia_id, periodo_id)
    )
    row = cursor.fetchone() or {}
    return int(row.get('fallas') or 0)


def _replace_placeholders_worksheet(ws, mapping):
    merged_ranges = list(ws.merged_cells.ranges)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                nuevo_valor = cell.value
                for token, valor in mapping.items():
                    if token in nuevo_valor:
                        nuevo_valor = nuevo_valor.replace(token, str(valor or ''))
                if nuevo_valor != cell.value:
                    row_idx, col_idx = cell.row, cell.column
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        for merged_range in merged_ranges:
                            if (
                                merged_range.min_row <= row_idx <= merged_range.max_row
                                and merged_range.min_col <= col_idx <= merged_range.max_col
                            ):
                                row_idx, col_idx = merged_range.min_row, merged_range.min_col
                                break
                    ws.cell(row=row_idx, column=col_idx, value=nuevo_valor)


def _safe_set_cell(ws, row, col, value):
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if (
                    merged_range.min_row <= row <= merged_range.max_row
                    and merged_range.min_col <= col <= merged_range.max_col
                ):
                    row, col = merged_range.min_row, merged_range.min_col
                    break
        ws.cell(row=row, column=col, value=value)
    except Exception:
        pass


def _ensure_materia_blocks(ws, total_materias):
    block_start = 9
    block_height = 4

    mat_rows = [
        cell.row
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and '{{MATERIA}}' in cell.value
    ]
    base_blocks = max(1, int(len(set(mat_rows)) / 2))
    if total_materias <= base_blocks:
        return block_start, block_height, base_blocks

    extra = total_materias - base_blocks

    template_merges = [
        rng for rng in ws.merged_cells.ranges
        if rng.min_row >= block_start and rng.max_row <= block_start + block_height - 1
    ]

    insert_at = block_start + base_blocks * block_height
    for _ in range(extra):
        ws.insert_rows(insert_at, amount=block_height)
        for i in range(block_height):
            src_row = block_start + i
            dst_row = insert_at + i
            ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
            for col in range(1, ws.max_column + 1):
                src = ws.cell(row=src_row, column=col)
                dst = ws.cell(row=dst_row, column=col)
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.fill = copy(src.fill)
                    dst.border = copy(src.border)
                    dst.alignment = copy(src.alignment)
                    dst.number_format = src.number_format
                    dst.protection = copy(src.protection)
                dst.value = src.value

        offset = insert_at - block_start
        for rng in template_merges:
            new_min_row = rng.min_row + offset
            new_max_row = rng.max_row + offset
            new_range = f"{get_column_letter(rng.min_col)}{new_min_row}:{get_column_letter(rng.max_col)}{new_max_row}"
            try:
                ws.merge_cells(new_range)
            except Exception:
                pass

        insert_at += block_height

    return block_start, block_height, base_blocks + extra


def _limpiar_bloque_materia(ws, row_start):
    for row in range(row_start, row_start + 4):
        for col in range(1, ws.max_column + 1):
            if col <= 12:
                _safe_set_cell(ws, row, col, '')


def _buscar_celda_por_texto(ws, texto):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and texto in cell.value:
                return cell
    return None


def _find_soffice_executable():
    env_path = os.getenv('SOFFICE_PATH') or os.getenv('LIBREOFFICE_PATH')
    if env_path and os.path.exists(env_path):
        return env_path

    for name in ('soffice', 'soffice.exe'):
        found = shutil.which(name)
        if found:
            return found

    if os.name == 'nt':
        candidates = [
            r"C:\\Program Files\\LibreOffice\\program\\soffice.exe",
            r"C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe",
        ]
        for path in candidates:
            if os.path.exists(path):
                return path

    return None


def _convert_xlsx_a_pdf(xlsx_path):
    """Convierte XLSX a PDF si LibreOffice esta disponible. Retorna ruta PDF o None."""
    soffice = _find_soffice_executable()
    if not soffice:
        return None
    output_dir = os.path.dirname(xlsx_path)
    try:
        subprocess.run(
            [soffice, '--headless', '--convert-to', 'pdf', '--outdir', output_dir, xlsx_path],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=60
        )
    except Exception:
        return None
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    pdf_path = os.path.join(output_dir, f"{base}.pdf")
    return pdf_path if os.path.exists(pdf_path) else None


def _limpiar_nombre_archivo(texto):
    if not texto:
        return ''
    limpio = []
    for ch in str(texto):
        if ch.isalnum() or ch in ('_', '-'):
            limpio.append(ch)
        elif ch.isspace():
            limpio.append('_')
    return ''.join(limpio).strip('_')


def _boletin_signature(cursor, grupo_id, periodo_id):
    cursor.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM notas n
        JOIN actividades a ON a.id_actividad = n.id_actividad
        WHERE a.id_grupo = %s AND n.id_periodo = %s
        """,
        (grupo_id, periodo_id)
    )
    notas_cnt = (cursor.fetchone() or {}).get('cnt') or 0

    cursor.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM asistencias_por_periodo ap
        JOIN estudiantes e ON e.id_estudiante = ap.id_estudiante
        WHERE e.id_grupo = %s AND ap.id_periodo = %s
        """,
        (grupo_id, periodo_id)
    )
    fallas_cnt = (cursor.fetchone() or {}).get('cnt') or 0

    cursor.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM estudiantes
        WHERE id_grupo = %s AND estado = 'Activo'
        """,
        (grupo_id,)
    )
    est_cnt = (cursor.fetchone() or {}).get('cnt') or 0

    return f"{notas_cnt}|{fallas_cnt}|{est_cnt}"


def _leer_boletin_log():
    log_path = os.path.join(ESCRITORIO, '.boletin_sync_log.txt')
    data = {}
    if os.path.exists(log_path):
        try:
            with open(log_path, 'r', encoding='utf-8') as f:
                for line in f:
                    partes = line.strip().split('|')
                    if len(partes) >= 3:
                        key = partes[0]
                        data[key] = partes[1]
        except Exception:
            pass
    return data


def _guardar_boletin_log(log_data):
    log_path = os.path.join(ESCRITORIO, '.boletin_sync_log.txt')
    try:
        with open(log_path, 'w', encoding='utf-8') as f:
            for key, firma in log_data.items():
                f.write(f"{key}|{firma}|{datetime.now().isoformat()}\n")
    except Exception:
        pass


def _calcular_puestos(cursor, grupo_id, periodos, materias, notas_map):
    cursor.execute(
        """
        SELECT id_estudiante, nombre, apellido
        FROM estudiantes
        WHERE id_grupo = %s AND estado = 'Activo'
        """,
        (grupo_id,)
    )
    estudiantes = cursor.fetchall() or []

    periodos_nums = [int(p['numero_periodo']) for p in periodos]
    periodos_map = {int(p['numero_periodo']): int(p['id_periodo']) for p in periodos}

    def promedio_periodo(estudiante_id, periodo_num):
        periodo_id = periodos_map.get(periodo_num)
        notas = []
        for m in materias:
            nota = _calcular_nota_periodo(cursor, notas_map, estudiante_id, grupo_id, m['id_materia'], periodo_id)
            if nota is not None:
                notas.append(nota)
        if not notas:
            return None
        return round(sum(notas) / len(notas), 2)

    promedios = {p: [] for p in periodos_nums}
    pf_promedios = []

    for est in estudiantes:
        est_id = est['id_estudiante']
        per_vals = {}
        for periodo_num in periodos_nums:
            per_vals[periodo_num] = promedio_periodo(est_id, periodo_num)
            if per_vals[periodo_num] is not None:
                promedios[periodo_num].append((est_id, per_vals[periodo_num]))

        vals = [v for v in per_vals.values() if v is not None]
        pf_val = round(sum(vals) / len(vals), 2) if vals else None
        if pf_val is not None:
            pf_promedios.append((est_id, pf_val))

    def rank_dense(items):
        orden = sorted(items, key=lambda x: x[1], reverse=True)
        ranks = {}
        last_score = None
        rank = 0
        for est_id, score in orden:
            if last_score is None or score < last_score:
                rank += 1
            ranks[est_id] = rank
            last_score = score
        return ranks

    ranks_periodo = {p: rank_dense(promedios[p]) for p in promedios}
    ranks_pf = rank_dense(pf_promedios)

    return ranks_periodo, ranks_pf


def _generar_boletin_excel(cursor, estudiante_id, periodo_id, *, save_path=None):
    plantilla = _resolver_ruta_plantilla_boletin()
    if not plantilla:
        raise FileNotFoundError('No se encontró PlantillaBoletin en templates.')

    cursor.execute(
        """
        SELECT e.*, gr.codigo_grupo, g.numero_grado, g.nombre_grado
        FROM estudiantes e
        JOIN grupos gr ON gr.id_grupo = e.id_grupo
        JOIN grados g ON g.id_grado = gr.id_grado
        WHERE e.id_estudiante = %s
        LIMIT 1
        """,
        (estudiante_id,)
    )
    estudiante = cursor.fetchone()
    if not estudiante:
        raise ValueError('Estudiante no encontrado')

    grupo_id = estudiante['id_grupo']

    cursor.execute(
        """
        SELECT id_periodo, numero_periodo, nombre_periodo
        FROM periodos
        ORDER BY numero_periodo
        """
    )
    periodos = cursor.fetchall() or []
    if not periodos:
        raise ValueError('No hay periodos definidos')

    periodo_actual = next((p for p in periodos if int(p['id_periodo']) == int(periodo_id)), None)
    if not periodo_actual:
        raise ValueError('Periodo no encontrado')

    cursor.execute(
        """
        SELECT DISTINCT m.id_materia, m.nombre_materia, m.codigo_materia, m.intensidad_horaria,
               GROUP_CONCAT(DISTINCT CONCAT(u.nombre, ' ', u.apellido) ORDER BY u.apellido SEPARATOR ', ') AS docentes
        FROM asignaciones_docente a
        JOIN materias m ON m.id_materia = a.id_materia
        LEFT JOIN usuarios u ON u.id_usuario = a.id_usuario
        WHERE a.id_grupo = %s AND a.estado = 'Activa'
        GROUP BY m.id_materia
        ORDER BY m.nombre_materia
        """,
        (grupo_id,)
    )
    materias = cursor.fetchall() or []
    if not materias:
        raise ValueError('No hay materias asignadas a este grupo')

    cursor.execute(
        """
        SELECT n.id_estudiante, n.id_actividad, n.id_periodo, n.puntaje_obtenido
        FROM notas n
        JOIN actividades a ON a.id_actividad = n.id_actividad
        WHERE a.id_grupo = %s
          AND n.id_periodo IN ({})
        """.format(','.join(['%s'] * len(periodos))),
        [grupo_id] + [p['id_periodo'] for p in periodos]
    )
    notas_rows = cursor.fetchall() or []
    notas_map = {
        (int(n['id_estudiante']), int(n['id_actividad']), int(n['id_periodo'])): n.get('puntaje_obtenido')
        for n in notas_rows
    }

    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active

    nombre_estudiante = f"{estudiante.get('nombre', '')} {estudiante.get('apellido', '')}".strip()
    documento = estudiante.get('documento') or ''
    nombre_grado = estudiante.get('nombre_grado') or f"Grado {estudiante.get('numero_grado', '')}"
    codigo_grupo = estudiante.get('codigo_grupo') or ''
    fecha_str = datetime.now().strftime('%d/%m/%Y')

    info_usuario = {}
    if session.get('user_id'):
        cursor.execute("SELECT * FROM usuarios WHERE id_usuario = %s LIMIT 1", (session.get('user_id'),))
        info_usuario = cursor.fetchone() or {}

    mapping = {
        '{{LOGO ISNTITUCION}}': '',
        '{{NOMBRE COMPLETO DE LA INSTITUCIÓN}}': info_usuario.get('institucion') or 'DocstrY',
        '{{ID IDENTIFICACIÓN}}': documento,
        '{{NOMBRE COMPLETO ESTUDIANTE}}': nombre_estudiante,
        '{{GRADO}}': nombre_grado,
        '{{GRUPO}}': codigo_grupo,
        '{{JORNADA}}': info_usuario.get('jornada') or '',
        '{{FECHA}}': fecha_str,
    }
    _replace_placeholders_worksheet(ws, mapping)

    block_start, block_height, total_blocks = _ensure_materia_blocks(ws, len(materias))

    period_cols = {1: 8, 2: 9, 3: 10, 4: 11}
    periodos_map = {int(p['numero_periodo']): int(p['id_periodo']) for p in periodos}

    ranks_periodo, ranks_pf = _calcular_puestos(cursor, grupo_id, periodos, materias, notas_map)

    periodos_nums = [int(p['numero_periodo']) for p in periodos if int(p['numero_periodo']) in period_cols]

    promedios_periodo = {p: [] for p in periodos_nums}
    promedio_estudiante = {}
    fallas_periodo = {p: 0 for p in periodos_nums}

    for idx, materia in enumerate(materias):
        row_start = block_start + idx * block_height
        _safe_set_cell(ws, row_start, 1, materia.get('nombre_materia', ''))
        docente_txt = materia.get('docentes') or ''
        _safe_set_cell(ws, row_start + 1, 1, f"{materia.get('nombre_materia', '')} - {docente_txt}".strip(' -'))

        ih = materia.get('intensidad_horaria')
        _safe_set_cell(ws, row_start, 7, ih if ih is not None else '')

        notas_periodos = {}
        for periodo_num in periodos_nums:
            periodo_db = periodos_map.get(periodo_num)
            nota = _calcular_nota_periodo(cursor, notas_map, estudiante_id, grupo_id, materia['id_materia'], periodo_db)
            notas_periodos[periodo_num] = nota
            if nota is not None:
                promedios_periodo[periodo_num].append(nota)

            fallas = _obtener_fallas_periodo(cursor, estudiante_id, materia['id_materia'], periodo_db)
            fallas_periodo[periodo_num] += fallas

            col = period_cols[periodo_num]
            _safe_set_cell(ws, row_start, col, nota if nota is not None else '')
            _safe_set_cell(ws, row_start + 2, col, fallas)

        notas_validas = [v for v in notas_periodos.values() if v is not None]
        pf = round(sum(notas_validas) / len(notas_validas), 2) if notas_validas else None
        _safe_set_cell(ws, row_start, 12, pf if pf is not None else '')

    for idx in range(len(materias), total_blocks):
        row_start = block_start + idx * block_height
        _limpiar_bloque_materia(ws, row_start)

    promedio_cell = _buscar_celda_por_texto(ws, 'Promedio nota del Periodo')
    fallas_cell = _buscar_celda_por_texto(ws, 'Total falla del Periodo')
    puesto_cell = _buscar_celda_por_texto(ws, 'Puesto x Periodo')

    for periodo_num in periodos_nums:
        notas = promedios_periodo.get(periodo_num) or []
        promedio_estudiante[periodo_num] = round(sum(notas) / len(notas), 2) if notas else None

    if promedio_cell:
        base_row = promedio_cell.row
        for periodo_num in periodos_nums:
            col = period_cols[periodo_num]
            promedio = promedio_estudiante.get(periodo_num)
            _safe_set_cell(ws, base_row, col, promedio if promedio is not None else '')
        # PF
        pf_vals = [
            ws.cell(row=block_start + idx * block_height, column=12).value
            for idx in range(len(materias))
        ]
        pf_vals = [v for v in pf_vals if isinstance(v, (int, float))]
        pf_prom = round(sum(pf_vals) / len(pf_vals), 2) if pf_vals else None
        _safe_set_cell(ws, base_row, 12, pf_prom if pf_prom is not None else '')
    else:
        pf_vals = []
        for periodo_num in periodos_nums:
            val = promedio_estudiante.get(periodo_num)
            if val is not None:
                pf_vals.append(val)
        pf_prom = round(sum(pf_vals) / len(pf_vals), 2) if pf_vals else None

    if fallas_cell:
        base_row = fallas_cell.row
        for periodo_num in periodos_nums:
            col = period_cols[periodo_num]
            _safe_set_cell(ws, base_row, col, fallas_periodo.get(periodo_num, 0))
        total_fallas_pf = sum(fallas_periodo.values())
        _safe_set_cell(ws, base_row, 12, total_fallas_pf)

    if puesto_cell:
        base_row = puesto_cell.row
        for periodo_num in periodos_nums:
            col = period_cols[periodo_num]
            puesto = (ranks_periodo.get(periodo_num) or {}).get(estudiante_id)
            _safe_set_cell(ws, base_row, col, puesto if puesto is not None else '')
        pf_puesto = ranks_pf.get(estudiante_id)
        _safe_set_cell(ws, base_row, 12, pf_puesto if pf_puesto is not None else '')

    # Guardar puestos en tabla (si existe)
    _asegurar_tabla_boletin_puestos(cursor)
    for periodo_num in periodos_nums:
        periodo_db = periodos_map.get(periodo_num)
        promedio = promedio_estudiante.get(periodo_num)
        puesto = (ranks_periodo.get(periodo_num) or {}).get(estudiante_id)
        total_fallas = fallas_periodo.get(periodo_num, 0)
        try:
            cursor.execute(
                """
                INSERT INTO boletin_puestos (id_estudiante, id_grupo, id_periodo, es_definitiva, promedio, total_fallas, puesto)
                VALUES (%s, %s, %s, 0, %s, %s, %s)
                ON DUPLICATE KEY UPDATE promedio = VALUES(promedio), total_fallas = VALUES(total_fallas), puesto = VALUES(puesto), fecha_generado = CURRENT_TIMESTAMP
                """,
                (estudiante_id, grupo_id, periodo_db, promedio, total_fallas, puesto)
            )
        except Exception:
            pass

    try:
        cursor.execute(
            """
            INSERT INTO boletin_puestos (id_estudiante, id_grupo, id_periodo, es_definitiva, promedio, total_fallas, puesto)
            VALUES (%s, %s, 0, 1, %s, %s, %s)
            ON DUPLICATE KEY UPDATE promedio = VALUES(promedio), total_fallas = VALUES(total_fallas), puesto = VALUES(puesto), fecha_generado = CURRENT_TIMESTAMP
            """,
            (estudiante_id, grupo_id, pf_prom, sum(fallas_periodo.values()), ranks_pf.get(estudiante_id))
        )
    except Exception:
        pass

    if save_path:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        wb.save(save_path)
        return {
            'estudiante': estudiante,
            'periodo': periodo_actual,
            'materias': materias
        }

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@reportes_bp.route('/api/reportes/enviar-correo', methods=['POST'])
def api_enviar_reporte_correo():
    """Envía un reporte por correo a uno o varios destinatarios"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'No autenticado'}), 401

        data = request.get_json()
        remitente = data.get('remitente', '')
        tipo_reporte = data.get('tipoReporte', '')
        correos = data.get('correos', [])
        mensaje_usuario = data.get('mensaje', '')

        if not remitente or not tipo_reporte or not correos or not mensaje_usuario:
            return jsonify({'error': 'Faltan datos requeridos'}), 400

        if not isinstance(correos, list) or len(correos) == 0:
            return jsonify({'error': 'Correos inválidos'}), 400

        asuntos = {
            'boletin': 'Boletín de Calificaciones',
            'consolidado': 'Consolidado de Notas',
            'asistencia_diaria': 'Reporte de Asistencia Diaria',
            'estudiantes': 'Lista de Estudiantes'
        }
        asunto = asuntos.get(tipo_reporte, 'Reporte de la Institución')

        success = enviar_correo_smtp(remitente, correos, asunto, mensaje_usuario)

        if success:
            log_action(session['user_id'], 'Export',
                       f'Envió {tipo_reporte} a {len(correos)} destinatario(s)')
            return jsonify({
                'status': 'ok',
                'message': f'Reporte enviado a {len(correos)} destinatario(s)',
                'destinatarios': len(correos)
            }), 200
        else:
            return jsonify({'error': 'Error al enviar el correo'}), 502

    except Exception as e:
        print(f"Error en api_enviar_reporte_correo: {e}")
        return _error_interno(e)


@reportes_bp.route('/api/reportes/enviar-correo-adjunto', methods=['POST'])
def api_enviar_reporte_correo_adjunto():
    """Envía un reporte por correo con archivo adjunto (PDF)."""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'No autenticado'}), 401

        remitente = request.form.get('remitente', '')
        tipo_reporte = request.form.get('tipoReporte', '')
        mensaje_usuario = request.form.get('mensaje', '')
        correos_raw = request.form.get('correos', '')

        if not remitente or not tipo_reporte or not mensaje_usuario or not correos_raw:
            return jsonify({'error': 'Faltan datos requeridos'}), 400

        try:
            correos = json.loads(correos_raw) if correos_raw.strip().startswith('[') else [c.strip() for c in correos_raw.split(',') if c.strip()]
        except Exception:
            correos = [c.strip() for c in correos_raw.split(',') if c.strip()]

        if not isinstance(correos, list) or len(correos) == 0:
            return jsonify({'error': 'Correos inválidos'}), 400

        archivo = request.files.get('archivo_pdf') or request.files.get('archivo')
        if not archivo:
            return jsonify({'error': 'No se adjuntó archivo PDF'}), 400

        nombre = archivo.filename or 'reporte.pdf'
        if not nombre.lower().endswith('.pdf'):
            return jsonify({'error': 'Solo se permiten archivos PDF'}), 400

        archivo_bytes = archivo.read()
        adjuntos = [{
            'filename': nombre,
            'data': archivo_bytes,
            'mimetype': 'application/pdf'
        }]

        asuntos = {
            'boletin': 'Boletín de Calificaciones',
            'consolidado': 'Consolidado de Notas',
            'asistencia_diaria': 'Reporte de Asistencia Diaria',
            'estudiantes': 'Lista de Estudiantes'
        }
        asunto = asuntos.get(tipo_reporte, 'Reporte de la Institución')

        success = enviar_correo_smtp_con_adjuntos(remitente, correos, asunto, mensaje_usuario, adjuntos)

        if success:
            log_action(session['user_id'], 'Export',
                       f'Envió {tipo_reporte} (adjunto) a {len(correos)} destinatario(s)')
            return jsonify({
                'status': 'ok',
                'message': f'Reporte enviado a {len(correos)} destinatario(s)',
                'destinatarios': len(correos)
            }), 200
        return jsonify({'error': 'Error al enviar el correo'}), 502

    except Exception as e:
        print(f"Error en api_enviar_reporte_correo_adjunto: {e}")
        return _error_interno(e)


def _ruta_boletin_archivo(estudiante, numero_periodo, ext):
    grado = estudiante.get('numero_grado') or ''
    grupo = estudiante.get('codigo_grupo') or ''
    nombre = _limpiar_nombre_archivo(f"{estudiante.get('apellido', '')}_{estudiante.get('nombre', '')}")
    documento = _limpiar_nombre_archivo(estudiante.get('documento') or '')
    base_dir = os.path.join(
        BOLETINES_DIR,
        f"Periodo_{numero_periodo}",
        f"Grado_{grado}",
        f"Grupo_{grupo}"
    )
    filename = f"Boletin_{nombre}_{documento}_G{grado}_{grupo}.{ext}"
    return os.path.join(base_dir, filename)


@reportes_bp.route('/api/boletines/generar', methods=['POST'])
def api_boletin_generar_archivo():
    """Genera boletin en disco y devuelve PDF/XLSX."""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'No autenticado'}), 401

        data = request.get_json() or {}
        estudiante_id = data.get('id_estudiante')
        periodo_id = data.get('periodo_id')
        formato = (data.get('formato') or 'pdf').lower()

        if not estudiante_id or not periodo_id:
            return jsonify({'error': 'Faltan datos (id_estudiante, periodo_id)'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT e.id_estudiante, e.nombre, e.apellido, e.documento, gr.codigo_grupo, g.numero_grado
            FROM estudiantes e
            JOIN grupos gr ON gr.id_grupo = e.id_grupo
            JOIN grados g ON g.id_grado = gr.id_grado
            WHERE e.id_estudiante = %s
            LIMIT 1
            """,
            (estudiante_id,)
        )
        estudiante = cursor.fetchone()
        if not estudiante:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Estudiante no encontrado'}), 404

        cursor.execute("SELECT id_periodo, numero_periodo FROM periodos WHERE id_periodo = %s", (periodo_id,))
        periodo = cursor.fetchone()
        if not periodo:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Periodo no encontrado'}), 404

        numero_periodo = periodo.get('numero_periodo') or periodo_id
        ruta_xlsx = _ruta_boletin_archivo(estudiante, numero_periodo, 'xlsx')
        _generar_boletin_excel(cursor, int(estudiante_id), int(periodo_id), save_path=ruta_xlsx)
        conn.commit()
        cursor.close()
        conn.close()

        if formato == 'pdf':
            ruta_pdf = _convert_xlsx_a_pdf(ruta_xlsx)
            if not ruta_pdf:
                return jsonify({
                    'error': (
                        'No se pudo generar PDF. LibreOffice no se encontro. '
                        'Agrega "C:\\Program Files\\LibreOffice\\program" al PATH '
                        'o define SOFFICE_PATH con la ruta de soffice.exe.'
                    )
                }), 500
            return send_file(
                ruta_pdf,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=os.path.basename(ruta_pdf)
            )

        return send_file(
            ruta_xlsx,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=os.path.basename(ruta_xlsx)
        )

    except Exception as e:
        print(f"Error en api_boletin_generar_archivo: {e}")
        return _error_interno(e)


@reportes_bp.route('/api/boletines/sincronizar-grupo', methods=['POST'])
def api_boletin_sincronizar_grupo():
    """Genera/actualiza boletines para un grupo y periodo, con deteccion de cambios."""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'No autenticado'}), 401

        data = request.get_json() or {}
        grupo_id = data.get('grupo_id')
        periodo_id = data.get('periodo_id')
        force = bool(data.get('force'))

        print(f"[boletines] Sync grupo={grupo_id} periodo={periodo_id}", flush=True)

        if not grupo_id or not periodo_id:
            return jsonify({'error': 'Faltan datos (grupo_id, periodo_id)'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT id_periodo, numero_periodo FROM periodos WHERE id_periodo = %s", (periodo_id,))
        periodo = cursor.fetchone()
        if not periodo:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Periodo no encontrado'}), 404

        firma = _boletin_signature(cursor, grupo_id, periodo_id)
        log_data = _leer_boletin_log()
        key = f"{grupo_id}_{periodo_id}"

        if not force and log_data.get(key) == firma:
            cursor.close()
            conn.close()
            return jsonify({'status': 'ok', 'message': 'Sin cambios detectados', 'archivos_actualizados': 0}), 200

        cursor.execute(
            """
            SELECT e.id_estudiante, e.nombre, e.apellido, e.documento, e.id_grupo, gr.codigo_grupo, g.numero_grado
            FROM estudiantes e
            JOIN grupos gr ON gr.id_grupo = e.id_grupo
            JOIN grados g ON g.id_grado = gr.id_grado
            WHERE e.id_grupo = %s AND e.estado = 'Activo'
            ORDER BY e.apellido, e.nombre
            """,
            (grupo_id,)
        )
        estudiantes = cursor.fetchall() or []

        if not estudiantes:
            cursor.close()
            conn.close()
            return jsonify({'error': 'No hay estudiantes activos en el grupo'}), 400

        print(f"[boletines] Estudiantes: {len(estudiantes)}", flush=True)

        stats = {'archivos_creados': 0, 'archivos_actualizados': 0}

        for est in estudiantes:
            ruta_xlsx = _ruta_boletin_archivo(est, periodo.get('numero_periodo') or periodo_id, 'xlsx')
            existe = os.path.exists(ruta_xlsx)
            _generar_boletin_excel(cursor, est['id_estudiante'], periodo_id, save_path=ruta_xlsx)
            if existe:
                stats['archivos_actualizados'] += 1
            else:
                stats['archivos_creados'] += 1

            _convert_xlsx_a_pdf(ruta_xlsx)

        conn.commit()
        log_data[key] = firma
        _guardar_boletin_log(log_data)
        cursor.close()
        conn.close()

        if session.get('user_id'):
            log_action(
                session.get('user_id'),
                'Export',
                (
                    f"Sincronizacion boletines grupo={grupo_id} periodo={periodo_id}. "
                    f"Nuevos={stats['archivos_creados']} Actualizados={stats['archivos_actualizados']}"
                ),
                tabla_afectada='boletines',
                registro_id=int(grupo_id)
            )

        print(
            f"[boletines] Terminado grupo. Nuevos={stats['archivos_creados']} "
            f"Actualizados={stats['archivos_actualizados']}",
            flush=True
        )

        return jsonify({'status': 'ok', 'message': 'Boletines sincronizados', **stats}), 200

    except Exception as e:
        print(f"Error en api_boletin_sincronizar_grupo: {e}")
        return _error_interno(e)


@reportes_bp.route('/api/boletines/enviar-grupo', methods=['POST'])
def api_boletin_enviar_grupo():
    """Genera boletines del grupo y los envia por correo a estudiantes con email valido."""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'No autenticado'}), 401

        data = request.get_json() or {}
        grupo_id = data.get('grupo_id')
        periodo_id = data.get('periodo_id')
        remitente = data.get('remitente', '')
        mensaje = data.get('mensaje', '')

        if not grupo_id or not periodo_id or not remitente or not mensaje:
            return jsonify({'error': 'Faltan datos (grupo_id, periodo_id, remitente, mensaje)'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT id_periodo, numero_periodo FROM periodos WHERE id_periodo = %s", (periodo_id,))
        periodo = cursor.fetchone()
        if not periodo:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Periodo no encontrado'}), 404

        cursor.execute(
            """
            SELECT e.id_estudiante, e.nombre, e.apellido, e.documento, e.correo, e.id_grupo,
                   gr.codigo_grupo, g.numero_grado
            FROM estudiantes e
            JOIN grupos gr ON gr.id_grupo = e.id_grupo
            JOIN grados g ON g.id_grado = gr.id_grado
            WHERE e.id_grupo = %s AND e.estado = 'Activo'
            ORDER BY e.apellido, e.nombre
            """,
            (grupo_id,)
        )
        estudiantes = cursor.fetchall() or []
        if not estudiantes:
            cursor.close()
            conn.close()
            return jsonify({'error': 'No hay estudiantes activos en el grupo'}), 400

        enviados = 0
        omitidos = 0
        errores = []

        for est in estudiantes:
            correo = est.get('correo') or ''
            if not _email_valido_simple(correo):
                omitidos += 1
                continue

            ruta_xlsx = _ruta_boletin_archivo(est, periodo.get('numero_periodo') or periodo_id, 'xlsx')
            _generar_boletin_excel(cursor, est['id_estudiante'], periodo_id, save_path=ruta_xlsx)
            ruta_pdf = _convert_xlsx_a_pdf(ruta_xlsx)
            if not ruta_pdf:
                errores.append({'estudiante': est['id_estudiante'], 'error': 'No se pudo generar PDF'})
                continue

            with open(ruta_pdf, 'rb') as f:
                adjuntos = [{
                    'filename': os.path.basename(ruta_pdf),
                    'data': f.read(),
                    'mimetype': 'application/pdf'
                }]

            asunto = 'Boletín de Calificaciones'
            ok = enviar_correo_smtp_con_adjuntos(remitente, [correo], asunto, mensaje, adjuntos)
            if ok:
                enviados += 1
            else:
                errores.append({'estudiante': est['id_estudiante'], 'error': 'No se pudo enviar'})

        conn.commit()
        cursor.close()
        conn.close()
        log_action(session['user_id'], 'Export', f'Envió boletines a {enviados} destinatario(s)')

        return jsonify({
            'status': 'ok',
            'enviados': enviados,
            'omitidos': omitidos,
            'errores': errores
        }), 200

    except Exception as e:
        print(f"Error en api_boletin_enviar_grupo: {e}")
        return _error_interno(e)


@reportes_bp.route('/api/boletines/sincronizar-todo', methods=['POST'])
def api_boletin_sincronizar_todo():
    """Genera/actualiza boletines para todos los grupos y periodos."""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'No autenticado'}), 401

        data = request.get_json() or {}
        force = bool(data.get('force'))

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT id_periodo, numero_periodo FROM periodos ORDER BY numero_periodo")
        periodos = cursor.fetchall() or []
        if not periodos:
            cursor.close()
            conn.close()
            return jsonify({'error': 'No hay periodos definidos'}), 400

        cursor.execute(
            """
            SELECT g.id_grado, g.numero_grado, gr.id_grupo, gr.codigo_grupo
            FROM grupos gr
            JOIN grados g ON gr.id_grado = g.id_grado
            ORDER BY g.numero_grado, gr.codigo_grupo
            """
        )
        grupos = cursor.fetchall() or []
        if not grupos:
            cursor.close()
            conn.close()
            return jsonify({'error': 'No hay grupos definidos'}), 400

        print(f"[boletines] Sync total: periodos={len(periodos)} grupos={len(grupos)}", flush=True)

        log_data = _leer_boletin_log()
        stats = {
            'archivos_creados': 0,
            'archivos_actualizados': 0,
            'grupos_procesados': 0,
            'periodos_procesados': len(periodos)
        }

        for periodo in periodos:
            periodo_id = periodo['id_periodo']
            numero_periodo = periodo['numero_periodo']

            for grupo in grupos:
                grupo_id = grupo['id_grupo']
                firma = _boletin_signature(cursor, grupo_id, periodo_id)
                key = f"{grupo_id}_{periodo_id}"

                if not force and log_data.get(key) == firma:
                    continue

                print(f"[boletines] Procesando grupo={grupo_id} periodo={periodo_id}", flush=True)

                cursor.execute(
                    """
                    SELECT e.id_estudiante, e.nombre, e.apellido, e.documento, e.id_grupo,
                           gr.codigo_grupo, g.numero_grado
                    FROM estudiantes e
                    JOIN grupos gr ON gr.id_grupo = e.id_grupo
                    JOIN grados g ON g.id_grado = gr.id_grado
                    WHERE e.id_grupo = %s AND e.estado = 'Activo'
                    ORDER BY e.apellido, e.nombre
                    """,
                    (grupo_id,)
                )
                estudiantes = cursor.fetchall() or []
                if not estudiantes:
                    continue

                for est in estudiantes:
                    ruta_xlsx = _ruta_boletin_archivo(est, numero_periodo or periodo_id, 'xlsx')
                    existe = os.path.exists(ruta_xlsx)
                    _generar_boletin_excel(cursor, est['id_estudiante'], periodo_id, save_path=ruta_xlsx)
                    if existe:
                        stats['archivos_actualizados'] += 1
                    else:
                        stats['archivos_creados'] += 1
                    _convert_xlsx_a_pdf(ruta_xlsx)

                log_data[key] = firma
                stats['grupos_procesados'] += 1

        conn.commit()
        _guardar_boletin_log(log_data)
        cursor.close()
        conn.close()

        if session.get('user_id'):
            log_action(
                session.get('user_id'),
                'Export',
                (
                    f"Sincronizacion boletines total. Nuevos={stats['archivos_creados']} "
                    f"Actualizados={stats['archivos_actualizados']} Grupos={stats['grupos_procesados']}"
                ),
                tabla_afectada='boletines'
            )

        print(
            f"[boletines] Sync completo. Nuevos={stats['archivos_creados']} "
            f"Actualizados={stats['archivos_actualizados']} Grupos={stats['grupos_procesados']}",
            flush=True
        )

        return jsonify({'status': 'ok', 'message': 'Boletines sincronizados', **stats}), 200

    except Exception as e:
        print(f"Error en api_boletin_sincronizar_todo: {e}")
        return _error_interno(e)
