# ═══════════════════════════════════════════════════════════════════════════════
# ASISTENCIA — Módulo independiente de control de asistencia
# ═══════════════════════════════════════════════════════════════════════════════
from flask import Blueprint, jsonify, request, session, send_file
from utils.database import get_db
from utils.helpers import _error_interno
from datetime import datetime
import io
import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from uuid import uuid4
from werkzeug.utils import secure_filename

asistencia_bp = Blueprint('asistencia', __name__)

_ROLES_ADMIN_EDICION = {'server_admin', 'admin_server', 'admin'}
_ROLES_APROBACION_JUSTIFICANTES = {'server_admin', 'admin_server', 'admin', 'rector', 'coordinador'}
_TIPOS_JUSTIFICANTE = {'Médico', 'Familiar', 'Administrativo', 'Otro'}
_EXTENSIONES_PERMITIDAS_JUSTIFICANTE = {'.pdf', '.png', '.jpg', '.jpeg', '.webp', '.doc', '.docx'}

_BASE_BACKEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'backend'))
_JUSTIFICANTES_UPLOAD_DIR = os.path.join(_BASE_BACKEND_DIR, 'uploads', 'justificantes')
_PLANTILLA_ASISTENCIAS_PATH = os.path.join(_BASE_BACKEND_DIR, 'templates', 'PlantillaAsistencias_ReporteGeneral.xlsx')


def _resolver_ruta_plantilla_asistencias():
    """Resuelve la ruta de la plantilla base de asistencias en diferentes layouts del proyecto."""
    parent = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    rutas_posibles = [
        _PLANTILLA_ASISTENCIAS_PATH,
        os.path.join(parent, 'templates', 'PlantillaAsistencias_ReporteGeneral.xlsx'),
        os.path.join(parent, 'backend', 'templates', 'PlantillaAsistencias_ReporteGeneral.xlsx'),
    ]
    for ruta in rutas_posibles:
        if os.path.exists(ruta):
            return ruta
    return None


def _set_cell_value_safe(ws, cell_ref, value):
    """Escribe en celdas incluso si forman parte de un rango combinado."""
    cell = ws[cell_ref]
    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
        cell.value = value
        return

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
            return

    # Fallback defensivo
    ws[cell_ref] = value


def _asegurar_carpeta_justificantes():
    os.makedirs(_JUSTIFICANTES_UPLOAD_DIR, exist_ok=True)


def _extension_permitida(nombre_archivo):
    _, ext = os.path.splitext(nombre_archivo or '')
    return ext.lower() in _EXTENSIONES_PERMITIDAS_JUSTIFICANTE


def _ruta_relativa_justificante(nombre_archivo):
    return f"uploads/justificantes/{nombre_archivo}"


def _ruta_absoluta_desde_relativa(ruta_relativa):
    if not ruta_relativa:
        return None

    ruta_normalizada = str(ruta_relativa).replace('\\', '/').lstrip('/')
    return os.path.abspath(os.path.join(_BASE_BACKEND_DIR, ruta_normalizada))


def _reemplazar_etiquetas_hoja(ws, reemplazos):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            valor = cell.value
            if not isinstance(valor, str):
                continue
            nuevo = valor
            for tag, texto in reemplazos.items():
                if tag in nuevo:
                    nuevo = nuevo.replace(tag, str(texto or ''))
            if nuevo != valor:
                cell.value = nuevo


def _a_estado_asistencia_legacy(estado_detalle):
    mapping = {
        'presente': 'Presente',
        'ausente': 'Ausente',
        'tardio': 'Tardío',
        'no_registrado': 'Ausente'
    }
    return mapping.get(str(estado_detalle or '').lower(), 'Ausente')


def _upsert_asistencia_legacy(cursor, id_estudiante, id_grupo, id_periodo, fecha, asistio, comentario):
    estado_legacy = _a_estado_asistencia_legacy(asistio)

    cursor.execute("""
        INSERT INTO asistencia (id_estudiante, id_grupo, id_periodo, fecha_asistencia, estado, observaciones)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            estado = VALUES(estado),
            observaciones = VALUES(observaciones),
            updated_at = NOW()
    """, (id_estudiante, id_grupo, id_periodo, fecha, estado_legacy, comentario))

    cursor.execute("""
        SELECT id_asistencia FROM asistencia
        WHERE id_estudiante = %s AND id_grupo = %s AND id_periodo = %s AND fecha_asistencia = %s
        LIMIT 1
    """, (id_estudiante, id_grupo, id_periodo, fecha))
    row = cursor.fetchone()
    return row['id_asistencia'] if row else None


def _es_rol_admin_edicion():
    rol = str(session.get('user_role') or '').strip().lower()
    return rol in _ROLES_ADMIN_EDICION


def _bloquear_docente_si_periodo_cerrado(cursor, id_periodo):
    cursor.execute("SELECT estado FROM periodos WHERE id_periodo = %s LIMIT 1", (id_periodo,))
    periodo = cursor.fetchone()
    if not periodo:
        return jsonify({'error': 'Período no encontrado'}), 404

    estado = str(periodo.get('estado') or '').strip().lower()
    if estado == 'cerrado' and not _es_rol_admin_edicion():
        return jsonify({'error': 'El período está cerrado. Solo el administrador puede realizar cambios.'}), 403

    return None

@asistencia_bp.route('/api/asistencia/filtros', methods=['GET'])
def api_asistencia_filtros():
    """Filtros para asistencia: grados + periodos (reusa misma lógica que calificaciones)."""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        user_id = session.get('user_id')
        user_role = session.get('user_role', '').lower()

        if user_role in ('server_admin', 'rector', 'coordinador'):
            cursor.execute("SELECT id_grado, nombre_grado, numero_grado FROM grados ORDER BY numero_grado")
            grados = cursor.fetchall()
        else:
            cursor.execute("""
                SELECT DISTINCT g.id_grado, g.nombre_grado, g.numero_grado FROM (
                    SELECT a.id_grado FROM asignaciones_docente a
                    WHERE a.id_usuario = %s AND a.estado = 'Activa'
                    UNION
                    SELECT gr.id_grado FROM actividades act
                    JOIN grupos gr ON act.id_grupo = gr.id_grupo
                    WHERE act.id_usuario = %s
                ) src
                JOIN grados g ON src.id_grado = g.id_grado
                ORDER BY g.numero_grado
            """, (user_id, user_id))
            grados = cursor.fetchall()

        cursor.execute("SELECT * FROM periodos ORDER BY numero_periodo")
        periodos = cursor.fetchall()
        for p in periodos:
            if p.get('fecha_inicio'):
                p['fecha_inicio'] = p['fecha_inicio'].isoformat()
            if p.get('fecha_fin'):
                p['fecha_fin'] = p['fecha_fin'].isoformat()

        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'grados': grados, 'periodos': periodos}), 200
    except Exception as e:
        return _error_interno(e)


@asistencia_bp.route('/api/asistencia/grupos/<int:grado_id>', methods=['GET'])
def api_asistencia_grupos(grado_id):
    """Grupos de un grado para asistencia."""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        user_id = session.get('user_id')
        user_role = session.get('user_role', '').lower()

        if user_role in ('server_admin', 'rector', 'coordinador'):
            cursor.execute("SELECT id_grupo, codigo_grupo FROM grupos WHERE id_grado = %s ORDER BY codigo_grupo", (grado_id,))
        else:
            cursor.execute("""
                SELECT DISTINCT gr.id_grupo, gr.codigo_grupo
                FROM grupos gr WHERE gr.id_grado = %s AND (
                    gr.id_grupo IN (SELECT a.id_grupo FROM asignaciones_docente a WHERE a.id_usuario = %s AND a.id_grado = %s AND a.estado = 'Activa')
                    OR gr.id_grupo IN (SELECT act.id_grupo FROM actividades act WHERE act.id_usuario = %s)
                ) ORDER BY gr.codigo_grupo
            """, (grado_id, user_id, grado_id, user_id))
        grupos = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'grupos': grupos}), 200
    except Exception as e:
        return _error_interno(e)


@asistencia_bp.route('/api/asistencia/materias/<int:grupo_id>', methods=['GET'])
def api_asistencia_materias(grupo_id):
    """Materias para un grupo (filtradas por docente si aplica)."""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        user_id = session.get('user_id')
        user_role = session.get('user_role', '').lower()

        if user_role in ('server_admin', 'rector', 'coordinador'):
            cursor.execute("""
                SELECT DISTINCT m.id_materia, m.nombre_materia FROM (
                    SELECT a.id_materia
                    FROM asignaciones_docente a
                    WHERE a.id_grupo = %s AND a.estado = 'Activa'
                    UNION
                    SELECT act.id_materia
                    FROM actividades act
                    WHERE act.id_grupo = %s
                ) src
                JOIN materias m ON src.id_materia = m.id_materia
                ORDER BY m.nombre_materia
            """, (grupo_id, grupo_id))
        else:
            cursor.execute("""
                SELECT DISTINCT m.id_materia, m.nombre_materia FROM (
                    SELECT act.id_materia FROM actividades act
                    WHERE act.id_grupo = %s AND act.id_usuario = %s
                    UNION
                    SELECT a.id_materia FROM asignaciones_docente a
                    WHERE a.id_grupo = %s AND a.id_usuario = %s AND a.estado = 'Activa'
                ) src
                JOIN materias m ON src.id_materia = m.id_materia
                ORDER BY m.nombre_materia
            """, (grupo_id, user_id, grupo_id, user_id))
        materias = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'materias': materias}), 200
    except Exception as e:
        return _error_interno(e)


@asistencia_bp.route('/api/asistencia/tabla', methods=['GET'])
def api_asistencia_tabla():
    """Cargar tabla de asistencia para una fecha. Devuelve estudiantes + estado actual si ya existe sesión."""
    try:
        grupo_id = request.args.get('grupo_id')
        materia_id = request.args.get('materia_id')
        fecha = request.args.get('fecha')
        periodo_id = request.args.get('periodo_id')

        if not grupo_id or not materia_id or not fecha:
            return jsonify({'error': 'Faltan parámetros: grupo_id, materia_id, fecha'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        # Estudiantes del grupo
        cursor.execute("""
            SELECT id_estudiante, nombre, apellido
            FROM estudiantes
            WHERE id_grupo = %s AND estado = 'Activo'
            ORDER BY apellido, nombre
        """, (grupo_id,))
        estudiantes = cursor.fetchall()

        # Buscar sesión existente para esta fecha/materia/grupo
        cursor.execute("""
            SELECT ad.id_asistencia_diaria, ad.estado as estado_sesion
            FROM asistencias_diarias ad
            WHERE ad.id_grupo = %s AND ad.id_materia = %s AND ad.fecha_registro = %s
                AND ad.estado = 'activa'
            LIMIT 1
        """, (grupo_id, materia_id, fecha))
        sesion = cursor.fetchone()

        registros = {}
        ids_asistencia_legacy = {}
        if sesion:
            # Cargar detalle existente
            cursor.execute("""
                SELECT da.id_estudiante, da.asistio, da.comentario, da.justificante_id,
                       IFNULL(ja.aprobado, 0) AS justificante_aprobado
                FROM detalle_asistencia da
                LEFT JOIN justificantes_ausencia ja ON ja.id_justificante = da.justificante_id
                WHERE da.id_asistencia_diaria = %s
            """, (sesion['id_asistencia_diaria'],))
            for r in cursor.fetchall():
                registros[r['id_estudiante']] = {
                    'asistio': r['asistio'],
                    'comentario': r['comentario'] or '',
                    'justificante_id': r.get('justificante_id'),
                    'justificante_aprobado': bool(r.get('justificante_aprobado'))
                }

        if periodo_id:
            cursor.execute("""
                SELECT id_estudiante, id_asistencia
                FROM asistencia
                WHERE id_grupo = %s AND id_periodo = %s AND fecha_asistencia = %s
            """, (grupo_id, periodo_id, fecha))
            for row in cursor.fetchall():
                ids_asistencia_legacy[row['id_estudiante']] = row['id_asistencia']

        cursor.close()
        conn.close()
        return jsonify({
            'status': 'ok',
            'estudiantes': estudiantes,
            'sesion_existente': sesion is not None,
            'id_asistencia_diaria': sesion['id_asistencia_diaria'] if sesion else None,
            'registros': registros,
            'ids_asistencia_legacy': ids_asistencia_legacy
        }), 200
    except Exception as e:
        return _error_interno(e)


@asistencia_bp.route('/api/asistencia/guardar', methods=['POST'])
def api_asistencia_guardar():
    """Guardar/actualizar asistencia para una fecha.
    Crea la sesión (asistencias_diarias) si no existe, luego inserta/actualiza detalle_asistencia.
    Recalcula asistencias_por_periodo al finalizar."""
    try:
        data = request.get_json()
        grupo_id = data.get('grupo_id')
        materia_id = data.get('materia_id')
        periodo_id = data.get('periodo_id')
        fecha = data.get('fecha')
        registros = data.get('registros', [])

        if not grupo_id or not materia_id or not periodo_id or not fecha or not registros:
            return jsonify({'error': 'Datos incompletos'}), 400

        user_id = session.get('user_id')
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        bloqueo = _bloquear_docente_si_periodo_cerrado(cursor, periodo_id)
        if bloqueo:
            cursor.close()
            conn.close()
            return bloqueo

        # Obtener o crear plantilla para este docente/materia/periodo
        cursor.execute("""
            SELECT id_plantilla FROM plantillas_asistencias
            WHERE id_usuario = %s AND id_materia = %s AND id_periodo = %s
            LIMIT 1
        """, (user_id, materia_id, periodo_id))
        plantilla = cursor.fetchone()

        if not plantilla:
            # Obtener nombre de materia para la plantilla
            cursor.execute("SELECT nombre_materia FROM materias WHERE id_materia = %s", (materia_id,))
            mat = cursor.fetchone()
            nombre_mat = mat['nombre_materia'] if mat else 'Materia'

            cursor.execute("""
                INSERT INTO plantillas_asistencias (id_usuario, nombre_plantilla, descripcion, id_materia, id_periodo)
                VALUES (%s, %s, %s, %s, %s)
            """, (user_id, f'Asistencia {nombre_mat}', f'Control de asistencia - {nombre_mat}', materia_id, periodo_id))
            conn.commit()
            id_plantilla = cursor.lastrowid
        else:
            id_plantilla = plantilla['id_plantilla']

        # Buscar sesión existente
        cursor.execute("""
            SELECT id_asistencia_diaria FROM asistencias_diarias
            WHERE id_grupo = %s AND id_materia = %s AND fecha_registro = %s AND estado = 'activa'
            LIMIT 1
        """, (grupo_id, materia_id, fecha))
        sesion = cursor.fetchone()

        if not sesion:
            cursor.execute("""
                INSERT INTO asistencias_diarias (id_plantilla, id_materia, id_docente, id_grupo, fecha_registro)
                VALUES (%s, %s, %s, %s, %s)
            """, (id_plantilla, materia_id, user_id, grupo_id, fecha))
            conn.commit()
            id_sesion = cursor.lastrowid
        else:
            id_sesion = sesion['id_asistencia_diaria']

        # Insertar/actualizar detalle por cada estudiante
        guardados = 0
        ids_asistencia_legacy = {}
        for reg in registros:
            id_est = reg.get('id_estudiante')
            asistio = reg.get('asistio', 'no_registrado')
            comentario = reg.get('comentario', '').strip()

            # Validar enum
            if asistio not in ('presente', 'ausente', 'tardio', 'no_registrado'):
                continue

            cursor.execute("""
                INSERT INTO detalle_asistencia (id_asistencia_diaria, id_estudiante, asistio, comentario)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE asistio = VALUES(asistio), comentario = VALUES(comentario)
            """, (id_sesion, id_est, asistio, comentario))

            id_asistencia_legacy = _upsert_asistencia_legacy(
                cursor,
                id_estudiante=id_est,
                id_grupo=grupo_id,
                id_periodo=periodo_id,
                fecha=fecha,
                asistio=asistio,
                comentario=comentario
            )
            if id_asistencia_legacy:
                ids_asistencia_legacy[id_est] = id_asistencia_legacy

            guardados += 1

        conn.commit()

        # Recalcular consolidado en asistencias_por_periodo para cada estudiante
        ids_estudiantes = [r['id_estudiante'] for r in registros]
        _recalcular_asistencias_periodo(cursor, conn, ids_estudiantes, materia_id, periodo_id, grupo_id)

        # Generar/actualizar reportes de inasistencia si corresponde
        _actualizar_reportes_inasistencia(cursor, conn, ids_estudiantes, materia_id, periodo_id)

        cursor.close()
        conn.close()
        return jsonify({
            'status': 'ok',
            'guardados': guardados,
            'id_asistencia_diaria': id_sesion,
            'ids_asistencia_legacy': ids_asistencia_legacy
        }), 200
    except Exception as e:
        return _error_interno(e)


def _recalcular_asistencias_periodo(cursor, conn, ids_estudiantes, materia_id, periodo_id, grupo_id):
    """Recalcula asistencias_por_periodo para una lista de estudiantes."""
    for id_est in ids_estudiantes:
        cursor.execute("""
            SELECT
                SUM(da.asistio = 'presente') as presencias,
                SUM(
                    da.asistio = 'ausente' AND NOT EXISTS (
                        SELECT 1
                        FROM asistencia a
                        JOIN justificantes_ausencia ja ON ja.id_asistencia = a.id_asistencia
                        WHERE a.id_estudiante = da.id_estudiante
                            AND a.id_grupo = ad.id_grupo
                            AND a.id_periodo = %s
                            AND a.fecha_asistencia = ad.fecha_registro
                            AND ja.aprobado = 1
                    )
                ) as ausencias,
                SUM(da.asistio = 'tardio') as tardios,
                SUM(da.asistio = 'no_registrado') as no_reg,
                COUNT(*) as total
            FROM detalle_asistencia da
            JOIN asistencias_diarias ad ON da.id_asistencia_diaria = ad.id_asistencia_diaria
            WHERE da.id_estudiante = %s AND ad.id_materia = %s AND ad.estado = 'activa'
                AND ad.id_grupo = %s
        """, (periodo_id, id_est, materia_id, grupo_id))
        stats = cursor.fetchone()

        presencias = int(stats['presencias'] or 0)
        ausencias = int(stats['ausencias'] or 0)
        tardios = int(stats['tardios'] or 0)
        no_reg = int(stats['no_reg'] or 0)
        total = int(stats['total'] or 0)

        pct = (presencias / total * 100) if total > 0 else 0

        if pct >= 95:
            estado = 'excelente'
        elif pct >= 85:
            estado = 'bueno'
        elif pct >= 75:
            estado = 'regular'
        elif pct >= 60:
            estado = 'deficiente'
        else:
            estado = 'critico'

        cursor.execute("""
            INSERT INTO asistencias_por_periodo
                (id_estudiante, id_materia, id_periodo, total_presencias, total_ausencias,
                 total_tardios, total_no_registrados, total_clases_programadas,
                 porcentaje_asistencia, estado_asistencia)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                total_presencias = VALUES(total_presencias),
                total_ausencias = VALUES(total_ausencias),
                total_tardios = VALUES(total_tardios),
                total_no_registrados = VALUES(total_no_registrados),
                total_clases_programadas = VALUES(total_clases_programadas),
                porcentaje_asistencia = VALUES(porcentaje_asistencia),
                estado_asistencia = VALUES(estado_asistencia)
        """, (id_est, materia_id, periodo_id, presencias, ausencias, tardios, no_reg, total, round(pct, 2), estado))

    conn.commit()


def _actualizar_reportes_inasistencia(cursor, conn, ids_estudiantes, materia_id, periodo_id):
    """Genera o actualiza reportes de inasistencia cuando un estudiante supera el 20% de ausencias."""
    for id_est in ids_estudiantes:
        cursor.execute("""
            SELECT total_presencias, total_ausencias, total_clases_programadas, porcentaje_asistencia
            FROM asistencias_por_periodo
            WHERE id_estudiante = %s AND id_materia = %s AND id_periodo = %s
        """, (id_est, materia_id, periodo_id))
        row = cursor.fetchone()
        if not row:
            continue

        total = int(row['total_clases_programadas'] or 0)
        ausencias = int(row['total_ausencias'] or 0)
        pct_inasist = (ausencias / total * 100) if total > 0 else 0
        es_critica = 1 if pct_inasist > 20 else 0

        if ausencias > 0:
            # reportes_inasistencias no tiene UNIQUE KEY → usar SELECT + INSERT/UPDATE
            cursor.execute("""
                SELECT id_reporte_inasistencia FROM reportes_inasistencias
                WHERE id_estudiante = %s AND id_materia = %s AND id_periodo = %s
                LIMIT 1
            """, (id_est, materia_id, periodo_id))
            reporte_existente = cursor.fetchone()

            if reporte_existente:
                cursor.execute("""
                    UPDATE reportes_inasistencias
                    SET total_inasistencias = %s,
                        inasistencias_sin_justificar = %s,
                        porcentaje_inasistencia = %s,
                        es_critica = %s
                    WHERE id_reporte_inasistencia = %s
                """, (ausencias, ausencias, round(pct_inasist, 2), es_critica, reporte_existente['id_reporte_inasistencia']))
            else:
                cursor.execute("""
                    INSERT INTO reportes_inasistencias
                        (id_estudiante, id_materia, id_periodo, total_inasistencias,
                         inasistencias_sin_justificar, porcentaje_inasistencia, es_critica)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (id_est, materia_id, periodo_id, ausencias, ausencias, round(pct_inasist, 2), es_critica))

    conn.commit()


@asistencia_bp.route('/api/asistencias/justificante', methods=['POST'])
@asistencia_bp.route('/api/asistencia/justificante', methods=['POST'])
def api_asistencia_crear_justificante():
    """Crea o actualiza justificante de ausencia y vincula su id al detalle de asistencia."""
    try:
        id_asistencia = request.form.get('id_asistencia')
        id_estudiante = request.form.get('id_estudiante')
        id_asistencia_diaria = request.form.get('id_asistencia_diaria')
        id_periodo = request.form.get('periodo_id')
        tipo_justificante = (request.form.get('tipo_justificante') or 'Otro').strip()
        descripcion = (request.form.get('descripcion') or '').strip()
        fecha_documento = request.form.get('fecha_documento')
        archivo = request.files.get('archivo')

        if not id_estudiante:
            return jsonify({'error': 'id_estudiante es obligatorio'}), 400

        if tipo_justificante not in _TIPOS_JUSTIFICANTE:
            return jsonify({'error': 'tipo_justificante inválido'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        asistencia_row = None
        if id_asistencia:
            cursor.execute("""
                SELECT id_asistencia, id_estudiante, id_grupo, id_periodo, fecha_asistencia
                FROM asistencia
                WHERE id_asistencia = %s
                LIMIT 1
            """, (id_asistencia,))
            asistencia_row = cursor.fetchone()

        if not asistencia_row and id_asistencia_diaria and id_periodo:
            cursor.execute("""
                SELECT ad.id_grupo, ad.fecha_registro, da.asistio, da.comentario
                FROM asistencias_diarias ad
                JOIN detalle_asistencia da ON da.id_asistencia_diaria = ad.id_asistencia_diaria
                WHERE ad.id_asistencia_diaria = %s AND da.id_estudiante = %s
                LIMIT 1
            """, (id_asistencia_diaria, id_estudiante))
            detalle = cursor.fetchone()

            if detalle:
                id_asistencia = _upsert_asistencia_legacy(
                    cursor,
                    id_estudiante=id_estudiante,
                    id_grupo=detalle['id_grupo'],
                    id_periodo=id_periodo,
                    fecha=detalle['fecha_registro'],
                    asistio=detalle['asistio'],
                    comentario=(detalle.get('comentario') or '')
                )
                conn.commit()
                cursor.execute("""
                    SELECT id_asistencia, id_estudiante, id_grupo, id_periodo, fecha_asistencia
                    FROM asistencia
                    WHERE id_asistencia = %s
                    LIMIT 1
                """, (id_asistencia,))
                asistencia_row = cursor.fetchone()

        if not asistencia_row:
            cursor.close()
            conn.close()
            return jsonify({'error': 'No se encontró la asistencia para asociar el justificante'}), 404

        archivo_relativo = None
        if archivo and archivo.filename:
            nombre_original = secure_filename(archivo.filename)
            if not nombre_original:
                cursor.close()
                conn.close()
                return jsonify({'error': 'Nombre de archivo inválido'}), 400

            if not _extension_permitida(nombre_original):
                cursor.close()
                conn.close()
                return jsonify({'error': 'Tipo de archivo no permitido'}), 400

            _asegurar_carpeta_justificantes()
            extension = os.path.splitext(nombre_original)[1].lower()
            nombre_guardado = f"just_{asistencia_row['id_asistencia']}_{uuid4().hex}{extension}"
            ruta_abs = os.path.join(_JUSTIFICANTES_UPLOAD_DIR, nombre_guardado)
            archivo.save(ruta_abs)
            archivo_relativo = _ruta_relativa_justificante(nombre_guardado)

        cursor.execute("""
            SELECT id_justificante, archivo_path
            FROM justificantes_ausencia
            WHERE id_asistencia = %s
            ORDER BY created_at DESC
            LIMIT 1
        """, (asistencia_row['id_asistencia'],))
        existente = cursor.fetchone()

        if existente:
            campos_update = [
                "tipo_justificante = %s",
                "descripcion = %s",
                "fecha_documento = %s"
            ]
            params = [tipo_justificante, descripcion, fecha_documento or None]

            if archivo_relativo:
                campos_update.append("archivo_path = %s")
                params.append(archivo_relativo)

            params.append(existente['id_justificante'])
            cursor.execute(f"""
                UPDATE justificantes_ausencia
                SET {', '.join(campos_update)}
                WHERE id_justificante = %s
            """, tuple(params))
            id_justificante = existente['id_justificante']
        else:
            cursor.execute("""
                INSERT INTO justificantes_ausencia
                    (id_asistencia, id_estudiante, tipo_justificante, archivo_path, descripcion, fecha_documento)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                asistencia_row['id_asistencia'],
                id_estudiante,
                tipo_justificante,
                archivo_relativo,
                descripcion,
                fecha_documento or None
            ))
            id_justificante = cursor.lastrowid

        if id_asistencia_diaria:
            cursor.execute("""
                UPDATE detalle_asistencia
                SET justificante_id = %s
                WHERE id_asistencia_diaria = %s AND id_estudiante = %s
            """, (id_justificante, id_asistencia_diaria, id_estudiante))

        conn.commit()

        # Recalcular consolidado y reportes del estudiante para reflejar una posible aprobación posterior
        cursor.execute("""
            SELECT ad.id_materia, ad.id_grupo
            FROM asistencias_diarias ad
            JOIN detalle_asistencia da ON da.id_asistencia_diaria = ad.id_asistencia_diaria
            WHERE da.id_estudiante = %s AND ad.fecha_registro = %s AND ad.id_grupo = %s
            LIMIT 1
        """, (asistencia_row['id_estudiante'], asistencia_row['fecha_asistencia'], asistencia_row['id_grupo']))
        sesion_ctx = cursor.fetchone()
        if sesion_ctx:
            _recalcular_asistencias_periodo(
                cursor,
                conn,
                [asistencia_row['id_estudiante']],
                sesion_ctx['id_materia'],
                asistencia_row['id_periodo'],
                sesion_ctx['id_grupo']
            )
            _actualizar_reportes_inasistencia(
                cursor,
                conn,
                [asistencia_row['id_estudiante']],
                sesion_ctx['id_materia'],
                asistencia_row['id_periodo']
            )

        cursor.close()
        conn.close()
        return jsonify({
            'status': 'ok',
            'id_justificante': id_justificante,
            'id_asistencia': asistencia_row['id_asistencia'],
            'archivo_path': archivo_relativo,
            'aprobado': False
        }), 200
    except Exception as e:
        return _error_interno(e)


@asistencia_bp.route('/api/asistencias/justificantes', methods=['GET'])
@asistencia_bp.route('/api/asistencia/justificantes', methods=['GET'])
def api_asistencia_listar_justificantes():
    """Lista justificantes por grupo/materia/periodo."""
    try:
        grupo_id = request.args.get('grupo_id')
        materia_id = request.args.get('materia_id')
        periodo_id = request.args.get('periodo_id')

        if not grupo_id or not materia_id or not periodo_id:
            return jsonify({'error': 'Faltan parámetros: grupo_id, materia_id, periodo_id'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                ja.id_justificante,
                ja.id_asistencia,
                ja.id_estudiante,
                e.nombre,
                e.apellido,
                a.fecha_asistencia,
                ja.fecha_documento,
                ja.tipo_justificante,
                ja.archivo_path,
                ja.descripcion,
                ja.aprobado,
                ja.aprobado_por,
                ja.fecha_aprobacion,
                u.nombre as aprobado_por_nombre
            FROM justificantes_ausencia ja
            JOIN asistencia a ON a.id_asistencia = ja.id_asistencia
            JOIN estudiantes e ON e.id_estudiante = ja.id_estudiante
            LEFT JOIN usuarios u ON u.id_usuario = ja.aprobado_por
            JOIN detalle_asistencia da ON da.justificante_id = ja.id_justificante
            JOIN asistencias_diarias ad ON ad.id_asistencia_diaria = da.id_asistencia_diaria
            WHERE ad.id_grupo = %s
                AND ad.id_materia = %s
                AND a.id_periodo = %s
            ORDER BY COALESCE(ja.fecha_documento, a.fecha_asistencia) DESC, e.apellido, e.nombre
        """, (grupo_id, materia_id, periodo_id))
        rows = cursor.fetchall()

        for row in rows:
            row['aprobado'] = bool(row.get('aprobado'))
            if row.get('fecha_documento'):
                row['fecha_documento'] = row['fecha_documento'].isoformat()
            if row.get('fecha_asistencia'):
                row['fecha_asistencia'] = row['fecha_asistencia'].isoformat()
            if row.get('fecha_aprobacion'):
                row['fecha_aprobacion'] = row['fecha_aprobacion'].isoformat()

        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'justificantes': rows}), 200
    except Exception as e:
        return _error_interno(e)


@asistencia_bp.route('/api/asistencias/justificantes/<int:id_justificante>/aprobar', methods=['PUT'])
@asistencia_bp.route('/api/asistencia/justificantes/<int:id_justificante>/aprobar', methods=['PUT'])
def api_asistencia_aprobar_justificante(id_justificante):
    """Aprueba o desaprueba justificante y recalcula consolidado de asistencia."""
    try:
        rol = str(session.get('user_role') or '').strip().lower()
        if rol not in _ROLES_APROBACION_JUSTIFICANTES:
            return jsonify({'error': 'No tienes permisos para aprobar justificantes'}), 403

        data = request.get_json() or {}
        aprobado = bool(data.get('aprobado', False))
        user_id = session.get('user_id')

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                ja.id_justificante,
                ja.id_asistencia,
                ja.id_estudiante,
                a.id_periodo,
                a.id_grupo,
                a.fecha_asistencia,
                ad.id_materia
            FROM justificantes_ausencia ja
            JOIN asistencia a ON a.id_asistencia = ja.id_asistencia
            LEFT JOIN detalle_asistencia da ON da.justificante_id = ja.id_justificante
            LEFT JOIN asistencias_diarias ad ON ad.id_asistencia_diaria = da.id_asistencia_diaria
            WHERE ja.id_justificante = %s
            LIMIT 1
        """, (id_justificante,))
        row = cursor.fetchone()
        if not row:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Justificante no encontrado'}), 404

        if aprobado:
            cursor.execute("""
                UPDATE justificantes_ausencia
                SET aprobado = 1,
                    aprobado_por = %s,
                    fecha_aprobacion = NOW()
                WHERE id_justificante = %s
            """, (user_id, id_justificante))
            cursor.execute("""
                UPDATE asistencia
                SET estado = 'Justificado'
                WHERE id_asistencia = %s
            """, (row['id_asistencia'],))
        else:
            cursor.execute("""
                UPDATE justificantes_ausencia
                SET aprobado = 0,
                    aprobado_por = NULL,
                    fecha_aprobacion = NULL
                WHERE id_justificante = %s
            """, (id_justificante,))
            cursor.execute("""
                UPDATE asistencia
                SET estado = 'Ausente'
                WHERE id_asistencia = %s
            """, (row['id_asistencia'],))

        conn.commit()

        if row.get('id_materia') and row.get('id_periodo') and row.get('id_grupo'):
            _recalcular_asistencias_periodo(
                cursor,
                conn,
                [row['id_estudiante']],
                row['id_materia'],
                row['id_periodo'],
                row['id_grupo']
            )
            _actualizar_reportes_inasistencia(
                cursor,
                conn,
                [row['id_estudiante']],
                row['id_materia'],
                row['id_periodo']
            )

        cursor.close()
        conn.close()
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        return _error_interno(e)


@asistencia_bp.route('/api/asistencias/justificantes/<int:id_justificante>/archivo', methods=['GET'])
@asistencia_bp.route('/api/asistencia/justificantes/<int:id_justificante>/archivo', methods=['GET'])
def api_asistencia_ver_archivo_justificante(id_justificante):
    """Descarga o visualiza el archivo soporte de un justificante."""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT archivo_path
            FROM justificantes_ausencia
            WHERE id_justificante = %s
            LIMIT 1
        """, (id_justificante,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if not row or not row.get('archivo_path'):
            return jsonify({'error': 'Soporte no encontrado'}), 404

        ruta_abs = _ruta_absoluta_desde_relativa(row['archivo_path'])
        if not ruta_abs or not os.path.isfile(ruta_abs):
            return jsonify({'error': 'Archivo no disponible en el servidor'}), 404

        return send_file(ruta_abs, as_attachment=False)
    except Exception as e:
        return _error_interno(e)


@asistencia_bp.route('/api/asistencia/resumen', methods=['GET'])
def api_asistencia_resumen():
    """Resumen consolidado de asistencia por estudiante para un grupo/materia/periodo."""
    try:
        grupo_id = request.args.get('grupo_id')
        materia_id = request.args.get('materia_id')
        periodo_id = request.args.get('periodo_id')

        if not grupo_id or not materia_id or not periodo_id:
            return jsonify({'error': 'Faltan parámetros'}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT e.id_estudiante, e.nombre, e.apellido,
                   ap.total_presencias, ap.total_ausencias, ap.total_tardios,
                   ap.total_clases_programadas, ap.porcentaje_asistencia, ap.estado_asistencia,
                   COALESCE(just.total_justificados, 0) AS total_justificados,
                   ri.es_critica, ri.estado_reporte, ri.porcentaje_inasistencia,
                   ri.acciones_tomadas, ri.id_reporte_inasistencia
            FROM estudiantes e
            LEFT JOIN asistencias_por_periodo ap ON e.id_estudiante = ap.id_estudiante
                AND ap.id_materia = %s AND ap.id_periodo = %s
            LEFT JOIN (
                SELECT
                    da.id_estudiante,
                    COUNT(*) AS total_justificados
                FROM detalle_asistencia da
                JOIN asistencias_diarias ad ON ad.id_asistencia_diaria = da.id_asistencia_diaria
                JOIN asistencia a ON a.id_estudiante = da.id_estudiante
                    AND a.id_grupo = ad.id_grupo
                    AND a.id_periodo = %s
                    AND a.fecha_asistencia = ad.fecha_registro
                JOIN justificantes_ausencia ja ON ja.id_asistencia = a.id_asistencia
                WHERE ad.id_materia = %s
                    AND ad.id_grupo = %s
                    AND ad.estado = 'activa'
                    AND da.asistio = 'ausente'
                    AND ja.aprobado = 1
                GROUP BY da.id_estudiante
            ) just ON just.id_estudiante = e.id_estudiante
            LEFT JOIN reportes_inasistencias ri ON e.id_estudiante = ri.id_estudiante
                AND ri.id_materia = %s AND ri.id_periodo = %s
            WHERE e.id_grupo = %s AND e.estado = 'Activo'
            ORDER BY e.apellido, e.nombre
        """, (materia_id, periodo_id, periodo_id, materia_id, grupo_id, materia_id, periodo_id, grupo_id))
        estudiantes = cursor.fetchall()

        for e in estudiantes:
            e['porcentaje_asistencia'] = float(e['porcentaje_asistencia'] or 0)
            e['porcentaje_inasistencia'] = float(e['porcentaje_inasistencia'] or 0) if e.get('porcentaje_inasistencia') else 0
            e['es_critica'] = bool(e.get('es_critica'))

        cursor.close()
        conn.close()
        return jsonify({'status': 'ok', 'estudiantes': estudiantes}), 200
    except Exception as e:
        return _error_interno(e)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTES EXCEL DE ASISTENCIA
# ═══════════════════════════════════════════════════════════════════════════════

@asistencia_bp.route('/api/asistencia/reportes/general', methods=['GET'])
@asistencia_bp.route('/api/asistencia/reportes/consolidado', methods=['GET'])
def api_reporte_general_asistencia():
    """Genera reporte de asistencia rellenando la plantilla institucional PlantillaAsistencias_ReporteGeneral.xlsx"""
    try:
        grado_id = request.args.get('grado_id')
        grupo_id = request.args.get('grupo_id')
        materia_id = request.args.get('materia_id')
        periodo_id = request.args.get('periodo_id')

        if not all([grado_id, grupo_id, materia_id, periodo_id]):
            return jsonify({'error': 'Faltan parámetros: grado_id, grupo_id, materia_id, periodo_id'}), 400

        # Resolver ruta de plantilla
        plantilla_path = _resolver_ruta_plantilla_asistencias()
        if not plantilla_path or not os.path.exists(plantilla_path):
            return jsonify({'error': 'Plantilla no encontrada en backend/backend/templates'}), 404

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        # Cargar metadatos: grado, grupo, materia, período
        cursor.execute("""
            SELECT
                g.numero_grado,
                gr.codigo_grupo,
                m.nombre_materia,
                p.numero_periodo,
                p.nombre_periodo,
                YEAR(p.fecha_inicio) AS anio_lectivo
            FROM grados g
            JOIN grupos gr ON gr.id_grado = g.id_grado
            JOIN materias m ON m.id_materia = %s
            JOIN periodos p ON p.id_periodo = %s
            WHERE g.id_grado = %s AND gr.id_grupo = %s
            LIMIT 1
        """, (materia_id, periodo_id, grado_id, grupo_id))
        info = cursor.fetchone()

        if not info:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Datos no encontrados: grado/grupo/materia/período'}), 404

        # Datos del docente
        cursor.execute(
            "SELECT CONCAT(COALESCE(apellido, ''), ' ', COALESCE(nombre, '')) AS nombre_completo FROM usuarios WHERE id_usuario = %s LIMIT 1",
            (session.get('user_id'),)
        )
        docente_row = cursor.fetchone() or {}
        docente = (docente_row.get('nombre_completo') or '').strip() or str(session.get('user_name') or '').strip()

        # Datos institucionales opcionales (si existe tabla configuracion)
        institucion = 'DOCSTRY'
        sede = ''
        jornada = ''
        try:
            cursor.execute("""
                SELECT
                    COALESCE(nombre_colegio, 'DOCSTRY') AS institucion,
                    COALESCE(sede, '') AS sede,
                    COALESCE(jornada, '') AS jornada
                FROM configuracion
                LIMIT 1
            """)
            cfg = cursor.fetchone() or {}
            institucion = (cfg.get('institucion') or 'DOCSTRY').strip() or 'DOCSTRY'
            sede = (cfg.get('sede') or '').strip()
            jornada = (cfg.get('jornada') or '').strip()
        except Exception:
            # Mantener valores por defecto cuando no exista la tabla en el esquema.
            pass

        # Cargar TODOS los estudiantes del grupo
        cursor.execute("""
            SELECT
                e.id_estudiante,
                e.nombre,
                e.apellido,
                COALESCE(ap.total_presencias, 0) AS total_presencias,
                COALESCE(ap.total_ausencias, 0) AS total_ausencias,
                COALESCE(ap.total_tardios, 0) AS total_tardios,
                COALESCE(ap.total_clases_programadas, 0) AS total_clases_programadas,
                COALESCE(ap.estado_asistencia, '') AS estado_asistencia,
                COALESCE(just.total_justificados, 0) AS total_justificados
            FROM estudiantes e
            LEFT JOIN asistencias_por_periodo ap ON e.id_estudiante = ap.id_estudiante
                AND ap.id_materia = %s AND ap.id_periodo = %s
            LEFT JOIN (
                SELECT da.id_estudiante, COUNT(*) AS total_justificados
                FROM detalle_asistencia da
                JOIN asistencias_diarias ad ON ad.id_asistencia_diaria = da.id_asistencia_diaria
                JOIN asistencia a ON a.id_estudiante = da.id_estudiante
                    AND a.id_grupo = ad.id_grupo
                    AND a.id_periodo = %s
                    AND a.fecha_asistencia = ad.fecha_registro
                JOIN justificantes_ausencia ja ON ja.id_asistencia = a.id_asistencia
                WHERE ad.id_grupo = %s
                    AND ad.id_materia = %s
                    AND ad.estado = 'activa'
                    AND da.asistio = 'ausente'
                    AND ja.aprobado = 1
                GROUP BY da.id_estudiante
            ) just ON just.id_estudiante = e.id_estudiante
            WHERE e.id_grupo = %s AND e.estado = 'Activo'
            ORDER BY e.apellido, e.nombre
        """, (materia_id, periodo_id, periodo_id, grupo_id, materia_id, grupo_id))
        estudiantes = cursor.fetchall() or []

        cursor.close()
        conn.close()

        # Cargar plantilla Excel
        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active

        # Fecha de impresión
        fecha_impresion = datetime.now().strftime('%Y-%m-%d')
        anio_valor = info.get('anio_lectivo') or datetime.now().year

        # MAPEO EXACTO DE CELDAS SEGÚN PLANTILLA ACTUAL
        # Encabezados - usando función segura para manejar merged cells
        _set_cell_value_safe(ws, 'A1', institucion)  # A1: Institución
        _set_cell_value_safe(ws, 'A2', sede)  # A2: Sede
        _set_cell_value_safe(ws, 'A3', jornada)  # A3: Jornada
        _set_cell_value_safe(ws, 'F1', f"PERIODO: {info.get('nombre_periodo') or info.get('numero_periodo') or ''} -- {anio_valor}")  # F1
        _set_cell_value_safe(ws, 'F2', f"GRADO: {info.get('numero_grado') or ''}")  # F2
        _set_cell_value_safe(ws, 'H2', f"GRUPO: {info.get('codigo_grupo') or ''}")  # H2
        _set_cell_value_safe(ws, 'J2', f"FECHA: {fecha_impresion}")  # J2
        materia_texto = f"{info.get('nombre_materia') or ''}".strip()
        # La celda de materia en la plantilla esta en O2:P3 (merged).
        _set_cell_value_safe(ws, 'O2', materia_texto)
        ws['O2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
        _set_cell_value_safe(ws, 'F3', f"DOCENTE: {docente}")  # F3

        # RELLENO DE ESTUDIANTES DESDE FILA 8
        # Se limita a las filas formateadas de la plantilla para no desplazar diseño.
        fila_inicio = 6
        fila_fin_plantilla = 45
        total_filas = max(0, fila_fin_plantilla - fila_inicio + 1)

        for idx, est in enumerate(estudiantes):
            if idx >= total_filas:
                break  # No escribir más allá de 50 filas

            fila = fila_inicio + idx
            presentes = int(est.get('total_presencias') or 0)
            ausentes = int(est.get('total_ausencias') or 0)
            tardios = int(est.get('total_tardios') or 0)
            justificados = int(est.get('total_justificados') or 0)
            total_clases = int(est.get('total_clases_programadas') or 0)
            nombre_completo = f"{est.get('apellido', '')} {est.get('nombre', '')}".strip()
            codigo_estudiante = (est.get('documento') or '').strip() or str(est.get('id_estudiante') or '')
            estado_asistencia = (est.get('estado_asistencia') or '').strip()

            # Mapeo de columnas - usando función segura para manejar merged cells:
            # A: Código del estudiante
            _set_cell_value_safe(ws, f"A{fila}", codigo_estudiante)
            # B: Nombre completo
            _set_cell_value_safe(ws, f"B{fila}", nombre_completo)
            # D: Presentes
            _set_cell_value_safe(ws, f"D{fila}", presentes)
            # E: Ausentes
            _set_cell_value_safe(ws, f"E{fila}", ausentes)
            # F: Tardíos
            _set_cell_value_safe(ws, f"F{fila}", tardios)
            # G: Justificados
            _set_cell_value_safe(ws, f"G{fila}", justificados)
            # J: Total Clases
            _set_cell_value_safe(ws, f"J{fila}", total_clases)
            # K: % Asistencia (fórmula)
            _set_cell_value_safe(ws, f"K{fila}", f"=IF(J{fila}>0,((D{fila}+G{fila})/J{fila})*100,0)")
            # L: Estado
            _set_cell_value_safe(ws, f"L{fila}", estado_asistencia)

        # Limpiar filas restantes (si hay menos estudiantes que 50)
        for idx in range(len(estudiantes), total_filas):
            fila = fila_inicio + idx
            for col_letter in ['A', 'B', 'D', 'E', 'F', 'G', 'J', 'K', 'L']:
                _set_cell_value_safe(ws, f"{col_letter}{fila}", None)

        # Guardar en BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nombre_archivo = f"Asistencia_G{info.get('numero_grado', '')}{info.get('codigo_grupo', '')}_P{info.get('numero_periodo', '')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )

    except Exception as e:
        return _error_interno(e)